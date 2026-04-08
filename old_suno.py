import time
import os
import logging
import openpyxl
import re
import json
from browser_controller import BrowserController

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class SunoGenerator:
    def __init__(self, project_file, output_dir="data/output", delay=10, startup_delay=5, browser=None,
                 audio_influence=25, vocal_gender="Default", lyrics_mode="Default", 
                 weirdness=50, style_influence=50, persona_link=""):
        self.project_file = project_file
        self.metadata_path = project_file # Backward compat
        
        self.output_dir = output_dir
        self.delay = delay
        self.startup_delay = startup_delay
        self.browser = browser if browser else BrowserController()
        self.tab = self.browser.get_page("suno")
        self.base_url = "https://suno.com/create"
        
        # Advanced Params
        self.audio_influence = audio_influence
        self.vocal_gender = vocal_gender
        self.lyrics_mode = lyrics_mode
        self.weirdness = weirdness
        self.style_influence = style_influence
        self.persona_link = str(persona_link).strip()
        
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        self.stop_requested = False

    def run(self, max_count=None, target_ids=None, progress_callback=None, force_update=False):
        try:
            if not self.browser.context:
                self.browser.start()
            
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Bekleniyor: {self.startup_delay}sn (Başlangıç Gecikmesi)...")
                time.sleep(self.startup_delay)
            
            # --- Voice Workflow (Before reaching /create) ---
            if self.persona_link:
                success = self._setup_persona_workflow(progress_callback)
                if not success:
                    logger.warning("Voice workflow failed, falling back to direct navigation.")
                    self.browser.goto(self.base_url, page=self.tab)
            else:
                self.browser.goto(self.base_url, page=self.tab)
            
            time.sleep(5) # Wait for page structure

            # Always Ensure v5 for reliability
            self._ensure_v5_active()

            if "login" in self.tab.url or self.browser.is_visible("text='Log In'", page=self.tab):
                logger.warning("--- LOGIN REQUIRED ---")
                if progress_callback: progress_callback("global", "Login Required! Please check Chrome.")
                
                # Check repeatedly for login
                for _ in range(60): # Wait up to 5 mins for user login
                    if "create" in self.tab.url and "login" not in self.tab.url:
                        break
                    time.sleep(5)
            
            if "create" not in self.tab.url:
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)

            if not os.path.exists(self.metadata_path):
                logger.error("Results file not found.")
                return 0

            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): cell.column - 1 for cell in ws[1] if cell.value}
            
            rows_data = []
            
            # Normalize target_ids for robust matching
            target_ids_set = set(str(t).strip().lower() for t in target_ids) if target_ids else None

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rid_orig = row[headers.get('id', 0)] if 'id' in headers else ""
                rid = str(rid_orig).strip().lower()
                
                # Check target_ids
                if target_ids_set and rid not in target_ids_set:
                    continue
                
                status = str(row[headers.get('status', 0)]).lower() if 'status' in headers else ""
                
                # Determine if song already has music (M1 or M2 presence)
                # In scan_materials, we check files. Here we can also check the status column.
                is_done = "done" in status or "generated" in status
                
                if not force_update and is_done:
                    logger.info(f"Song {rid_orig} already generated on Suno. Skipping.")
                    if progress_callback: progress_callback(str(rid_orig), "Skip: Already Generated ✅")
                    continue
                
                row_dict = {key: row[idx] for key, idx in headers.items() if idx < len(row)}
                row_dict['_row_idx'] = i 
                rows_data.append(row_dict)
            
            if not rows_data:
                logger.info("Nothing to generate on Suno.")
                return 0

            if max_count and max_count > 0:
                rows_data = rows_data[:max_count]

            processed_count = 0
            for i, row_dict in enumerate(rows_data): 
                rid = str(row_dict.get('id', ''))
                
                if i > 0:
                    if progress_callback: progress_callback(rid, f"Waiting {self.delay}s...")
                    time.sleep(self.delay)
                
                self.update_row_status(row_dict['_row_idx'], "Processing")
                if progress_callback: progress_callback(rid, "Generating Music... 🎵")
                
                try:
                    success = self.process_row(row_dict, progress_callback=progress_callback)
                    if success:
                        self.update_row_status(row_dict['_row_idx'], "Generated")
                        if progress_callback: progress_callback(rid, "Music Generated! 🎵")
                        processed_count += 1
                    else:
                        self.update_row_status(row_dict['_row_idx'], "Failed")
                        if progress_callback: progress_callback(rid, "Generation Failed ❌")
                except Exception as row_error:
                    logger.error(f"Row error: {row_error}")
                    self.update_row_status(row_dict['_row_idx'], "Failed")
                    if progress_callback: progress_callback(rid, "Error ❌")

            return processed_count

        except Exception as e:
            logger.error(f"Suno error: {e}")
            raise e
        finally:
            logger.info("Suno finished.")

    def run_batch(self, target_ids=None, progress_callback=None, force_update=False, op_mode="full"):
        """
        Orchestrates the BATCH workflow:
        1. Generate ALL requested songs (Phase 1)
        2. Wait for & Download ALL requested songs (Phase 2)
        """
        try:
            if not self.browser.context:
                self.browser.start()
            
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Bekleniyor: {self.startup_delay}sn (Başlangıç Gecikmesi)...")
                time.sleep(self.startup_delay)
            
            # --- Common Setup (Voice / Login / v5) ---
            # Reuse logic from run(), or extract to common _init_session()
            if self.persona_link:
                success = self._setup_persona_workflow(progress_callback)
                if not success:
                    self.browser.goto(self.base_url, page=self.tab)
            else:
                self.browser.goto(self.base_url, page=self.tab)
            
            time.sleep(5)
            self._ensure_v5_active()

            if "login" in self.tab.url or self.browser.is_visible("text='Log In'", page=self.tab):
                logger.warning("--- LOGIN REQUIRED ---")
                if progress_callback: progress_callback("global", "Login Required! Please check Chrome.")
                for _ in range(60): 
                    if "create" in self.tab.url and "login" not in self.tab.url: break
                    time.sleep(5)
            
            if "create" not in self.tab.url:
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)

            # --- Prepare Data ---
            if not os.path.exists(self.metadata_path): return 0
            
            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): cell.column - 1 for cell in ws[1] if cell.value}
            
            rows_data = []
            target_ids_set = set(str(t).strip().lower() for t in target_ids) if target_ids else None

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rid_orig = row[headers.get('id', 0)] if 'id' in headers else ""
                rid = str(rid_orig).strip().lower()
                
                if target_ids_set and rid not in target_ids_set: continue
                
                status = str(row[headers.get('status', 0)]).lower() if 'status' in headers else ""
                is_done = "done" in status or "generated" in status
                
                # Check for existing MP3 files to double-verify (for Download phase skip)
                # But here we mostly care about Generation Phase skip
                
                if not force_update and is_done:
                    continue
                
                row_dict = {key: row[idx] for key, idx in headers.items() if idx < len(row)}
                row_dict['_row_idx'] = i 
                rows_data.append(row_dict)
            
            if not rows_data:
                logger.info("Nothing to batch generate.")
                return 0

            generated_ids = []
            
            # --- PHASE 1: BATCH GENERATION ---
            if op_mode in ["full", "gen_only"]:
                logger.info("--- Starting Batch Generation Phase ---")
                if progress_callback: progress_callback("global", "Toplu Üretim Başlatılıyor... 🚀")
                
                for i, row_dict in enumerate(rows_data):
                    if self.stop_requested: break
                    
                    rid = str(row_dict.get('id', ''))
                    
                    # Ensure a fresh state for every song in batch mode by reloading
                    if i >= 0: # Reload for every song to be absolutely sure
                        if self.persona_link:
                            logger.info(f"Batch: Re-activating Voice for song {rid}...")
                            self._setup_persona_workflow(progress_callback)
                        else:
                            logger.info(f"Batch: Reloading /create for song {rid}...")
                            self.browser.goto(self.base_url, page=self.tab)
                        
                        time.sleep(3)
                        self._ensure_v5_active()
                    
                    if i > 0:
                        time.sleep(self.delay)
                    
                    self.update_row_status(row_dict['_row_idx'], "Processing")
                    if progress_callback: progress_callback(rid, "Sıraya Alınıyor... 🎵")
                    
                    success = self._generate_single_no_wait(row_dict, progress_callback)
                    if success:
                        generated_ids.append(rid)
                        self.update_row_status(row_dict['_row_idx'], "Generating...") 
                    else:
                        self.update_row_status(row_dict['_row_idx'], "Failed")
                
                logger.info(f"Batch Generation Phase Complete. {len(generated_ids)} songs queued.")
            
            # --- PHASE 1.5: Identify Resume Items (For DL Only or Full Resume) ---
            # If we skipped generation, or if we want to pick up items that are already "Generating..."
            # Scan rows_data for items with status "Generating..." or "Processing"
            if op_mode in ["full", "dl_only"]:
                for row_dict in rows_data:
                    rid = str(row_dict.get('id', ''))
                    status = str(row_dict.get('status', '')).lower()
                    
                    # If it's a target ID, we definitely want it
                    is_target = target_ids and (rid in target_ids)
                    
                    if rid not in generated_ids:
                        if is_target or "generating" in status or "processing" in status or "suno bekleniyor" in status:
                             logger.info(f"Adding {rid} to download queue (is_target: {is_target}, status: {status})")
                             generated_ids.append(rid)

            # --- PHASE 2: BATCH DOWNLOAD ---
            if op_mode in ["full", "dl_only"] and generated_ids:
                logger.info("--- Starting Batch Download Phase ---")
                if progress_callback: progress_callback("global", "Toplu İndirme Bekleniyor... ⬇️")
                
                # PRE-DOWNLOAD PHASE: SMART WAIT ROOM
                # We wait until ALL items are ready before we start downloading any
                pending_ids = list(generated_ids)
                start_wait = time.time()
                total_timeout = max(600, len(pending_ids) * 180) 
                
                logger.info(f"Targeted Batch: Waiting for {len(pending_ids)} songs to complete before download...")
                
                while pending_ids and (time.time() - start_wait < total_timeout):
                    if self.stop_requested: break
                    still_generating = []
                    
                    for rid in pending_ids:
                        r_data = next((r for r in rows_data if str(r.get('id', '')).strip().lower() == rid), None)
                        if not r_data: continue
                        
                        title = r_data.get("title", "Song")
                        suno_title = f"{rid}_{title}"
                        
                        # Just checking readiness, not downloading yet
                        if not self._check_if_ready(suno_title, rid, suffix="1") or \
                           not self._check_if_ready(suno_title, rid, suffix="2"):
                            still_generating.append(rid)
                        else:
                            if progress_callback: progress_callback(rid, "Hazır! Beklemede... ✅")
                    
                    pending_ids = still_generating
                    if not pending_ids: break
                    
                    if progress_callback: progress_callback("global", f"Üretim Bekleniyor: {len(pending_ids)} şarkı kaldı... ⏳")
                    time.sleep(15)
                    
                    # Refresh page occasionally
                    if (time.time() - start_wait) % 90 < 15:
                        self.tab.reload()
                        time.sleep(5)
                        self._ensure_v5_active()

                # ACTUAL DOWNLOAD PHASE
                # Now that everything (or mostly everything) is ready, download them
                download_queue = list(generated_ids)
                completed_ids = []
                
                for rid in download_queue:
                    if self.stop_requested: break
                    r_data = next((r for r in rows_data if str(r.get('id', '')).strip().lower() == rid), None)
                    if not r_data: continue
                    
                    title = r_data.get("title", "Song")
                    suno_title = f"{rid}_{title}"
                    
                    if progress_callback: progress_callback(rid, "İndiriliyor... ⬇️")
                    
                    s1 = self._download_specific(suno_title, rid, suffix="1")
                    s2 = self._download_specific(suno_title, rid, suffix="2")
                    
                    if s1 or s2:
                        self.update_row_status(r_data['_row_idx'], "Generated")
                        if progress_callback: progress_callback(rid, "İndirildi! ✅")
                        completed_ids.append(rid)
                    else:
                        if progress_callback: progress_callback(rid, "İndirme Hatası! ❌")

                return len(completed_ids)

            return len(completed_ids)

        except Exception as e:
            logger.error(f"Batch Run Error: {e}")
            raise e
        finally:
            logger.info("Batch Run Finished.")

    def _generate_single_no_wait(self, row, progress_callback=None):
        """
        Fills the form and clicks Create, then verifies 'Generating' state appeared.
        Does NOT wait for completion.
        """
        prompt = row.get("prompt", "")
        lyrics = row.get("lyrics", "")
        style = row.get("suno_style", "")
        if not str(style).strip():
            style = row.get("style", "")
        
        rid = str(row.get("id", "song"))
        title = row.get("title", "Song")
        suno_title = f"{rid}_{title}"
        
        content = lyrics if str(lyrics).strip() else prompt
        
        try:
            # Custom Mode
            if not self.tab.locator("text='Lyrics'").first.is_visible():
                custom_btn = self.tab.locator("button:has-text('Custom')").first
                if custom_btn.is_visible():
                    custom_btn.click()
                    time.sleep(3)
            
            initial_count = self.tab.locator("div.clip-row").count()

            # Fills
            textareas = [t for t in self.tab.locator("textarea").all() if t.is_visible()]
            if not textareas: return False

            # Title Logic
            input_title = None
            title_loc = self.tab.locator("input[placeholder*='title' i]")
            for i in range(title_loc.count()):
                if title_loc.nth(i).is_visible():
                    input_title = title_loc.nth(i); break
            
            if not input_title:
                title_loc_alt = self.tab.locator("input[aria-label*='title' i]")
                for i in range(title_loc_alt.count()):
                    if title_loc_alt.nth(i).is_visible():
                        input_title = title_loc_alt.nth(i); break

            def fill_field(el, val):
                if not el or val is None: return
                try:
                    el.scroll_into_view_if_needed()
                    self.tab.evaluate(r"""(el) => {
                        const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value')?.set ||
                                           Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value')?.set;
                        if (nativeSetter) nativeSetter.call(el, '');
                        else el.value = '';
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }""", el)
                    time.sleep(0.1)
                except Exception: pass
                try:
                    self.browser.humanizer.type_text(self.tab, el, val)
                except Exception:
                    el.fill(str(val))

            # Lyrics: find by placeholder containing 'lyric'
            lyrics_textarea = self.tab.locator("textarea[placeholder*='lyric' i]").first
            if lyrics_textarea.is_visible():
                fill_field(lyrics_textarea, content)
            else:
                fill_field(textareas[0], content)

            # Style: find the textarea near 'Exclude styles' input (sibling in Styles section)
            style_textarea = self.tab.evaluate(r"""() => {
                const excludeInput = document.querySelector('input[placeholder*="Exclude" i]');
                if (excludeInput) {
                    let container = excludeInput.parentElement;
                    for (let d = 0; d < 8 && container; d++) {
                        const tas = container.querySelectorAll('textarea');
                        for (const ta of tas) {
                            if (ta.offsetParent !== null && !ta.placeholder.toLowerCase().includes('lyric') && !ta.placeholder.toLowerCase().includes('enhance')) {
                                return true;
                            }
                        }
                        container = container.parentElement;
                    }
                }
                return false;
            }""")
            if style_textarea:
                self.tab.evaluate(r"""(styleText) => {
                    const excludeInput = document.querySelector('input[placeholder*="Exclude" i]');
                    if (!excludeInput) return;
                    let container = excludeInput.parentElement;
                    for (let d = 0; d < 8 && container; d++) {
                        const tas = container.querySelectorAll('textarea');
                        for (const ta of tas) {
                            if (ta.offsetParent !== null && !ta.placeholder.toLowerCase().includes('lyric') && !ta.placeholder.toLowerCase().includes('enhance')) {
                                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value')?.set;
                                if (nativeSetter) nativeSetter.call(ta, styleText);
                                else ta.value = styleText;
                                ta.dispatchEvent(new Event('input', { bubbles: true }));
                                ta.dispatchEvent(new Event('change', { bubbles: true }));
                                return;
                            }
                        }
                        container = container.parentElement;
                    }
                }""", str(style))
            else:
                if len(textareas) > 2: fill_field(textareas[2], style)
                elif len(textareas) > 1: fill_field(textareas[1], style)

            if input_title:
                try: fill_field(input_title, suno_title)
                except: pass

            # Adv Options
            self._setup_lyrics_mode()
            self._setup_advanced_options()

            # Click Create
            create_btn = self.tab.locator("button:has-text('Create')").filter(has_not_text="Custom").last
            if not create_btn.is_visible():
                create_btn = self.tab.get_by_role("button", name="Create").last

            if create_btn.is_enabled():
                create_btn.click()
                time.sleep(2) 

                # Wait for clip count increase OR "Generating"
                for i in range(15):
                    if self._detect_captcha():
                        logger.warning("CAPTCHA DETECTED DURING BATCH!")
                        if progress_callback: progress_callback(rid, "⚠️ CAPTCHA! Çözün... 🕒")
                        self._play_alert()
                        while self._detect_captcha() and not self.stop_requested:
                            time.sleep(3)
                    
                    current_count = self.tab.locator("div.clip-row").count()
                    if current_count > initial_count:
                        logger.info(f"Batch: Queued {rid}")
                        return True
                    
                    time.sleep(2)
                
                return False
            return False
        except Exception as e:
            logger.error(f"Generate Batch Error {rid}: {e}")
            return False

    def _check_if_ready(self, title, rid, suffix="1"):
        """Checks if a song with title/rid is ready (not generating)."""
        try:
            rows = self.tab.locator("div.clip-row")
            count = rows.count()
            
            occurrence = 0
            target_occur = 0 if suffix == "1" else 1

            for i in range(min(20, count)):
                row = rows.nth(i)
                row_text = row.inner_text().lower()
                
                if i < 3:
                    logger.debug(f"[_check_if_ready] Row {i} Text Preview: {repr(row_text[:60])}...")
                
                if str(title).lower() in row_text or (str(rid) + "_" in row_text):
                    logger.info(f"[_check_if_ready] Matched {rid} in row {i}. FULL TEXT: {repr(row_text)}")
                    if occurrence == target_occur:
                        # Hover to ensure buttons are visible
                        try: row.hover(timeout=500)
                        except: pass
                        
                        is_gen = row.locator("text='Generating'").is_visible(timeout=500)
                        
                        # Broader more button search - Use count() to be less strict than is_visible()
                        more_btn = row.locator("button[data-context-menu-trigger='true'], button.context-menu-button, [aria-label*='More' i]").first
                        is_more = more_btn.count() > 0
                        
                        logger.info(f"[_check_if_ready] Status: Generating={is_gen}, MoreBtnCount={more_btn.count()}")
                        if not is_gen and is_more:
                            return True
                        return False 
                    occurrence += 1
            return False
        except: return False

    def _download_specific(self, title, rid, suffix="1"):
        """Downloads a specific occurrence of a song title."""
        try:
            # Ensure safe start state
            try: self.tab.keyboard.press("Escape")
            except: pass
            time.sleep(0.5)

            rows = self.tab.locator("div.clip-row")
            count = rows.count()
            if count == 0:
                logger.warning("[_download_specific] No rows found!")
                return False

            occurrence = 0
            target_occur = 0 if suffix == "1" else 1
            
            target_row = None
            
            for i in range(min(30, count)): # Increased scan range
                row = rows.nth(i)
                row_text = row.inner_text().lower()
                
                if i < 3:
                    logger.debug(f"[_download_specific] Row {i} Text Preview: {repr(row_text[:60])}...")
                
                if str(title).lower() in row_text or (str(rid) + "_" in row_text):
                    logger.info(f"[_download_specific] Found Match: {rid}_{title} in row {i}")
                    if occurrence == target_occur:
                        target_row = row
                        break
                    occurrence += 1
            
            if not target_row: return False
            
            # --- EXACT LOGIC FROM _wait_and_download (Lines 998-1040) ---
            target_row.scroll_into_view_if_needed()
            target_row.click()
            time.sleep(1)

            # Match locator logic: first try aria-label 'More', then .context-menu-button
            target_more = target_row.locator("button[aria-label*='More' i]").first
            if not target_more.is_visible():
                target_more = target_row.locator("button.context-menu-button").last # Matches _wait_and_download line 1007

            if not target_more.is_visible():
                logger.warning(f"[_download_specific] 'More' button not visible for {rid}")
                return False

            target_more.scroll_into_view_if_needed()
            time.sleep(1)
            target_more.click()
            time.sleep(2)
            
            target_dl = self.tab.locator("button:has-text('Download')").first
            if not target_dl.is_visible():
                target_dl = self.tab.get_by_text("Download", exact=True).first
            
            if not target_dl.is_visible():
                logger.warning(f"[_download_specific] 'Download' button NOT visible globally for {rid}")
                return False

            target_dl.hover()
            time.sleep(1.5)
            
            target_audio = None
            for _ in range(5):
                loc = self.tab.locator("button[aria-label*='WAV' i]").first
                if loc.is_visible():
                    target_audio = loc
                    break
                time.sleep(1)
            
            if not target_audio or not target_audio.is_visible():
                target_audio = self.tab.locator("button[aria-label*='MP3' i]").first
            
            if not target_audio or not target_audio.is_visible():
                target_audio = self.tab.locator("button:has-text('Audio')").first

            if not target_audio or not target_audio.is_visible():
                logger.warning(f"[_download_specific] Format (WAV/MP3) not found for {rid}")
                return False

            # End of identical logic - Proceed to save
            logger.info(f"[_download_specific] Selecting Format for {rid}")
            ext = "wav" if "wav" in (target_audio.get_attribute("aria-label") or "").lower() or "wav" in target_audio.inner_text().lower() else "mp3"
            
            # Click format button (WAV/MP3) -> Opens Popup
            target_audio.click()
            time.sleep(2)

            # --- POPUP HANDLING ---
            # User reported a popup opens with a "Download File" button
            try:
                # Wait for modal/dialog
                popup = self.tab.locator("div[role='dialog'], div.chakra-modal__content").last
                if popup.is_visible(timeout=5000):
                    logger.info(f"[_download_specific] Popup detected for {rid}")
                    
                    # Find 'Download File' button inside popup
                    # Common patterns: "Download File", "Download", or icon
                    dl_confirm_btn = popup.locator("button").filter(has_text="Download").last
                    
                    # Be more specific if possible
                    if not dl_confirm_btn.is_visible():
                         dl_confirm_btn = popup.locator("button:has-text('Download File')").first
                    
                    if dl_confirm_btn.is_visible():
                        logger.info(f"[_download_specific] Clicking 'Download File' confirmation for {rid}")
                        with self.tab.expect_download(timeout=60000) as download_info:
                            dl_confirm_btn.click()
                        
                        download = download_info.value
                        clean_title = re.sub(r'[^\w\s-]', '', title).strip()
                        clean_title = re.sub(r'[-\s]+', '_', clean_title)
                        filename = f"{clean_title}_{suffix}.{ext}"
                        save_path = os.path.join(self.output_dir, filename)
                        download.save_as(save_path)
                        logger.info(f"[_download_specific] Saved {filename}")
                        
                        # Close popup if it didn't close automatically
                        try: self.tab.keyboard.press("Escape")
                        except: pass
                        
                        return True
                    else:
                        logger.warning(f"[_download_specific] Popup found but 'Download File' button NOT visible for {rid}")
                        # Fallback: maybe the first click triggered download directly?
            except Exception as e:
                logger.warning(f"[_download_specific] Popup handling error: {e}")

            # Fallback if no popup or direct download
            # Check if download event happened on first click? (Unlikely given user report)
            logger.error(f"[_download_specific] Download failed - Popup flow incomplete for {rid}")
            return False
            
            return False
        except Exception as e:
            logger.error(f"Download Specific Error {rid}: {e}")
            return False



    def _setup_persona_workflow(self, progress_callback=None):
        """Navigates to voice page, clicks 'Create with Voice', and handles v5 modal."""
        try:
            logger.info(f"Navigating to Voice Link: {self.persona_link}")
            if progress_callback: progress_callback("global", "Voice Profili Seçiliyor... 👤")
            self.browser.goto(self.persona_link, page=self.tab)
            time.sleep(5)

            def find_and_click():
                return self.tab.evaluate("""() => {
                    const btns = Array.from(document.querySelectorAll('button'));
                    const btn = btns.find(b => {
                        const txt = (b.innerText || "").toLowerCase();
                        return txt.includes('create') && txt.includes('voice');
                    });
                    
                    if (btn) {
                        const style = window.getComputedStyle(btn);
                        const isVisible = btn.offsetWidth > 0 && btn.offsetHeight > 0 && style.visibility !== 'hidden' && style.display !== 'none';
                        const isEnabled = !btn.disabled && btn.getAttribute('aria-disabled') !== 'true';
                        
                        if (isVisible && isEnabled) {
                            btn.click();
                            return "clicked";
                        }
                        return "found_but_inactive";
                    }
                    return "not_found";
                }""")

            # Click "Create with Voice" with retries
            clicked = False
            logger.info("Waiting for 'Create with Voice' button...")
            
            # First attempt loop (Wait up to ~15s)
            for attempt in range(5): 
                status = find_and_click()
                if status == "clicked":
                    clicked = True
                    break
                elif status == "found_but_inactive":
                    logger.info(f"Button found but not ready yet (Attempt {attempt+1}/5)...")
                time.sleep(3)

            # If not found, try refresh (User request for soft 404/slow load)
            if not clicked:
                logger.warning("'Create with Voice' button not found or inactive. Refreshing page in 3s...")
                time.sleep(3)
                self.browser.goto(self.persona_link, page=self.tab)
                time.sleep(5) # Wait for reload
                
                # Second attempt loop (Wait up to ~30s)
                logger.info("Searching for button again after refresh...")
                for attempt in range(10):
                    status = find_and_click()
                    if status == "clicked":
                        clicked = True
                        break
                    time.sleep(3)

            if clicked:
                logger.info("Clicked 'Create with Voice'. Waiting for /create page...")
                time.sleep(5)
                # Handle "Store as voice" or "Switch to v5" modals if they appear
                self._handle_v5_switch_modal()
                return True
            else:
                logger.warning("'Create with Voice' button not found after refresh and retries.")
                return False
        except Exception as e:
            logger.error(f"Voice workflow failed: {e}")
            return False

    def _handle_v5_switch_modal(self):
        """Clicks 'Switch to v5' if it appears in a modal."""
        try:
            # Modal buttons often have "Switch to v5" text
            switched = self.tab.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button'));
                const switchBtn = btns.find(b => {
                    const txt = b.innerText || "";
                    return txt.includes('Switch to v5');
                });
                if (switchBtn) {
                    switchBtn.click();
                    return true;
                }
                return false;
            }""")
            if switched:
                logger.info("Handled 'Switch to v5' modal.")
                time.sleep(2)
        except: pass

    def _ensure_v5_active(self):
        """Checks if v5 is active, if not, tries to switch."""
        try:
            # Look for version selector
            v_selector = self.tab.locator("button:has-text('v4'), button:has-text('v3')").first
            if v_selector.is_visible():
                logger.info("Non-v5 model detected. Switching to v5...")
                v_selector.click()
                time.sleep(1)
                v5_option = self.tab.locator("div[role='menuitem']:has-text('v5'), button:has-text('v5')").first
                if v5_option.is_visible():
                    v5_option.click()
                    time.sleep(2)
        except: pass


    def process_row(self, row, progress_callback=None):
        prompt = row.get("prompt", "")
        lyrics = row.get("lyrics", "")
        style = row.get("suno_style", "")
        if not str(style).strip():
            style = row.get("style", "")
        
        # [Requirement 7] Title Format: ID_Title
        rid = str(row.get("id", "song"))
        title = row.get("title", "Song")
        suno_title = f"{rid}_{title}"
        
        content = lyrics if str(lyrics).strip() else prompt
        
        try:
            # --- Ensure Custom Mode ---
            if not self.tab.locator("text='Lyrics'").first.is_visible():
                logger.info("Custom mode not detected, clicking 'Custom'...")
                custom_btn = self.tab.locator("button:has-text('Custom')").first
                if custom_btn.is_visible():
                    custom_btn.click()
                    time.sleep(3)
            
            # Record current clip count to prevent stale downloads
            initial_count = self.tab.locator("div.clip-row").count()
            logger.info(f"Current clip count before creation: {initial_count}")

            # Ensure we are indeed in custom mode
            textareas = [t for t in self.tab.locator("textarea").all() if t.is_visible()]
            if not textareas: 
                logger.error("No textareas found on Suno.")
                return False

            # Fills
            input_title = None
            title_loc = self.tab.locator("input[placeholder*='title' i]")
            for i in range(title_loc.count()):
                if title_loc.nth(i).is_visible():
                    input_title = title_loc.nth(i)
                    break
            
            if not input_title:
                title_loc_alt = self.tab.locator("input[aria-label*='title' i]")
                for i in range(title_loc_alt.count()):
                    if title_loc_alt.nth(i).is_visible():
                        input_title = title_loc_alt.nth(i)
                        break

            def fill_field(el, val):
                if not el or val is None: return
                try:
                    # Robust JS-based clear to ensure reactive state is reset
                    try:
                        el.scroll_into_view_if_needed()
                        self.tab.evaluate(r"""(el) => {
                            const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value')?.set || 
                                               Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value')?.set;
                            if (nativeSetter) {
                                nativeSetter.call(el, '');
                            } else {
                                el.value = '';
                            }
                            el.dispatchEvent(new Event('input', { bubbles: true }));
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                        }""", el)
                        time.sleep(0.2)
                    except: pass

                    logger.info(f"Filling field with: {str(val)[:20]}...")
                    self.browser.humanizer.type_text(self.tab, el, val)
                except Exception as fe:
                    logger.warning(f"Humanizer failed: {fe}")
                    try:
                        el.scroll_into_view_if_needed()
                        el.fill(str(val))
                    except: pass

            # Lyrics: find by placeholder containing 'lyric'
            lyrics_textarea = self.tab.locator("textarea[placeholder*='lyric' i]").first
            if lyrics_textarea.is_visible():
                fill_field(lyrics_textarea, content)
            else:
                fill_field(textareas[0], content)

            # Style: find the textarea near 'Exclude styles' input (sibling in Styles section)
            style_textarea = self.tab.evaluate(r"""() => {
                const excludeInput = document.querySelector('input[placeholder*="Exclude" i]');
                if (excludeInput) {
                    let container = excludeInput.parentElement;
                    for (let d = 0; d < 8 && container; d++) {
                        const tas = container.querySelectorAll('textarea');
                        for (const ta of tas) {
                            if (ta.offsetParent !== null && !ta.placeholder.toLowerCase().includes('lyric') && !ta.placeholder.toLowerCase().includes('enhance')) {
                                return true;
                            }
                        }
                        container = container.parentElement;
                    }
                }
                return false;
            }""")
            if style_textarea:
                self.tab.evaluate(r"""(styleText) => {
                    const excludeInput = document.querySelector('input[placeholder*="Exclude" i]');
                    if (!excludeInput) return;
                    let container = excludeInput.parentElement;
                    for (let d = 0; d < 8 && container; d++) {
                        const tas = container.querySelectorAll('textarea');
                        for (const ta of tas) {
                            if (ta.offsetParent !== null && !ta.placeholder.toLowerCase().includes('lyric') && !ta.placeholder.toLowerCase().includes('enhance')) {
                                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value')?.set;
                                if (nativeSetter) nativeSetter.call(ta, styleText);
                                else ta.value = styleText;
                                ta.dispatchEvent(new Event('input', { bubbles: true }));
                                ta.dispatchEvent(new Event('change', { bubbles: true }));
                                return;
                            }
                        }
                        container = container.parentElement;
                    }
                }""", str(style))
            else:
                if len(textareas) > 2: fill_field(textareas[2], style)
                elif len(textareas) > 1: fill_field(textareas[1], style)

            # Title fill
            if input_title:
                try:
                    fill_field(input_title, suno_title)
                except Exception:
                    pass
                actual = input_title.input_value() if input_title.is_visible() else ""
                if str(suno_title) not in actual:
                    self.tab.evaluate("""(t) => {
                        const inputs = document.querySelectorAll('input[placeholder*="Song Title" i]');
                        for (const inp of inputs) {
                            if (inp.offsetParent !== null) {
                                inp.scrollIntoView({ block: 'center' });
                                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                                nativeSetter.call(inp, t);
                                inp.dispatchEvent(new Event('input', { bubbles: true }));
                                inp.dispatchEvent(new Event('change', { bubbles: true }));
                                return;
                            }
                        }
                    }""", str(suno_title))
                    time.sleep(1)

            # --- Voice & Advanced Options Setup ---
            self._setup_lyrics_mode()
            self._setup_advanced_options()

            # Click Create
            create_btn = self.tab.locator("button:has-text('Create')").filter(has_not_text="Custom").last
            if not create_btn.is_visible():
                create_btn = self.tab.get_by_role("button", name="Create").last

            if create_btn.is_enabled():
                logger.info(f"Clicking Create for: {suno_title}")
                create_btn.click()
                time.sleep(2) 

                # [Requirement 7 + CAPTCHA Alert]
                new_clip_appeared = False
                for i in range(25): # ~40-50s maximum wait
                    # 1. Check for success indicators
                    current_count = self.tab.locator("div.clip-row").count()
                    if current_count > initial_count:
                        new_clip_appeared = True
                        logger.info(f"New clip detected! Count: {current_count}")
                        break
                    # Check for "Generating" text in first row
                    if self.tab.locator("div.clip-row").first.locator("text='Generating'").is_visible(timeout=500):
                        new_clip_appeared = True
                        break
                    
                    # 2. Check for CAPTCHA / Challenge
                    if self._detect_captcha():
                        logger.warning("⚠️ CAPTCHA/Challenge detected on Suno!")
                        if progress_callback: progress_callback(rid, "⚠️ CAPTCHA ALGILANDI! Lütfen tarayıcıda çözün... 🕒")
                        self._play_alert()
                        # Pause until cleared
                        while self._detect_captcha() and not self.stop_requested:
                            time.sleep(3)
                        logger.info("Captcha cleared or disappeared. Resuming wait...")

                    time.sleep(2)
                
                if not new_clip_appeared:
                    if progress_callback: progress_callback(rid, "Suno Üretimi Başarısız (CAPTCHA olabilir) ❌")
                    return False

                # Trigger dual download
                logger.info(f"Triggering dual download for {rid}...")
                s1 = self._wait_and_download(suno_title, rid, progress_callback, index=0, suffix="1")
                s2 = self._wait_and_download(suno_title, rid, progress_callback, index=1, suffix="2")
                
                return s1 or s2
            
            logger.error("Create button not enabled or not found.")
            return False
        except Exception as e:
            logger.error(f"process_row failed: {e}")
            return False

    def _wait_and_download(self, title, rid, progress_callback=None, index=0, suffix="1"):
        try:
            logger.info(f"Waiting for {title} (ID: {rid}, Suffix: {suffix}) to finish generating...")
            if progress_callback: progress_callback(rid, f"Suno Bekleniyor ({suffix})... ⏳")
            
            found_ready = False
            best_index = index # Fallback
            
            for attempt in range(120): # 120 * 3s = 360s (6 minutes for longer generations)
                try:
                    # Check top 10 rows for matches
                    rows = self.tab.locator("div.clip-row")
                    count = rows.count()
                    
                    match_found = False
                    occurrence_count = 0
                    
                    for i in range(min(10, count)):
                        row = rows.nth(i)
                        row_text = row.inner_text().lower()
                        
                        # Check if this row is OUR song (rid or title match)
                        if str(rid).lower() in row_text or str(title).lower() in row_text:
                            if occurrence_count == index:
                                # This is the occurrence we want
                                is_gen = row.locator("text='Generating'").is_visible(timeout=300)
                                if not is_gen:
                                    # Double check readiness by looking for More button or duration
                                    more_btn = row.locator("button[aria-label*='More' i], button.context-menu-button").first
                                    duration_loc = row.locator("text=/\\d{1,2}:\\d{2}/")
                                    
                                    if more_btn.is_visible(timeout=300) or (duration_loc.count() > 0 and duration_loc.first.is_visible(timeout=300)):
                                        logger.info(f"Occurrence {index} for {rid} is READY at row {i}.")
                                        best_index = i
                                        found_ready = True
                                        match_found = True
                                        break
                                
                                # If we found our occurrence but it's still generating, log and break search
                                if attempt % 5 == 0:
                                    logger.info(f"Found occurrence {index} for {rid} but it's still generating...")
                                break
                            occurrence_count += 1
                    
                    if match_found: break
                except Exception as e:
                    logger.debug(f"Search attempt error: {e}")
                
                time.sleep(3) # Faster polling
            
            if not found_ready:
                logger.warning(f"Timeout waiting for {title} (suffix {suffix})")
                return False

            if progress_callback: progress_callback(rid, f"Ses İndiriliyor ({suffix})... ⬇️")
            self.tab.keyboard.press("Escape")
            time.sleep(1)

            try:
                rows = self.tab.locator("div.clip-row")
                if rows.count() <= best_index: return True
                target_row = rows.nth(best_index)
                target_row.click()
                time.sleep(1)

                target_more = target_row.locator("button[aria-label*='More' i]").first
                if not target_more.is_visible():
                    target_more = target_row.locator("button.context-menu-button").last

                if not target_more.is_visible(): return False

                target_more.scroll_into_view_if_needed()
                time.sleep(1)
                target_more.click()
                time.sleep(2)
                
                target_dl = self.tab.locator("button:has-text('Download')").first
                if not target_dl.is_visible():
                    target_dl = self.tab.get_by_text("Download", exact=True).first
                
                if not target_dl.is_visible(): return False

                target_dl.hover()
                time.sleep(1.5)
                
                target_audio = None
                for _ in range(5):
                    loc = self.tab.locator("button[aria-label*='WAV' i]").first
                    if loc.is_visible():
                        target_audio = loc
                        break
                    time.sleep(1)
                
                if not target_audio or not target_audio.is_visible():
                    target_audio = self.tab.locator("button[aria-label*='MP3' i]").first
                
                if not target_audio or not target_audio.is_visible():
                    target_audio = self.tab.locator("button:has-text('Audio')").first

                if not target_audio.is_visible(): return False

                ext = "wav" if "wav" in (target_audio.get_attribute("aria-label") or "").lower() else "mp3"
                target_audio.click()
                time.sleep(2)
                
                final_btn = self.tab.locator("button:has-text('Download File')").first
                target_click = final_btn if final_btn.is_visible() else None

                with self.tab.expect_download(timeout=120000) as download_info:
                    if target_click: target_click.click()
                
                download = download_info.value
                clean_title = re.sub(r'[^\w\s-]', '', title).strip()
                clean_title = re.sub(r'[-\s]+', '_', clean_title)
                # 'title' already contains the ID
                filename = f"{clean_title}_{suffix}.{ext}"
                save_path = os.path.join(self.output_dir, filename)
                download.save_as(save_path)
                return True
                
            except Exception as e:
                logger.error(f"Error during Suno download: {e}")
                return False
                
        except Exception as e:
            logger.error(f"Critical error in _wait_and_download: {e}")
            return False

    def close(self): pass

    def _setup_lyrics_mode(self):
        if self.lyrics_mode == "Default": return
        try:
            self.tab.evaluate(f"""(mode) => {{
                const btns = Array.from(document.querySelectorAll('button'));
                const target = btns.find(b => b.innerText.trim() === mode);
                if (target) target.click();
            }}""", self.lyrics_mode)
            time.sleep(1)
        except: pass

    def _setup_advanced_options(self):
        try:
            logger.info("Expanding More Options panel...")

            for attempt in range(3):
                # Check if already expanded
                exclude_visible = self.tab.evaluate("""() => {
                    const inp = document.querySelector('input[placeholder*="Exclude" i]');
                    return inp && inp.offsetParent !== null;
                }""")

                if exclude_visible:
                    logger.info("More Options already expanded.")
                    break

                # Click the More Options button — supports both old and new Suno UI names
                clicked = self.tab.evaluate("""() => {
                    const candidates = Array.from(document.querySelectorAll('div[role="button"], button'));
                    const btn = candidates.find(el => {
                        const txt = el.textContent.trim();
                        return txt === 'More Options' || txt === 'Advanced Options';
                    });
                    if (btn) {
                        btn.scrollIntoView({ block: 'center' });
                        btn.click();
                        return true;
                    }
                    return false;
                }""")

                if clicked:
                    logger.info(f"Clicked More Options (attempt {attempt+1})")
                    time.sleep(3)
                else:
                    logger.warning(f"More Options button not found (attempt {attempt+1})")
                    time.sleep(2)

            # Verify
            exclude_visible = self.tab.evaluate("""() => {
                const inp = document.querySelector('input[placeholder*="Exclude" i]');
                return inp && inp.offsetParent !== null;
            }""")
            if not exclude_visible:
                logger.warning("More Options panel did NOT expand.")
                return

            # Vocal Gender
            if self.vocal_gender != "Default":
                self.tab.evaluate(f"""(gender) => {{
                    const btns = Array.from(document.querySelectorAll('button'));
                    const target = btns.find(b => b.innerText.trim() === gender);
                    if (target) target.click();
                }}""", self.vocal_gender)
                time.sleep(1)

            # Slider helper: dispatch JS mouse events to open inline input, type value
            # Slider helper: dispatch JS mouse events to open inline input, type value
            def set_numeric_value(label_text, target_val):
                if target_val == "Default" or target_val is None: return
                try:
                    logger.info(f"Setting {label_text} to {target_val}%...")
                    
                    # 1. Wait for the label and percentage to be READY in DOM
                    ready = self.tab.evaluate(r"""(args) => {
                        return new Promise((resolve) => {
                            const check = () => {
                                const allEls = Array.from(document.querySelectorAll('div, span'));
                                const labelEl = allEls.find(el => 
                                    el.childNodes.length === 1 && 
                                    el.textContent.trim() === args.label &&
                                    el.offsetParent !== null
                                );
                                if (!labelEl) return false;
                                let row = labelEl.parentElement;
                                for (let depth = 0; depth < 10 && row; depth++) {
                                    const pcts = Array.from(row.querySelectorAll('div, span'))
                                        .filter(el => /^\d+%$/.test(el.textContent.trim()) && el.offsetParent !== null);
                                    if (pcts.length === 1) return true;
                                    row = row.parentElement;
                                }
                                return false;
                            };
                            if (check()) resolve(true);
                            else {
                                let count = 0;
                                const interval = setInterval(() => {
                                    if (check() || ++count > 20) {
                                        clearInterval(interval);
                                        resolve(check());
                                    }
                                }, 500);
                            }
                        });
                    }""", {"label": label_text})
                    
                    if not ready:
                        logger.warning(f"Slider/Label for {label_text} not found or not ready.")
                        return

                    # Verification loop: try up to 3 times to get the final value right
                    for main_retry in range(3):
                        # Clear focus
                        self.tab.keyboard.press("Escape")
                        time.sleep(0.3)
                        
                        # Scroll label into view
                        self.tab.evaluate(r"""(args) => {
                            const allEls = Array.from(document.querySelectorAll('div, span'));
                            const labelEl = allEls.find(el => 
                                el.childNodes.length === 1 && el.textContent.trim() === args.label && el.offsetParent !== null
                            );
                            if (labelEl) labelEl.scrollIntoView({ block: 'center' });
                        }""", {"label": label_text})
                        time.sleep(1)
                        
                        # Dispatch dblclick
                        input_ready = False
                        for retry in range(4):
                            self.tab.evaluate(r"""(args) => {
                                const allEls = Array.from(document.querySelectorAll('div, span'));
                                const labelEl = allEls.find(el => 
                                    el.childNodes.length === 1 && el.textContent.trim() === args.label && el.offsetParent !== null
                                );
                                if (!labelEl) return;
                                let row = labelEl.parentElement;
                                for (let depth = 0; depth < 10 && row; depth++) {
                                    const pcts = Array.from(row.querySelectorAll('div, span'))
                                        .filter(el => /^\d+%$/.test(el.textContent.trim()) && el.offsetParent !== null);
                                    if (pcts.length === 1) {
                                        const el = pcts[0];
                                        const rect = el.getBoundingClientRect();
                                        const cx = rect.x + rect.width/2;
                                        const cy = rect.y + rect.height/2;
                                        const opts = {bubbles: true, cancelable: true, view: window, clientX: cx, clientY: cy, detail: 1};
                                        el.dispatchEvent(new MouseEvent('mousedown', {...opts}));
                                        el.dispatchEvent(new MouseEvent('mouseup', {...opts}));
                                        el.dispatchEvent(new MouseEvent('click', {...opts}));
                                        const opts2 = {...opts, detail: 2};
                                        el.dispatchEvent(new MouseEvent('mousedown', opts2));
                                        el.dispatchEvent(new MouseEvent('mouseup', opts2));
                                        el.dispatchEvent(new MouseEvent('click', opts2));
                                        el.dispatchEvent(new MouseEvent('dblclick', opts2));
                                        return;
                                    }
                                    row = row.parentElement;
                                }
                            }""", {"label": label_text})
                            time.sleep(0.5)
                            
                            active_info = self.tab.evaluate("""() => {
                                const a = document.activeElement;
                                return a ? { tag: a.tagName } : null;
                            }""")
                            
                            if active_info and active_info['tag'] == "INPUT":
                                input_ready = True
                                break
                            
                            if active_info and active_info['tag'] == "DIV":
                                self.tab.keyboard.press("Escape")
                                time.sleep(0.3)
                        
                        if input_ready:
                            self.tab.keyboard.press("Meta+A")
                            time.sleep(0.1)
                            self.tab.keyboard.type(str(target_val))
                            self.tab.keyboard.press("Enter")
                            time.sleep(2) # Wait for UI to update text
                            
                            # Verify if it updated
                            current_val = self.tab.evaluate(r"""(args) => {
                                const allEls = Array.from(document.querySelectorAll('div, span'));
                                const lEl = allEls.find(el => el.childNodes.length === 1 && el.textContent.trim() === args.label && el.offsetParent !== null);
                                if (!lEl) return null;
                                let row = lEl.parentElement;
                                for (let d = 0; d < 10 && row; d++) {
                                    const p = Array.from(row.querySelectorAll('div, span'))
                                        .filter(el => /^\d+%$/.test(el.textContent.trim()) && el.offsetParent !== null);
                                    if (p.length === 1) return p[0].textContent.trim();
                                    row = row.parentElement;
                                }
                                return null;
                            }""", {"label": label_text})
                            
                            if current_val == f"{target_val}%":
                                logger.info(f"Successfully set {label_text} to {target_val}%")
                                return
                            else:
                                logger.warning(f"Value mismatch for {label_text}: got {current_val}, expected {target_val}%. Retrying (attempt {main_retry+1})...")
                        else:
                            logger.warning(f"Could not activate input for {label_text} on attempt {main_retry+1}")
                    
                    logger.error(f"Failed to set {label_text} correctly after all retries.")
                except Exception as e:
                    logger.warning(f"Error setting {label_text}: {e}")

            if self.persona_link:
                set_numeric_value("Audio Influence", self.audio_influence)
            set_numeric_value("Weirdness", self.weirdness)
            set_numeric_value("Style Influence", self.style_influence)

            # FINAL VERIFICATION: Read back all values
            final_check = self.tab.evaluate(r"""() => {
                const results = {};
                const labels = ["Audio Influence", "Weirdness", "Style Influence"];
                const allEls = Array.from(document.querySelectorAll('div, span'));
                
                labels.forEach(label => {
                    const labelEl = allEls.find(el => el.childNodes.length === 1 && el.textContent.trim() === label && el.offsetParent !== null);
                    if (labelEl) {
                        let row = labelEl.parentElement;
                        for (let depth = 0; depth < 10 && row; depth++) {
                            const pcts = Array.from(row.querySelectorAll('div, span'))
                                .filter(el => /^\d+%$/.test(el.textContent.trim()) && el.offsetParent !== null);
                            if (pcts.length === 1) {
                                results[label] = pcts[0].textContent.trim();
                                break;
                            }
                            row = row.parentElement;
                        }
                    }
                });
                
                // Gender check
                const btns = Array.from(document.querySelectorAll('button'));
                const genderBtns = btns.filter(b => ['Male', 'Female'].includes(b.textContent.trim()) && b.offsetParent !== null);
                genderBtns.forEach(b => {
                    const bg = window.getComputedStyle(b).backgroundColor;
                    if (bg !== 'rgba(0, 0, 0, 0)' && bg !== 'transparent') {
                        results['Vocal Gender'] = b.textContent.trim();
                    }
                });
                
                return results;
            }""")
            logger.info(f"Advanced Options verification: {json.dumps(final_check)}")
        except Exception as e:
            logger.warning(f"Advanced options error: {e}")

    def update_row_status(self, row_idx, status):
        try:
            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            status_col = None
            for cell in ws[1]:
                if str(cell.value).lower() == "status": status_col = cell.column
            if not status_col: 
                status_col = ws.max_column + 1
                ws.cell(row=1, column=status_col, value="status")
            ws.cell(row=row_idx, column=status_col, value=status)
            wb.save(self.metadata_path)
        except: pass

    def _detect_captcha(self):
        """Checks for known Cloudflare / hCaptcha elements."""
        captcha_selectors = [
            "iframe[title*='challenge']",
            "div#turnstile-wrapper",
            "iframe[src*='cloudflare-static']",
            "iframe[src*='hcaptcha']",
            "iframe[src*='recaptcha']",
            "#cf-turnstile-wrapper",
            ".cf-turnstile-wrapper",
            "div:has-text('Verify you are human')"
        ]
        for sel in captcha_selectors:
            try:
                if self.tab.locator(sel).is_visible(timeout=300):
                    return True
            except: pass
        return False

    def _play_alert(self):
        """Plays a prominent alert sound on macOS."""
        import sys
        try:
            if sys.platform == "darwin":
                # Play 'Glass' or 'Basso' for attention
                os.system("afplay /System/Library/Sounds/Glass.aiff &")
        except: pass

if __name__ == "__main__":
    suno = SunoGenerator(project_file="test.xlsx")
    suno.run()
