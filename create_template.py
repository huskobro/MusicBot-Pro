
import os
import openpyxl
from openpyxl.styles import PatternFill, Font

def create_template():
    # Define Workspace
    workspace_dir = os.path.expanduser("~/Documents/MusicBot_Workspace")
    if not os.path.exists(workspace_dir):
        os.makedirs(workspace_dir)
        print(f"Created Workspace: {workspace_dir}")
        
    # Define Subfolders
    media_dir = os.path.join(workspace_dir, "output_media")
    images_dir = os.path.join(workspace_dir, "images")
    for d in [media_dir, images_dir]:
        if not os.path.exists(d): 
            os.makedirs(d)
            
    # Create Template
    template_path = os.path.join(workspace_dir, "New_Project_Template.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Music Project"
    
    # Headers
    headers = [
        "id", "prompt", "style", "title", "lyrics", "status", 
        "visual_prompt", "video_prompt", "cover_art_prompt", "cover_art_path"
    ]
    
    # Styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        # Adjust width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
        
    # Example Row
    ws.append(["EXAMPLE_01", "A happy song about coding", "Pop", "Code Joy", "", "Pending", "", "", "", ""])
    
    wb.save(template_path)
    print(f"Created Template: {template_path}")

if __name__ == "__main__":
    create_template()
