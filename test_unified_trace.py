
import os
import openpyxl
import logging
from execution.gui_launcher import MusicBotGUI
import tkinter as tk
import traceback

# Configure logging
logging.basicConfig(level=logging.INFO)

def create_test_project():
    path = os.path.abspath("test_project.xlsx")
    if os.path.exists(path): os.remove(path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "prompt", "style"])
    ws.append(["TEST_01", "A retro funky boogie", "Disco"])
    wb.save(path)
    return path

def test_unified_workflow():
    print("--- Testing Unified Workflow ---")
    
    root = tk.Tk()
    app = MusicBotGUI(root)
    
    # 1. Create Raw Project
    proj_path = create_test_project()
    print(f"Created raw project: {proj_path}")
    
    # 2. Simulate Load
    print("Simulating Load Project...")
    app.load_project_data(proj_path)
    
    # 3. Verify Column Initialization
    print("Verifying Column Initialization...")
    wb = openpyxl.load_workbook(proj_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value]
    
    required = ["lyrics", "status", "cover_art_path"]
    missing = [col for col in required if col not in headers]
    
    if not missing:
        print("SUCCESS: All required columns auto-initialized! ✅")
    else:
        print(f"FAILURE: Missing columns: {missing} ❌")

    # 4. Verify GUI State
    if "test_project.xlsx" in app.lbl_project.cget("text"):
        print("SUCCESS: GUI label updated correctly! ✅")
    else:
        print("FAILURE: GUI label not updated. ❌")
        
    root.destroy()
    # os.remove(proj_path)

if __name__ == "__main__":
    try:
        test_unified_workflow()
    except Exception:
        traceback.print_exc()
