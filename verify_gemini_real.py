
import os
import time
import openpyxl
import logging
from execution.browser_controller import BrowserController
from execution.gemini_prompter import GeminiPrompter

# Configure logging to see our new debug messages
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def create_dummy_excel(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    # Headers
    headers = ["id", "cover_art_prompt", "cover_art_path"]
    ws.append(headers)
    # Row 1
    ws.append(["TEST_001", "A cyberpunk cat hacker, neon lights, high detail", ""])
    # Row 2
    ws.append(["TEST_002", "A peaceful zen garden with a robot gardener, anime style", ""])
    wb.save(filename)
    return filename

def run_real_test():
    xlsx_path = os.path.abspath("test_gemini_input.xlsx")
    create_dummy_excel(xlsx_path)
    
    print(f"Created dummy input: {xlsx_path}")
    
    # 1. Initialize Controller (Headless=False so user can see)
    # Note: This will attempt to use the local 'chrome_profile' folder.
    # If the user normally runs the .app, they might need to log in again here, 
    # OR we can point to the Documents path if we want to be fancy. 
    # For now, standard script path.
    browser = BrowserController(headless=False)
    
    try:
        # 2. Initialize Prompter
        prompter = GeminiPrompter(output_path=xlsx_path, browser=browser)
        
        # 3. Run Art Generation for 2 items
        print("Starting Batch Art Generation (2 Songs)...")
        # We pass target_ids to be specific
        prompter.generate_art_images(target_ids=["TEST_001", "TEST_002"], progress_callback=lambda rid, msg: print(f"[{rid}] {msg}"))
        
        # 4. Verify Output
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        
        # Check Row 2 (TEST_001)
        path1 = ws.cell(row=2, column=3).value
        # Check Row 3 (TEST_002)
        path2 = ws.cell(row=3, column=3).value
        
        print("\n--- Results ---")
        print(f"TEST_001 Path: {path1}")
        print(f"TEST_002 Path: {path2}")
        
        if path1 and os.path.exists(path1) and path2 and os.path.exists(path2):
            print("SUCCESS: Both images generated and saved! ✅")
            # Cleanup
            # os.remove(xlsx_path)
        else:
            print("FAILURE: One or more images missing. ❌")

    except Exception as e:
        print(f"Test Failed with Error: {e}")
    finally:
        print("Closing browser in 5 seconds...")
        time.sleep(5)
        browser.stop()

if __name__ == "__main__":
    run_real_test()
