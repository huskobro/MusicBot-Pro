"""
Excel management mixin for SunoGenerator.
Handles reading/writing status, atomic saves, backups, and recovery.
"""
import os
import time
import shutil
import logging
import openpyxl

from logging.handlers import RotatingFileHandler

workspace_dir = os.path.expanduser("~/Documents/MusicBot_Workspace")
log_dir = os.path.join(workspace_dir, "logs")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, "suno_musicbot.log")

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
if not logger.hasHandlers():
    file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5, encoding="utf-8")
    console_handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)


class SunoExcelMixin:
    """Excel file management: status updates, atomic writes, backup & recovery."""

    def update_row_status(self, row_idx, **kwargs):
        """Updates status in memory cache. Flushes to disk periodically."""
        if not hasattr(self, "_excel_cache"):
            self._excel_cache = {}
            self._last_flush = time.time()
            
        if row_idx not in self._excel_cache:
            self._excel_cache[row_idx] = {}
            
        for key, val in kwargs.items():
            if val is not None:
                self._excel_cache[row_idx][key] = val
            
        # Flush if 15 seconds have passed
        if time.time() - getattr(self, "_last_flush", 0) > 15:
            self.flush_excel_cache()

    def flush_excel_cache(self):
        """Writes all pending status updates from cache to the Excel file atomically."""
        if not hasattr(self, "_excel_cache") or not self._excel_cache:
            return
            
        # Multi-concurrent safety
        if hasattr(self, "xlsx_lock") and self.xlsx_lock:
            with self.xlsx_lock:
                self._flush_excel_cache_internal()
        else:
            self._flush_excel_cache_internal()

    def _flush_excel_cache_internal(self):
        try:
            self._backup_excel()

            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            headers = {str(cell.value).lower(): cell.column for cell in ws[1] if cell.value}

            for row_idx, updates in self._excel_cache.items():
                for key, val in updates.items():
                    col = headers.get(key)
                    if not col:
                        col = ws.max_column + 1
                        ws.cell(row=1, column=col, value=key)
                        headers[key] = col
                    try:
                        target_row_idx = int(row_idx)
                        ws.cell(row=target_row_idx, column=col, value=val)
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid row_idx for Excel update: {row_idx}")

            temp_path = self.metadata_path + ".tmp"
            wb.save(temp_path)
            shutil.move(temp_path, self.metadata_path)
            
            self._excel_cache.clear()
            self._last_flush = time.time()

        except Exception as e:
            logger.error(f"Failed to flush Excel cache: {e}")
            if "not a zip file" in str(e).lower() or "BadZipFile" in str(e):
                self._recover_excel_from_backup()

    def _backup_excel(self):
        """Keep rotating backups of the Excel file (last 3 versions). Throttled to every 30s."""
        try:
            if not os.path.exists(self.metadata_path):
                return

            now = time.time()
            last_backup = getattr(self, '_last_backup_time', 0)
            if now - last_backup < self.config.backup_interval:
                return
            self._last_backup_time = now

            backup_dir = os.path.join(os.path.dirname(self.metadata_path), ".backups")
            os.makedirs(backup_dir, exist_ok=True)

            base = os.path.basename(self.metadata_path)

            for i in range(self.config.max_backups, 1, -1):
                old = os.path.join(backup_dir, f"{base}.bak{i-1}")
                new = os.path.join(backup_dir, f"{base}.bak{i}")
                if os.path.exists(old):
                    shutil.copy2(old, new)

            shutil.copy2(self.metadata_path, os.path.join(backup_dir, f"{base}.bak1"))
            logger.debug("Excel backup created (.bak1)")

        except Exception as e:
            logger.debug(f"Backup failed (non-critical): {e}")

    def _recover_excel_from_backup(self):
        """Attempt to recover Excel file from the most recent valid backup."""
        try:
            backup_dir = os.path.join(os.path.dirname(self.metadata_path), ".backups")
            base = os.path.basename(self.metadata_path)

            for i in range(1, self.config.max_backups + 1):
                backup_path = os.path.join(backup_dir, f"{base}.bak{i}")
                if os.path.exists(backup_path):
                    temp_test = backup_path + ".temp.xlsx"
                    try:
                        shutil.copy2(backup_path, temp_test)
                        openpyxl.load_workbook(temp_test)  # Verify it's valid
                        os.remove(temp_test)
                        
                        shutil.copy2(backup_path, self.metadata_path)
                        logger.info(f"✅ Excel recovered from backup .bak{i}")
                        return True
                    except Exception as e:
                        if os.path.exists(temp_test): os.remove(temp_test)
                        logger.debug(f"Failed to load backup {backup_path}: {e}")
                        continue

            logger.error("❌ No valid Excel backup found for recovery!")
            return False
        except Exception as e:
            logger.error(f"Recovery failed: {e}")
            return False
