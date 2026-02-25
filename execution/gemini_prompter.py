import time
import os
import logging
import openpyxl
import traceback
from browser_controller import BrowserController
from openpyxl.styles import PatternFill

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class GeminiPrompter:
    def __init__(self, project_file, output_dir=None, headless=False, 
                 use_gemini_lyrics=True, generate_visual=True, generate_video=True, generate_style=False, startup_delay=5, 
                 language="Turkish", browser=None, chat_mode="New Chat"):
        self.project_file = project_file
        self.output_dir = output_dir if output_dir else os.path.dirname(project_file)
        # Backward compatibility for internal methods
        self.output_path = project_file
        self.metadata_path = project_file
        
        self.use_gemini_lyrics = use_gemini_lyrics
        self.generate_visual = generate_visual
        self.generate_video = generate_video
        self.generate_style = generate_style
        self.startup_delay = startup_delay
        self.language = language
        self.chat_mode = chat_mode
        self.browser = browser if browser else BrowserController(headless=headless)
        self.tab = self.browser.get_page("default")
        self.base_url = "https://gemini.google.com/app"
        
        # Load Prompts
        self.prompts_path = os.path.join(os.path.dirname(self.metadata_path), "prompts.json")
        self.load_prompts()
        self.load_translations()

    def load_translations(self):
        import json
        import sys
        t_file = "translations_tr.json" if self.language == "Turkish" else "translations_en.json"
        
        if getattr(sys, 'frozen', False):
            base_dir = os.path.join(sys._MEIPASS, "execution")
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            
        t_path = os.path.join(base_dir, t_file)
        self.translations = {}
        if os.path.exists(t_path):
            try:
                with open(t_path, "r", encoding="utf-8") as f:
                    self.translations = json.load(f)
            except Exception as e: 
                logger.error(f"Gemini translations load error: {e}")
        else:
            logger.error(f"Translation file not found at: {t_path}")

    def t(self, key, default="", **kwargs):
        val = self.translations.get(key, default or key)
        try: return val.format(**kwargs)
        except Exception: return val

    def load_prompts(self):
        import json
        default_lyrics = """Sen profesyonel bir şarkı sözü yazarı ve müzik prodüktörüsün. Suno.ai modelinin en iyi şekilde besteleyebilmesi için, sana vereceğim temalarda içerik oluşturmanı istiyorum.
Tema: {theme}
Language: {language} (IMPORTANT: Use only {language} for lyrics and title)

Lütfen şu formatta yanıt ver (başka bir şey yazma, markdown kullanma):
Başlık: [{language} dilinde şarkı başlığı]
Sözler:
[{language} dilinde, vurucu ve kafiyeli şarkı sözleri... (Intro, Verse 1, Chorus, Verse 2, Bridge, Outro etiketleriyle)]
Stil: [Müzik tarzı, enstrümanlar ve tempo (Örn: Lo-fi, Melancholic Piano, 90bpm)]
Görsel Prompt: [Albüm kapağı için İngilizce, detaylı Stable Diffusion promptu]
Video Prompt: [Müzik videosu için İngilizce, detaylı video üretim promptu]
"""
        default_art = """Create a high-quality YouTube music thumbnail inspired by modern romantic/lofi/ballad compilation channels...
Main title: “{title}”
""" 

        if os.path.exists(self.prompts_path):
            try:
                with open(self.prompts_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.master_prompt_template = data.get("lyrics_master_prompt", default_lyrics)
                    self.visual_master_prompt = data.get("visual_master_prompt", "")
                    self.video_master_prompt = data.get("video_master_prompt", "")
                    self.art_master_prompt = data.get("art_master_prompt", default_art)
            except Exception as e:
                logger.error(f"Error loading prompts.json: {e}")
                self.master_prompt_template = default_lyrics
                self.visual_master_prompt = ""
                self.video_master_prompt = ""
                self.art_master_prompt = default_art
        else:
             self.master_prompt_template = default_lyrics
             self.visual_master_prompt = ""
             self.video_master_prompt = ""
             self.art_master_prompt = default_art

    def _start_new_chat(self):
        self._lyrics_injected_in_current_page = False
        try:
            logger.info(f"Starting Gemini Chat Mode: {self.chat_mode}")
            
            if "Geçici" in self.chat_mode or "Temp" in self.chat_mode:
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)
            else:
                try:
                    new_chat_btn = self.tab.locator("a[href*='/app'], a[aria-label*='New chat' i], a[aria-label*='Yeni sohbet' i]").first
                    if new_chat_btn.is_visible(timeout=2000):
                        new_chat_btn.click()
                        time.sleep(2)
                    else:
                        self.browser.goto(self.base_url, page=self.tab)
                        time.sleep(3)
                except Exception:
                    self.browser.goto(self.base_url, page=self.tab)
                    time.sleep(3)
            time.sleep(1)
        except Exception as e:
            logger.warning(f"Failed to start new chat smoothly, doing hard reload: {e}")
            try:
                self.browser.ensure_alive("default")
                self.tab = self.browser.get_page("default")
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)
            except Exception as severe_e:
                logger.error(f"Severe navigation failure in chat start: {severe_e}")

    def run(self, max_count=None, target_ids=None, progress_callback=None, force_update=False):
        self.progress_callback_internal = progress_callback
        try:
            if not self.browser.context:
                self.browser.start()
            
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Waiting {self.startup_delay}s (Startup Delay)...")
                time.sleep(self.startup_delay)
            
            self.browser.goto(self.base_url, page=self.tab)
            time.sleep(3) 

            # Check if login is needed
            if "accounts.google.com" in self.tab.url:
                logger.warning("--- LOGIN REQUIRED ---")
                if progress_callback: progress_callback("global", "Gemini Login Required! Please check Chrome.")
                time.sleep(1)

            if not os.path.exists(self.metadata_path):
                logger.error("Input file not found.")
                return 0

            # Load pending rows logic
            import openpyxl
            import uuid
            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            # Clean header mapping (Robust case-insensitive and whitespace handled)
            headers = {str(cell.value).strip().lower(): cell.column - 1 for cell in ws[1] if cell.value}
            
            # Column mapping for status check (Multiple variations support)
            lyr_col_idx = headers.get('lyrics')
            vis_col_idx = headers.get('visual_prompt') or headers.get('visual prompt') or headers.get('visual_prompts') or headers.get('visual prompts')
            vid_col_idx = headers.get('video_prompt') or headers.get('video prompt') or headers.get('video_prompts') or headers.get('video prompts')
            status_col_idx = headers.get('status')
            id_col_idx = headers.get('id')
            if id_col_idx is None: # Fallback to first column if no 'id' header found
                id_col_idx = 0
            prompt_col_idx = headers.get('prompt') or headers.get('lyrics_prompt') or 1
            style_col_idx = headers.get('style') or 2

            pending_rows = []
            updates_needed = False
            
            # Normalize target_ids for comparison
            target_ids_set = set(str(t).strip().lower() for t in target_ids) if target_ids else None

            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                rid_orig = row[id_col_idx].value
                # 1. Ensure ID exists
                if not rid_orig:
                    rid_orig = str(uuid.uuid4())[:8]
                    row[id_col_idx].value = rid_orig
                    updates_needed = True
                
                rid = str(rid_orig).strip().lower()
                if target_ids_set and rid not in target_ids_set:
                    continue
                
                # 2. Check if song needs processing
                status_val = ""
                if status_col_idx is not None and status_col_idx < len(row):
                    status_val = str(row[status_col_idx].value or "").lower()
                
                # We process if:
                # - Status is not 'done'
                # - OR any requested step is missing data
                needs_work = False
                if "done" not in status_val and "completed" not in status_val:
                    needs_work = True
                
                # Even if status is 'done', check for missing components
                lyrics_val = row[lyr_col_idx].value if lyr_col_idx is not None else ""
                visual_val = row[vis_col_idx].value if vis_col_idx is not None else ""
                video_val = row[vid_col_idx].value if vid_col_idx is not None else ""
                
                # Ensure values aren't just whitespace
                has_lyrics = lyrics_val and str(lyrics_val).strip()
                has_visual = visual_val and str(visual_val).strip()
                has_video = video_val and str(video_val).strip()

                if self.use_gemini_lyrics and not has_lyrics: needs_work = True
                if self.generate_visual and not has_visual: needs_work = True
                if self.generate_video and not has_video: needs_work = True
                
                if needs_work or force_update:
                    pending_rows.append({
                        'id': str(rid_orig), 
                        'row_idx': i, 
                        'theme': row[prompt_col_idx].value, 
                        'style': row[style_col_idx].value if style_col_idx is not None else "",
                        'existing_lyrics': lyrics_val,
                        'existing_visual': visual_val,
                        'existing_video': video_val
                    })
                else:
                    logger.info(f"Song {rid_orig} marked as DONE/SKIP (Data already exists).")
                    if progress_callback: progress_callback(str(rid_orig), "Skipping: Data exists ✅")
            
            if updates_needed:
                wb.save(self.metadata_path)

            if not pending_rows:
                logger.info("Nothing to process in Gemini.")
                return 0

            if max_count: pending_rows = pending_rows[:max_count]

            processed_count = 0
            def process_pending(rows, pass_num=1):
                nonlocal processed_count
                failed_ids = []
                
                for i, rowdata in enumerate(rows):
                    row_id = rowdata['id']
                    theme = rowdata['theme']
                    style_init = rowdata['style']
                    
                    logger.info(f"[Pass {pass_num}] Processing Song {i+1}/{len(rows)} ID: {row_id}")
                    self._prompt_injected_for_row_id = row_id
                    
                    # Prevent crashes: Ensure browser is still alive before processing
                    try:
                        is_alive = self.browser.ensure_alive("default")
                        if not is_alive: raise Exception("ensure_alive returned False")
                        self.tab = self.browser.get_page("default")
                    except Exception as e:
                        logger.error(f"Browser recovery failed for song {row_id}, skipping: {e}")
                        failed_ids.append(rowdata)
                        continue

                    # Refresh Chat State based on Settings before starting a song
                    self._start_new_chat()
                    
                    # --- Step 1: Lyrics ---
                    final_title = ""
                    final_style = style_init
                    if (self.use_gemini_lyrics and not rowdata.get('existing_lyrics')) or force_update:
                        if progress_callback: progress_callback(row_id, f"Step 1: Lyrics (Deneme {pass_num})... ✍️")
                        lyrics_res = self.generate_content(theme, style=style_init)
                        
                        if lyrics_res:
                            output_payload = lyrics_res.copy()
                            if "style" in lyrics_res:
                                output_payload["suno_style"] = lyrics_res["style"]
                            
                            self.update_output_data(row_id, output_payload)
                            
                            final_title = lyrics_res.get("title", "")
                            final_style = lyrics_res.get("style", style_init)
                            
                            if lyrics_res.get("status") == "YARIM":
                                logger.warning(f"Song {row_id} lyrics missing Outro. Marked as YARIM.")
                                if progress_callback: progress_callback(row_id, self.t("log_gemini_yarim"))
                                failed_ids.append(rowdata) # Append to failed list so Pass 2 will auto-retry
                            else:
                                if progress_callback: progress_callback(row_id, self.t("log_gemini_lyrics_saved"))
                        else:
                            if progress_callback: progress_callback(row_id, self.t("log_gemini_timeout", timeout=120))
                            logger.error(f"Gemini failed to respond for {row_id} after timeout. Forcing Hard Reset and continuing.")
                            failed_ids.append(rowdata)
                            
                            # Force a strict hard reset so the broken context doesn't infect the next song
                            try:
                                self.browser.stop()
                                time.sleep(2)
                                self.browser.start()
                                self.tab = self.browser.get_page("default")
                                self._lyrics_injected_in_current_page = False
                            except Exception as e:
                                logger.error(f"Failed to hard reset browser: {e}")
                            
                            time.sleep(2)
                            continue # Move to next song safely instead of crashing batch
                    else:
                        logger.info(f"   -> Lyrics already exist for {row_id}, skipping to prompt design.")
                        final_title = theme # Fallback

                    # --- Step 2: Visual Prompt ---
                    if self.generate_visual and self.visual_master_prompt and (not rowdata.get('existing_visual') or force_update):
                        if progress_callback: progress_callback(row_id, "Step 2: Visual Design... 🎨")
                        vis_res = self.generate_focused_prompt("visual", final_title, final_style)
                        if vis_res:
                            vis_res = vis_res.replace("**", "").replace("__", "").strip()
                            self.update_output_data(row_id, {"visual_prompt": vis_res})
                            if progress_callback: progress_callback(row_id, self.t("log_gemini_visual_saved"))
                        else:
                            if progress_callback: progress_callback(row_id, "Visual Failed ❌")

                    # --- Step 3: Video Prompt ---
                    if self.generate_video and self.video_master_prompt and (not rowdata.get('existing_video') or force_update):
                        if progress_callback: progress_callback(row_id, "Step 3: Video Design... 🎬")
                        vid_res = self.generate_focused_prompt("video", final_title, final_style)
                        if vid_res:
                            vid_res = vid_res.replace("**", "").replace("__", "").strip()
                            self.update_output_data(row_id, {"video_prompt": vid_res})
                            if progress_callback: progress_callback(row_id, self.t("log_gemini_video_saved"))
                        else:
                            if progress_callback: progress_callback(row_id, "Video Failed ❌")

                    processed_count += 1
                    if i < len(rows) - 1:
                        time.sleep(5)
                
                return failed_ids

            processed_count = 0
            # PASS 1
            failed_in_pass1 = process_pending(pending_rows, pass_num=1)
            
            # PASS 2 (Retry failed ones)
            if failed_in_pass1:
                logger.info(f"Retrying {len(failed_in_pass1)} failed songs in Pass 2...")
                process_pending(failed_in_pass1, pass_num=2)
            
            return processed_count

        except Exception as e:
            logger.error(f"Gemini error: {e}")
            if progress_callback: progress_callback("global", f"Gemini Process Error: {e}")
        finally:
             logger.info("Gemini finished.")

    def generate_content(self, theme, style=None):
        try:
            if getattr(self, "_page_injected", False):
                logger.warning("Prompt injected in this tab already. Forcing a page refresh for clean state...")
                if self.progress_callback_internal: self.progress_callback_internal("global", "Sayfa yenileniyor (Temizleme) 🔄")
                try:
                    self.browser.goto(self.base_url, page=self.tab)
                except Exception: pass
                self._page_injected = False
                time.sleep(3)

            input_box = None
            # Gemini uses Quill editor (ql-editor) which is a div contenteditable
            possible_selectors = ["div.ql-editor", "div[contenteditable='true']", "textarea[aria-label*='prompt']", "textarea"]
            for selector in possible_selectors:
                if self.browser.is_visible(selector, page=self.tab):
                    input_box = selector
                    break
            
            if not input_box: 
                logger.error("Could not find Gemini input box.")
                if self.progress_callback_internal: self.progress_callback_internal("global", self.t("log_gemini_no_input"))
                return None

            # Using .replace() instead of .format() for better robustness against unknown braces in template
            full_prompt = self.master_prompt_template.replace("{theme}", str(theme)).replace("{language}", str(self.language))
            
            # Inject style instruction if provided
            if style:
                full_prompt += f"\n\n[IMPORTANT] Target Music Style: {style}\nPlease ensure all outputs (lyrics, art, video, style) follow this style precisely."

            # Error detection - If Gemini has a visible error toast, the textbox might be stuck.
            # We try to clear it before injecting.
            error_selectors = [".error-message", "div[role='alert']", ".error-toast"]
            has_error = False
            for es in error_selectors:
                if self.tab.locator(es).is_visible(timeout=500):
                    has_error = True; break
            
            if has_error:
                logger.warning("Gemini page error detected. Attempting to scroll/clear.")
                if self.progress_callback_internal: self.progress_callback_internal("global", self.t("log_gemini_error_trigger"))
                self._start_new_chat() # Reset chat state
                time.sleep(2)

            logger.info("Injecting prompt into Gemini (Instant Block Mode)...")
            if self.progress_callback_internal: self.progress_callback_internal("global", self.t("log_gemini_injecting"))
            
            # CRITICAL WAIT: Give the new chat's DOM a chance to attach fully and settle
            try:
                self.tab.locator(input_box).wait_for(state="attached", timeout=5000)
                time.sleep(2.0)
            except Exception: pass
            
            self.tab.click(input_box)
            time.sleep(1)
            
            # BYPASS HUMANIZER: Use self.tab.fill for instant injection as requested by user.
            # This ensures Gemini sees the entire prompt as a single unit.
            self.tab.fill(input_box, full_prompt)
            self._page_injected = True
            
            # Dispatch events to ensure Quill editor internal state updates correctly
            self.tab.evaluate(f"""(sel) => {{
                const el = document.querySelector(sel);
                if (el) {{
                    el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                    el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                    // Specific to some versions of Quill/Gemini
                    if (el.innerText === "") el.innerText = " "; 
                }}
            }}""", input_box)
            
            time.sleep(2)
            
            # Click Send button
            send_btn_selector = "button.send-button:not(.stop)"
            if self.tab.locator(send_btn_selector).is_visible(timeout=3000):
                logger.info("Clicking Send button...")
                self.tab.click(send_btn_selector)
            else:
                logger.warning("Send button not active/visible, trying Enter key fallback.")
                self.tab.keyboard.press("Enter")
            
            response_text = self._wait_for_response(input_box=input_box)
            if not response_text:
                # Backup: Try to find any message that looks like a response if count logic failed
                candidates = self.tab.locator("message-content").all()
                if not candidates: candidates = self.tab.locator(".model-response-text").all()
                if candidates:
                    logger.info("Count logic failed, but candidates found. Attempting to use last candidate.")
                    response_text = candidates[-1].inner_text().strip()
            
            if not response_text:
                return None
            
            result = {}
            lines = response_text.split('\n')
            current_section = None
            buffer = []
            
            for line in lines:
                clean_line = line.strip()
                # Strip markdown bolding and other markers for tag detection
                tag_check = clean_line.replace("*", "").replace("#", "").strip().lower()
                
                if tag_check.startswith("başlık"):
                    if current_section: result[current_section] = "\n".join(buffer).strip()
                    parts = clean_line.split(":", 1)
                    val = parts[1].strip() if len(parts) > 1 else ""
                    result["title"] = val.replace("[", "").replace("]", "")
                    current_section = None
                    buffer = []
                elif tag_check.startswith("sözler"):
                    if current_section: result[current_section] = "\n".join(buffer).strip()
                    current_section = "lyrics"
                    parts = clean_line.split(":", 1)
                    buffer = [parts[1].strip()] if len(parts) > 1 and parts[1].strip() else []
                elif tag_check.startswith("stil"):
                    if current_section: result[current_section] = "\n".join(buffer).strip()
                    current_section = "style"
                    parts = clean_line.split(":", 1)
                    buffer = [parts[1].strip()] if len(parts) > 1 and parts[1].strip() else []
                elif tag_check.startswith("görsel prompt"):
                    if current_section: result[current_section] = "\n".join(buffer).strip()
                    current_section = "visual_prompt"
                    parts = clean_line.split(":", 1)
                    buffer = [parts[1].strip()] if len(parts) > 1 and parts[1].strip() else []
                elif tag_check.startswith("video prompt"):
                    if current_section: result[current_section] = "\n".join(buffer).strip()
                    current_section = "video_prompt"
                    parts = clean_line.split(":", 1)
                    buffer = [parts[1].strip()] if len(parts) > 1 and parts[1].strip() else []
                else:
                    if current_section: buffer.append(line)
            
            if current_section: result[current_section] = "\n".join(buffer).strip()
            
            # Check for Outro (YARIM Status trigger) + Ensure Style is generated
            if "lyrics" in result:
                l_text = result["lyrics"].lower()
                has_outro = "[outro]" in l_text or "outro" in l_text[-150:]
                has_style = "style" in result and len(result["style"]) > 3
                
                if not has_outro or not has_style:
                    logger.warning(f"Song marked as YARIM. Outro found: {has_outro}, Style found: {has_style}")
                    result["status"] = "YARIM"

            # Unify art prompt naming
            if "visual_prompt" in result:
                result["cover_art_prompt"] = result["visual_prompt"]
                
            return result
        except Exception as e:
            logger.error(f"Gemini Step Error: {e}")
            logger.error(traceback.format_exc())
            return None
    def generate_focused_prompt(self, ptype, title, style):
        """Generates a focused prompt (visual or video) as a separate interaction."""
        try:
            # FIX: Use string selector directly to avoid "Locator not JSON serializable"
            input_box_selector = 'div[contenteditable="true"]'
            
            template = self.visual_master_prompt if ptype == "visual" else self.video_master_prompt
            if not template: return None

            full_prompt = template.replace("{title}", str(title)).replace("{style}", str(style))
            
            # CRITICAL WAIT: Give the DOM a chance to settle
            try:
                self.tab.locator(input_box_selector).wait_for(state="attached", timeout=5000)
                time.sleep(2.0)
            except Exception: pass
            
            self.tab.click(input_box_selector)
            time.sleep(1)
            self.browser.fill(input_box_selector, full_prompt, page=self.tab)
            time.sleep(2)
            self.tab.keyboard.press("Enter")
            
            return self._wait_for_response(input_box=input_box_selector)
        except Exception as e:
            logger.error(f"Gemini {ptype} Step Error: {e}")
            return None

    def _wait_for_response(self, timeout=120, input_box=None):
        """Helper to wait for Gemini response to appear and stabilize."""
        logger.info(f"Waiting for Gemini response (Global Timeout: {timeout}s)...")
        if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
            self.progress_callback_internal("global", self.t("log_gemini_wait", timeout=timeout))
        
        start_time = time.time()
        
        # Capture initial state
        initial_candidates = self.tab.locator("message-content").all()
        if not initial_candidates: initial_candidates = self.tab.locator(".model-response-text").all()
        initial_count = len(initial_candidates)
        
        last_text = ""
        stable_count = 0
        has_started = False
        
        while time.time() - start_time < timeout:
            elapsed = time.time() - start_time
            
            # 1. Stuck Detection: If after 60 seconds (User requested 1 min) we still have no new message content, fail early
            if not has_started and elapsed > 60:
                logger.error(f"Gemini stuck detection: No response started after {elapsed:.1f}s. Failing early.")
                if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
                    self.progress_callback_internal("global", self.t("log_gemini_stuck", elapsed=int(elapsed)))
                return None

            # 1.1 Error Detection during wait
            error_selectors = [".error-message", "div[role='alert']", ".error-toast", ".toast-container"]
            for es in error_selectors:
                try:
                    if self.tab.locator(es).is_visible(timeout=200):
                        logger.error(f"Gemini reported an error during generation: {es}")
                        if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
                             self.progress_callback_internal("global", self.t("log_gemini_error_trigger"))
                        return None
                except Exception: pass
            
            # 1.2 Text Box Re-dump Detection (Gemini "Bir hata oluştu" invisible failure check)
            if input_box and has_started and elapsed > 5:
                try:
                    box_text = self.tab.locator(input_box).inner_text(timeout=500).strip()
                    if len(box_text) > 20: # If prompt is spat back out
                        logger.error("Gemini crashed and dumped prompt back into input box. Failing early.")
                        if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
                             self.progress_callback_internal("global", "Gemini Hatası (Erken İptal) ❌")
                        return None
                except Exception: pass

            # 2. Check for 'Stop generating' button
            stop_btn_visible = False
            try:
                stop_btn = self.tab.locator("button[aria-label*='Stop' i], button[aria-label*='Durdur' i]").first
                if stop_btn.is_visible(timeout=500):
                    stop_btn_visible = True
            except Exception: pass

            candidates = self.tab.locator("message-content").all() 
            if not candidates: candidates = self.tab.locator(".model-response-text").all()
            
            if len(candidates) > initial_count or (len(candidates) == initial_count and initial_count > 1):
                # Use the last candidate as our potential response
                current_text = candidates[-1].inner_text().strip()
                
                # If message is still empty or just has a loading pulse, keep waiting
                if not current_text or len(current_text) < 10:
                    time.sleep(2)
                    continue

                if not has_started:
                    logger.info(f"Gemini started responding after {elapsed:.1f}s.")
                    if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
                        self.progress_callback_internal("global", self.t("log_gemini_started", elapsed=int(elapsed)))
                    has_started = True

                if current_text == last_text:
                    if not stop_btn_visible:
                        stable_count += 1
                        if stable_count >= 2: # 4 seconds of stability
                            logger.info(f"Response finished and stabilized. Length: {len(current_text)}")
                            if hasattr(self, "progress_callback_internal") and self.progress_callback_internal:
                                self.progress_callback_internal("global", self.t("log_gemini_finished", length=len(current_text)))
                            return current_text
                    else:
                        logger.info("Text is stable but 'Stop' button is still visible. Gemini might be thinking...")
                        stable_count = 0 
                else:
                    last_text = current_text
                    stable_count = 0
            
            time.sleep(2)
        
        logger.error(f"Gemini response timed out after {timeout}s.")
        return None

    
    def generate_art_prompts(self, max_count=None, target_ids=None, progress_callback=None):
        """Step 3a: Generates ONLY the text prompt for Cover Art."""
        return self._run_art_step(mode="prompt", max_count=max_count, target_ids=target_ids, progress_callback=progress_callback)

    def generate_art_images(self, max_count=None, target_ids=None, progress_callback=None):
        """Step 3b: Generates Image from existing 'cover_art_prompt'."""
        return self._run_art_step(mode="image", max_count=max_count, target_ids=target_ids, progress_callback=progress_callback)

    def _run_art_step(self, mode="prompt", max_count=None, target_ids=None, progress_callback=None):
        try:
            logger.info(f"Starting Art Step (Mode: {mode})...")
            
            if not os.path.exists(self.output_path):
                # Attempt to initialize from input if available
                # Assuming input is in same dir as 'input_songs.xlsx'
                base_dir = os.path.dirname(self.output_path)
                input_path = os.path.join(base_dir, "input_songs.xlsx")
                
                if os.path.exists(input_path):
                    logger.warning(f"Output file not found at {self.output_path}. Initializing from {input_path}...")
                    import shutil
                    shutil.copy(input_path, self.output_path)
                    if progress_callback: progress_callback("global", "Initialized Output File from Input.")
                else:
                    logger.error(f"Output file not found at {self.output_path} and Input not found.")
                    if progress_callback: progress_callback("global", "Error: Missing Input/Output Files! Run Step 1.")
                    return 0
            
            wb = openpyxl.load_workbook(self.output_path, data_only=True)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(ws[1]) if cell.value}
            
            rows_to_process = []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):                # ID Logic (Robust)
                id_idx = headers.get('id')
                if id_idx is None: id_idx = 0 # Default to column A
                rid = str(row[id_idx]) if id_idx < len(row) and row[id_idx] is not None else ""
                if target_ids and rid not in target_ids: 
                    continue
                
                # Needs cover_art_prompt but NO cover_art_path
                p_idx = headers.get('cover_art_prompt')
                v_idx = headers.get('visual_prompt')
                i_idx = headers.get('cover_art_path')
                
                # Dynamic value lookup
                prompt_val = row[p_idx] if p_idx is not None else None
                # Fallback to visual_prompt if cover_art_prompt is empty
                if (not prompt_val or str(prompt_val).strip() == "") and v_idx is not None:
                     prompt_val = row[v_idx]
                
                path_val = row[i_idx] if i_idx is not None else None
                
                if mode == "prompt":
                    if not prompt_val or str(prompt_val).strip() == "":
                        rows_to_process.append({"id": rid, "row": row, "_row_idx": i, "headers": headers})
                    else:
                        logger.info(f"Skipping Art Prompt for {rid}: Prompt already exists (Found in cover_art_prompt or visual_prompt).")
                elif mode == "image":
                    if prompt_val and str(prompt_val).strip() != "" and not path_val:
                        rows_to_process.append({"id": rid, "row": row, "_row_idx": i, "headers": headers, "art_prompt": prompt_val})
                    elif not prompt_val or str(prompt_val).strip() == "":
                        logger.warning(f"Skipping Art Image for {rid}: No prompt found in 'cover_art_prompt' or 'visual_prompt'.")
                        if progress_callback: progress_callback(rid, "Error: No Art Prompt Found! ❌")
                    else:
                        logger.info(f"Skipping Art Image for {rid}: Image already exists or not needed.")

            if not rows_to_process:
                logger.info(f"Nothing to process for Step 3/4 (Mode: {mode}) for IDs: {target_ids}")
                return 0
            
            if max_count: rows_to_process = rows_to_process[:max_count]
            
            self.browser.start()
            self.tab.bring_to_front()
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Waiting {self.startup_delay}s (Startup Delay)...")
                time.sleep(self.startup_delay)
            self.browser.goto(self.base_url, page=self.tab)
            
            # Login Check
            if "accounts.google.com" in self.tab.url:
                 if progress_callback: progress_callback("global", "Login Required!")
                 time.sleep(5)

            count = 0
            for item in rows_to_process:
                rid = item['id']
                
                if mode == "prompt":
                    if progress_callback: progress_callback(rid, "Generating Art Prompt...")
                    # Get Data
                    row = item["row"]
                    headers = item["headers"]
                    title = row[headers.get('title', 0)] or ""
                    
                    # Fill Master Prompt
                    full_prompt = self.art_master_prompt.replace("{title}", str(title))
                    
                    # Send
                    result_text = self._send_and_get_text(full_prompt)
                    if result_text:
                        self.update_output_data(rid, {"cover_art_prompt": result_text})
                        if progress_callback: progress_callback(rid, "Prompt Generated ✅")
                        count += 1
                    else:
                        if progress_callback: progress_callback(rid, "Prompt Failed ❌")

                elif mode == "image":
                    if progress_callback: progress_callback(rid, "Generating Image... 🎨")
                    art_prompt = item["art_prompt"]
                    
                    # Trigger Image Gen
                    img_trigger = f"Generate an image based on this description:\n\n{art_prompt}"
                    save_path = self._generate_and_download_image(img_trigger, rid)
                    
                    if save_path:
                        self.update_output_data(rid, {"cover_art_path": save_path})
                        if progress_callback: progress_callback(rid, "Image Saved 🖼️")
                        count += 1
                    else:
                        if progress_callback: progress_callback(rid, "Image Failed ❌")
            
            return count
        except Exception as e:
            logger.error(f"Art Error: {e}")
            raise e
        finally:
            logger.info("Art Step Finished")

    def _send_and_get_text(self, prompt):
        try:
            input_box = None
            possible_selectors = ["div[contenteditable='true']", "textarea[aria-label*='prompt']", "textarea"]
            for selector in possible_selectors:
                if self.browser.is_visible(selector, page=self.tab):
                    input_box = selector
                    break
            
            if not input_box: return None
            
            self.browser.fill(input_box, prompt, page=self.tab)
            time.sleep(1)
            self.tab.keyboard.press("Enter")
            time.sleep(10) # Wait for text gen
            
            candidates = self.tab.locator("message-content").all()
            if candidates: return candidates[-1].inner_text().strip()
            return None
        except Exception: return None

    def _generate_and_download_image(self, prompt, rid):
        try:
            input_box = "div[contenteditable='true']"
            if not self.browser.is_visible(input_box, page=self.tab): 
                logger.warning("Gemini input box not found for image gen.")
                return None
            
            # Step 0: Count existing image containers to avoid stale detection
            initial_containers = self.tab.locator("div.image-set-container, sac-image-set, sac-single-image-set").count()
            logger.info(f"Initial image container count: {initial_containers}")

            logger.info("Sending image generation prompt to Gemini...")
            self.browser.fill(input_box, prompt, page=self.tab)
            self.tab.keyboard.press("Enter")
            
            # Step 1: Wait for generation to start (indicator to appear)
            logger.info("Waiting for generation indicator to appear...")
            indicator_selectors = [
                "text='Görüntüler oluşturuluyor...'", 
                "text='Generating images...'",
                ".loading-indicator", 
                "div:has-text('Nanobanana')"
            ]
            
            # Brief wait for indicator to appear
            time.sleep(3)
            
            # Step 2: Poll for generation completion (New container appearing)
            logger.info("Waiting for NEW images to be fully rendered (polling up to 90s)...")
            found_img = None
            
            for attempt in range(18): # 18 * 5s = 90s
                # Check if still generating
                is_loading = False
                for sel in indicator_selectors:
                    try:
                        if self.tab.locator(sel).first.is_visible():
                            is_loading = True
                            break
                    except Exception: pass
                
                if is_loading:
                    logger.info(f"   Still generating (Attempt {attempt+1}/18)...")
                else:
                    # Look for NEW image containers
                    current_containers = self.tab.locator("div.image-set-container, sac-image-set, sac-single-image-set")
                    current_count = current_containers.count()
                    
                    if current_count > initial_containers:
                        # New container found!
                        container = current_containers.last
                        if container.is_visible():
                            logger.info("Generation finished (New container detected). Finding primary image...")
                            time.sleep(3) # Extra wait for full high-res render
                            
                            # Find the largest image in the newest container
                            images = container.locator("img").all()
                            if images:
                                for img in reversed(images):
                                    box = img.bounding_box()
                                    if box and box['width'] > 300 and box['height'] > 300:
                                        src = img.get_attribute("src") or ""
                                        if "blob:" in src or "googleusercontent" in src:
                                            found_img = img
                                            break
                            if found_img: break
                    else:
                        logger.info(f"   Waiting for new container (Current: {current_count}, Initial: {initial_containers})...")
                
                time.sleep(5)
            
            if found_img:
                img_dir = self.output_dir
                os.makedirs(img_dir, exist_ok=True)
                save_path = os.path.join(img_dir, f"{rid}.png")
                logger.info(f"Found image. Saving with screenshot to: {save_path}")
                found_img.screenshot(path=save_path)
                return save_path
            
            logger.warning("Image generation timed out or no image element discovered.")
            return None
        except Exception as e:
            logger.error(f"Image gen error: {e}")
            return None

    
    
    def close(self):
        # We no longer stop the browser here, as it's shared/managed by the GUI
        pass

    def update_output_data(self, row_id, data):
        try:
            wb = openpyxl.load_workbook(self.output_path)
            ws = wb.active
            col_map = {str(cell.value).strip().lower(): cell.column for cell in ws[1] if cell.value}
            
            # Ensure columns exist
            for key in ["title", "lyrics", "visual_prompt", "video_prompt", "style", "suno_style", "status", "cover_art_prompt", "cover_art_path"]:
                if key not in col_map:
                    new_idx = ws.max_column + 1
                    ws.cell(row=1, column=new_idx, value=key)
                    col_map[key] = new_idx
            
            target_row = None
            id_col_num = col_map.get("id", 1)
            for row in range(2, ws.max_row + 1):
                cur_id = str(ws.cell(row=row, column=id_col_num).value or "").strip()
                if cur_id == str(row_id).strip():
                    target_row = row
                    break
            
            if not target_row: target_row = ws.max_row + 1
            
            fill_new = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # Helper to update cell if empty or force update(though we default to skip if exists)
            def update_cell(col_name, new_value):
                if col_name in col_map:
                    cell = ws.cell(row=target_row, column=col_map[col_name])
                    if not cell.value: # Only update if empty
                        cell.value = new_value
                        cell.fill = fill_new
            
            ws.cell(row=target_row, column=col_map["id"], value=row_id)
            
            if "title" in data: update_cell("title", data["title"])
            if self.use_gemini_lyrics and "lyrics" in data: update_cell("lyrics", data["lyrics"])
            if "suno_style" in data: update_cell("suno_style", data["suno_style"])
            if self.generate_style and "style" in data: update_cell("style", data["style"])
            if self.generate_visual and "visual_prompt" in data: update_cell("visual_prompt", data["visual_prompt"])
            if self.generate_video and "video_prompt" in data: update_cell("video_prompt", data["video_prompt"])
            if "cover_art_prompt" in data: update_cell("cover_art_prompt", data["cover_art_prompt"])
            if "cover_art_path" in data: update_cell("cover_art_path", data["cover_art_path"])
            
            wb.save(self.output_path)
        except Exception as e:
            logger.error(f"Save error: {e}")

if __name__ == "__main__":
    gemini = GeminiPrompter()
    gemini.run()
