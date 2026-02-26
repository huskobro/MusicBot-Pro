import time
import os
import logging
import openpyxl
import re
import json
from browser_controller import BrowserController

logger = logging.getLogger(__name__)

from suno_config import SunoConfig
from suno_excel import SunoExcelMixin
from suno_downloader import SunoDownloaderMixin
from suno_ui import SunoUIMixin

class UserStoppedException(Exception):
    pass

class SunoGenerator(SunoExcelMixin, SunoDownloaderMixin, SunoUIMixin):
    def __init__(self, project_file, output_dir="data/output", delay=10, startup_delay=5, browser=None,
                 audio_influence=25, vocal_gender="Default", lyrics_mode="Default", 
                 weirdness=50, style_influence=50, persona_link="", turbo=False, xlsx_lock=None):
        self.xlsx_lock = xlsx_lock
        self.config = SunoConfig()
        self.project_file = project_file
        self.metadata_path = project_file # Backward compat
        
        self.output_dir = output_dir
        try:
            self.delay = int(delay)
            self.startup_delay = int(startup_delay)
        except (ValueError, TypeError):
            self.delay = 10
            self.startup_delay = 5
            
        self.browser = browser or BrowserController()
        self.tab = self.browser.page
        self.base_url = "https://suno.com/create"
        self.stop_requested = False
        self.turbo = turbo
        
        # Advanced Params
        try:
            self.audio_influence = int(audio_influence)
            self.weirdness = int(weirdness)
            self.style_influence = int(style_influence)
        except (ValueError, TypeError):
            self.audio_influence = 25
            self.weirdness = 50
            self.style_influence = 50
            
        self.vocal_gender = vocal_gender
        self.lyrics_mode = lyrics_mode
        self.persona_link = str(persona_link).strip()
        
        import sys
        self.mod = "Meta" if sys.platform == "darwin" else "Control"

    def _check_stop(self):
        if self.stop_requested:
            logger.info("Kullanıcı tarafından durdurma isteği alındı.")
            raise UserStoppedException("User requested stop")
        
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
        self.stop_requested = False

    def run(self, max_count=None, target_ids=None, progress_callback=None, force_update=False):
        try:
            if not self.browser.context:
                self.browser.start()
            
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Bekleniyor: {self.startup_delay}sn (Başlangıç Gecikmesi)...")
                time.sleep(self.startup_delay)
            
            # --- Persona Workflow (Before reaching /create) ---
            if self.persona_link:
                success = self._setup_persona_workflow(progress_callback)
                if not success:
                    logger.warning("Persona akışı başarısız oldu, doğrudan yenilemeye geçiliyor.")
                    self.browser.goto(self.base_url, page=self.tab)
            else:
                self.browser.goto(self.base_url, page=self.tab)
            
            time.sleep(5) # Wait for page structure

            # Always Ensure v5 for reliability
            self._ensure_v5_active()

            if "login" in self.tab.url or self.browser.is_visible("text='Log In'", page=self.tab):
                logger.warning("--- GİRİŞ GEREKLİ ---")
                if progress_callback: progress_callback("global", "Giriş Gerekli! Lütfen Chrome'u kontrol edin.")
                
                # Check repeatedly for login
                for _ in range(60): # Wait up to 5 mins for user login
                    if "create" in self.tab.url and "login" not in self.tab.url:
                        break
                    time.sleep(5)
            
            if "create" not in self.tab.url:
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)

            if not os.path.exists(self.metadata_path):
                logger.error("Sonuç dosyası (Excel) bulunamadı.")
                return 0

            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): cell.column - 1 for cell in ws[1] if cell.value}
            
            rows_data = []
            
            # Normalize target_ids for robust matching
            target_ids_set = set(str(t).strip().lower() for t in target_ids) if target_ids else None

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rid_orig = row[headers.get('id', 0)] if 'id' in headers else ""
                rid = str(rid_orig).strip().lower()
                
                # Check target_ids
                if target_ids_set and rid not in target_ids_set:
                    continue
                
                status = str(row[headers.get('status', 0)]).lower() if 'status' in headers else ""
                
                # Determine if song already has music (M1 or M2 presence)
                # In scan_materials, we check files. Here we can also check the status column.
                is_done = "done" in status or "generated" in status
                
                if not force_update and is_done:
                    logger.info(f"{rid_orig} ID'li şarkı Suno'da zaten üretilmiş. Atlanıyor.")
                    if progress_callback: progress_callback(str(rid_orig), "Atlandı: Zaten Üretilmiş ✅")
                    continue
                
                row_dict = {key: row[idx] for key, idx in headers.items() if idx < len(row)}
                row_dict['_row_idx'] = i 
                rows_data.append(row_dict)
            
            if not rows_data:
                logger.info("Suno'da üretilecek şarkı bulunamadı.")
                return 0

            if max_count and max_count > 0:
                rows_data = rows_data[:max_count]

            processed_count = 0
            for i, row_dict in enumerate(rows_data): 
                rid = str(row_dict.get('id', ''))
                
                if i > 0:
                    if progress_callback: progress_callback(rid, f"Waiting {self.delay}s...")
                    time.sleep(self.delay)
                
                self.update_row_status(row_dict['_row_idx'], "Processing")
                if progress_callback: progress_callback(rid, "Generating Music... 🎵")
                
                try:
                    self.browser.ensure_alive()
                    success = self.process_row(row_dict, progress_callback=progress_callback)
                    if success:
                        self.update_row_status(row_dict['_row_idx'], "Generated")
                        if progress_callback: progress_callback(rid, "Music Generated! 🎵")
                        processed_count += 1
                    else:
                        self.update_row_status(row_dict['_row_idx'], "Failed")
                        if progress_callback: progress_callback(rid, "Generation Failed ❌")
                except Exception as row_error:
                    logger.error(f"Row error: {row_error}")
                    self.update_row_status(row_dict['_row_idx'], "Failed")
                    if progress_callback: progress_callback(rid, "Error ❌")

            return processed_count

        except Exception as e:
            logger.error(f"Suno error: {e}")
            raise e
        finally:
            if hasattr(self, "flush_excel_cache"):
                self.flush_excel_cache()
            logger.info("Suno finished.")

    def run_batch(self, target_ids=None, progress_callback=None, stats_callback=None, force_update=False, op_mode="full"):
        """
        Orchestrates the BATCH workflow:
        1. Generate ALL requested songs (Phase 1)
        2. Wait for & Download ALL requested songs (Phase 2)
        """
        try:
            batch_start_time = time.time()
            if not self.browser.context:
                self.browser.start()
            
            if self.startup_delay > 0:
                if progress_callback: progress_callback("global", f"Bekleniyor: {self.startup_delay}sn (Başlangıç Gecikmesi)...")
                time.sleep(self.startup_delay)
            
            # --- Common Setup (Persona / Login / v5) ---
            # Reuse logic from run(), or extract to common _init_session()
            if self.persona_link:
                success = self._setup_persona_workflow(progress_callback)
                if not success:
                    self.browser.goto(self.base_url, page=self.tab)
            else:
                self.browser.goto(self.base_url, page=self.tab)
            
            time.sleep(5)
            self._ensure_v5_active()

            if "login" in self.tab.url or self.browser.is_visible("text='Log In'", page=self.tab):
                logger.warning("--- GİRİŞ GEREKLİ ---")
                if progress_callback: progress_callback("global", "Login Required! Please check Chrome.")
                for _ in range(60): 
                    if "create" in self.tab.url and "login" not in self.tab.url: break
                    time.sleep(5)
            
            if "create" not in self.tab.url:
                self.browser.goto(self.base_url, page=self.tab)
                time.sleep(3)

            # --- Prepare Data ---
            if not os.path.exists(self.metadata_path): return 0
            
            wb = openpyxl.load_workbook(self.metadata_path)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): cell.column - 1 for cell in ws[1] if cell.value}
            
            rows_data = []
            target_ids_set = set(str(t).strip().lower() for t in target_ids) if target_ids else None
            
            self._batch_stats = {"total": 0, "success": 0, "failed": 0, "threads": 2}
            
            def trigger_dashboard_update():
                if stats_callback:
                    elapsed = time.time() - batch_start_time
                    t_str = f"{int(elapsed//60):02d}:{int(elapsed%60):02d}"
                    stats_callback(
                        total=self._batch_stats["total"],
                        success=self._batch_stats["success"],
                        failed=self._batch_stats["failed"],
                        time=t_str,
                        threads=self._batch_stats["threads"]
                    )
                    
            # Initial stats trigger
            trigger_dashboard_update()

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rid_orig = row[headers.get('id', 0)] if 'id' in headers else ""
                rid = str(rid_orig).strip().lower()
                
                if target_ids_set and rid not in target_ids_set: continue
                
                status = str(row[headers.get('status', 0)]).lower() if 'status' in headers else ""
                is_done = "done" in status or "generated" in status
                
                # Determine if song already has music (M1 or M2 presence)
                # In scan_materials, we check files. Here we can also check the status column.
                
                # [FIX] In dl_only or full mode, we DON'T skip generated items here 
                # because we need them in rows_data for the Phase 2 (Download).
                # We will skip them explicitly inside Phase 1 (Generation).
                if not force_update and is_done and op_mode == "gen_only":
                    continue
                
                row_dict = {key: row[idx] for key, idx in headers.items() if idx < len(row)}
                row_dict['_row_idx'] = i 
                rows_data.append(row_dict)
            
            if not rows_data:
                logger.info("Toplu üretim için şarkı bulunamadı.")
                return 0
                
            self._batch_stats["total"] = len(rows_data)
            trigger_dashboard_update()

            generated_ids = []
            
            # --- PHASE 1: BATCH GENERATION ---
            if op_mode in ["full", "gen_only"]:
                logger.info("--- Toplu Üretim Aşaması Başlıyor ---")
                if progress_callback: progress_callback("global", "Toplu Üretim Başlatılıyor... 🚀")
                
                for i, row_dict in enumerate(rows_data):
                    if self.stop_requested: break
                    
                    rid = str(row_dict.get('id', ''))
                    
                    # Ensure a fresh state for every song in batch mode by reloading
                    if i >= 0: # Reload for every song to be absolutely sure
                        if self.persona_link:
                            logger.info(f"Toplu Üretim: {rid} ID'li şarkı için Persona tekrar aktif ediliyor...")
                            self._setup_persona_workflow(progress_callback)
                        else:
                            logger.info(f"Toplu Üretim: {rid} ID'li şarkı için sayfa yenileniyor (/create)...")
                            self.browser.goto(self.base_url, page=self.tab)
                        
                        time.sleep(3)
                        self._ensure_v5_active()
                    
                    if i > 0:
                        time.sleep(self.delay)
                    
                    self.update_row_status(row_dict['_row_idx'], "Processing")
                    
                    # [Safety] Skip generation if already done (and not forced)
                    status = str(row_dict.get('status', '')).lower()
                    if not force_update and ("generated" in status or "done" in status):
                        logger.info(f"{rid} ID'li şarkı için Üretim Atlanıyor (Zaten Üretilmiş)")
                        generated_ids.append(rid) # Move straight to potential download queue
                        continue

                    if progress_callback: progress_callback(rid, "Sıraya Alınıyor... 🎵")
                    
                    success = self._generate_single_no_wait(row_dict, progress_callback)
                    if success:
                        generated_ids.append(rid)
                        self.update_row_status(row_dict['_row_idx'], "Generating...", music_status="Sıraya Alındı") 
                    else:
                        self.update_row_status(row_dict['_row_idx'], "Failed")
                
                logger.info(f"Toplu Üretim Aşaması Tamamlandı. {len(generated_ids)} şarkı sıraya eklendi.")
            
            from suno_config import DownloadContext, SongState
            download_queue = []

            # --- PHASE 1.5: Identify Resume Items ---
            if op_mode in ["full", "dl_only"]:
                for row_dict in rows_data:
                    rid = str(row_dict.get('id', ''))
                    title = str(row_dict.get('title', 'Song'))
                    status = str(row_dict.get('status', '')).lower()
                    
                    is_target = target_ids and (rid in target_ids)
                    
                    if rid not in generated_ids:
                        if is_target or "generating" in status or "processing" in status or "suno bekleniyor" in status:
                             logger.info(f"Adding {rid} to download queue (is_target: {is_target}, status: {status})")
                             download_queue.append(DownloadContext(rid=rid, title=title, row_idx=row_dict['_row_idx']))
                    else:
                         download_queue.append(DownloadContext(rid=rid, title=title, row_idx=row_dict['_row_idx']))

            # --- PHASE 1.9: Pre-flight Search Check ---
            if op_mode in ["full", "dl_only"] and download_queue:
                # [NEW] Pre-flight check to verify search clips locator is visible on the current page
                logger.info("Arama motoru (Search clips) kontrol ediliyor...")
                search_ready = False
                search_selectors = ["input[aria-label='Search clips']", "input[placeholder='Search']"]
                for sel in search_selectors:
                    if self.tab.locator(sel).first.is_visible(timeout=3000):
                        search_ready = True; break
                
                if not search_ready:
                    logger.warning("DİKKAT: Arama kutusu (Search clips) bulunamadı! Arama gerektiren şarkılar indirilemeyebilir.")
                    if progress_callback: progress_callback("global", "Uyarı: Arama kutusu bulunamadı! ⚠️")

            # --- PHASE 2: BATCH DOWNLOAD ---
            if op_mode in ["full", "dl_only"] and download_queue:
                logger.info("--- Toplu İndirme Aşaması Başlıyor ---")
                if progress_callback: progress_callback("global", "Toplu İndirme Bekleniyor... ⬇️")
                
                # PRE-DOWNLOAD PHASE: SMART WAIT ROOM
                # Only needed when songs are being GENERATED (full mode)
                # In dl_only mode, songs are already generated - skip straight to download
                missing_counts = {rid: 0 for rid in generated_ids}
                
                if op_mode == "full":
                    pending_ids = list(generated_ids)
                    start_wait = time.time()
                    total_timeout = max(600, len(pending_ids) * 180) 
                    
                    logger.info(f"Hedefli Toplu İşlem: İndirmeden önce {len(pending_ids)} şarkının tamamlanması bekleniyor...")
                    
                    max_row_reached = 0
                    loop_count = 0
                    missing_counts = {rid: 0 for rid in pending_ids} # Track missing loops
                    
                    while pending_ids and (time.time() - start_wait < total_timeout):
                        if self.stop_requested: break
                        loop_count += 1
                        
                        # SESSION HARDENING: Check if tab is alive, if not RECOVER
                        try:
                            # Simple check to see if tab is responsive
                            self.tab.url
                        except Exception:
                            logger.warning("Browser tab lost or crashed! Attempting recovery...")
                            try:
                                if self.persona_link: self._setup_persona_workflow(progress_callback)
                                else: self.browser.goto(self.base_url, page=self.tab)
                                time.sleep(5)
                                self._ensure_v5_active()
                            except Exception as re_e:
                                logger.error(f"Recovery failed: {re_e}")
                                time.sleep(10)
                                continue

                        still_generating = []
                        current_max_idx = 0
                        all_targets_found_in_view = True
                        
                        # 1. SCAN AND COLLECT STATUS
                        found_this_loop = {}
                        try:
                            rows = self.tab.locator("div.clip-row")
                            row_count = rows.count()
                            for i in range(min(300, row_count)):
                                row_text = rows.nth(i).inner_text().lower()
                                for rid in pending_ids:
                                    if str(rid).lower() in row_text:
                                        found_this_loop[rid] = {"index": i, "ready": "generating" not in row_text}
                                        current_max_idx = max(current_max_idx, i)
                                        # If any target found in view, we'll track its max index
                        except Exception as scan_e:
                            logger.debug(f"Scan interrupted: {scan_e}")

                        for ctx in pending_ctxs:
                            if ctx.state != SongState.QUEUED: continue
                            
                            if ctx.rid in found_this_loop:
                                missing_counts[ctx.rid] = 0
                                if found_this_loop[ctx.rid]["ready"]:
                                    suno_title = f"{ctx.rid}_{ctx.title}"
                                    if self._check_if_ready(suno_title, ctx.rid, suffix="1") and \
                                       self._check_if_ready(suno_title, ctx.rid, suffix="2"):
                                        logger.info(f"ID {ctx.rid} is ready for download phase.")
                                        if progress_callback: progress_callback(ctx.rid, "Hazır! Beklemede... ✅")
                                        ctx.state = SongState.FOUND
                            else:
                                missing_counts[ctx.rid] = missing_counts.get(ctx.rid, 0) + 1
                                if missing_counts[ctx.rid] > 12:
                                    logger.warning(f"Timeout waiting for {ctx.rid} to appear. Skipping.")
                                    self.update_row_status(ctx.row_idx, status="Failed", dl_status="failed")
                                    if progress_callback: progress_callback(ctx.rid, "Üretilemedi / Bulunamadı ❌")
                                    ctx.state = SongState.FAILED
                                else:
                                    all_targets_found_in_view = False
                        
                        max_row_reached = max(max_row_reached, current_max_idx)

                        if not all_targets_found_in_view:
                            if max_row_reached > 5:
                                try:
                                    logger.info(f"Waterfall Scroll: Pushing down from row {max_row_reached} to find missing songs...")
                                    rows = self.tab.locator("div.clip-row")
                                    if rows.count() > max_row_reached:
                                        rows.nth(min(max_row_reached, rows.count()-1)).scroll_into_view_if_needed()
                                        time.sleep(1)
                                        self.tab.mouse.wheel(0, 2500)
                                except Exception: pass
                            else:
                                try:
                                    self.tab.mouse.wheel(0, 1500)
                                    time.sleep(1)
                                except Exception: pass
                        else:
                            logger.debug("All pending songs found in current view. No scrolling needed.")

                        if not all_targets_found_in_view and loop_count > 8:
                            missing_ctxs = [c for c in pending_ctxs if c.state == SongState.QUEUED and c.rid not in found_this_loop]
                            if missing_ctxs:
                                target_ctx = missing_ctxs[0]
                                logger.info(f"Using Search Fallback for missing ID {target_ctx.rid}...")
                                try:
                                    if self._search_for_song(target_ctx.rid, target_ctx.title):
                                        time.sleep(3)
                                        s_rows = self.tab.locator("div.clip-row")
                                        for si in range(min(10, s_rows.count())):
                                            s_text = s_rows.nth(si).inner_text().lower()
                                            if target_ctx.rid in s_text:
                                                if "generating" not in s_text:
                                                    target_ctx.state = SongState.FOUND
                                                    logger.info(f"Found missing ID {target_ctx.rid} via search!")
                                                break
                                except Exception as e:
                                    logger.error(f"Search fallback error: {e}")
                                
                                try:
                                    search_input = self.tab.locator("input[aria-label='Search clips']").first
                                    if search_input.is_visible():
                                        search_input.fill("")
                                        self.tab.keyboard.press("Enter")
                                        time.sleep(2)
                                except Exception: pass

                        time.sleep(3)
                        
                        pending_count = len([c for c in pending_ctxs if c.state == SongState.QUEUED])
                        if progress_callback: progress_callback("global", f"Bekleniyor: {pending_count} şarkı kaldı... ⏳")
                        time.sleep(10)
                        
                        # Periodic reload to refresh state (Every 120-150s)
                        if (time.time() - start_wait) > 30 and (time.time() - start_wait) % 150 < 15:
                            try:
                                logger.info("Periodic reload to refresh state...")
                                self.tab.keyboard.press("Escape")
                                time.sleep(1)
                                self.tab.reload(timeout=60000)
                                time.sleep(5)
                                self._ensure_v5_active()
                            except Exception as e:
                                logger.warning(f"Reload failed: {e}")
                else:
                    logger.info(f"dl_only mode: Skipping wait phase. Going straight to download for {len(download_queue)} songs.")

                # ACTUAL DOWNLOAD PHASE
                # Everything is ready, now we execute the persistent download logic
                # Only exclude songs that were never found (timed out in wait phase)
                completed_ids = []
                
                logger.info(f"İndirme sırası: {[c.rid for c in download_queue if c.state != SongState.FAILED]} ({len([c for c in download_queue if c.state != SongState.FAILED])} şarkı)")
                
                # Clear any leftover search
                try:
                    search_input = self.tab.locator("input[aria-label='Search clips']").first
                    if search_input.is_visible() and search_input.input_value().strip():
                        search_input.fill("")
                        self.tab.keyboard.press("Enter")
                        time.sleep(2)
                except Exception: pass
                
                # ===== PRE-PHASE A: REFRESH STATE =====
                # Force a reload to guarantee the DOM isn't stale before we start scraping
                logger.info("Taze veri çekmek için arayüz yenileniyor...")
                if progress_callback: progress_callback("global", "Sayfa yenileniyor... 🔄")
                try:
                    self.tab.reload(timeout=45000)
                    time.sleep(6)
                    self._ensure_v5_active()
                except Exception as e:
                    logger.warning(f"Aşama A öncesi yenileme hatası: {e}")

                # ===== PHASE A: FULL WATERFALL SCROLL =====
                # Scroll through the entire song list to load everything into DOM
                # Dynamically calculate passes based on queue size (approx 15-20 songs per scroll)
                target_passes = max(5, int(len(download_queue) / 10) + 2)
                
                logger.info(f"Aşama A: {len(download_queue)} şarkı için {target_passes} denemelik şelale kaydırması yapılıyor...")
                if progress_callback: progress_callback("global", "Şarkı listesi yükleniyor... ⬇️")
                try:
                    for scroll_pass in range(target_passes):
                        rows = self.tab.locator("div.clip-row")
                        current_count = rows.count()
                        if current_count > 0:
                            # Use locator.nth() and .scroll_into_view_if_needed() safely
                            rows.nth(current_count - 1).scroll_into_view_if_needed()
                            time.sleep(0.5)
                            self.tab.mouse.wheel(0, 3000)
                            time.sleep(2) # Give it slightly more time to load
                            new_count = self.tab.locator("div.clip-row").count()
                            logger.info(f"Kaydırma geçişi {scroll_pass+1}: {current_count} → {new_count} satır")
                            if new_count <= current_count:
                                break  # No more rows loading
                except Exception as e:
                    logger.warning(f"Şelale kaydırma hatası: {e}")
                
                # Scroll back to top 
                try:
                    self.tab.keyboard.press("Home")
                    time.sleep(1)
                except Exception: pass
                
                # ===== PHASE B: HYBRID WAIT ROOM & SINGLE-PASS INDEX =====
                # Poll DOM for song readiness, log detailed metrics, and index loaded ones
                logger.info("Aşama B: Akıllı Bekleme Odası & Tek Geçişli İndeksleme...")
                if progress_callback: progress_callback("global", "Şarkıların hazır olması bekleniyor... ⏳")
                
                song_index = {}  # {rid: [row_element_1, row_element_2, ...]}
                valid_ctxs = [c for c in download_queue if c.state != SongState.FAILED]
                pending_ctxs = list(valid_ctxs)
                seen_ctx_rids = set()
                wait_loop_count = 0
                
                start_wait = time.time()
                total_timeout = max(600, len(pending_ctxs) * 180)
                
                while pending_ctxs and (time.time() - start_wait < total_timeout):
                    if self.stop_requested: break
                    still_generating = []
                    wait_loop_count += 1
                    
                    rows = self.tab.locator("div.clip-row")
                    count = rows.count()
                    
                    for i in range(count):
                        try:
                            row = rows.nth(i)
                            row.scroll_into_view_if_needed()
                            row_text = row.inner_text().lower()
                            
                            is_gen = row.locator("text='Generating'").is_visible()
                            duration_loc = row.locator("text=/\\d{1,2}:\\d{2}/")
                            duration = duration_loc.first.inner_text().strip() if duration_loc.count() > 0 and duration_loc.first.is_visible() else "Mevcut Değil"
                            
                            for ctx in pending_ctxs:
                                # Loose match: "1234_" or both ID and title
                                if (ctx.rid.lower() + "_" in row_text) or (ctx.rid.lower() in row_text and str(ctx.title).lower() in row_text):
                                    seen_ctx_rids.add(ctx.rid)
                                    logger.info(f"[Aşama B Bekleme Odası] ID: {ctx.rid} | Başlık: {ctx.title} | Satır: {i} | Süre: {duration} | Üretiliyor: {is_gen}")
                                    
                                    # If it's NOT generating anymore, it's done enough to click. 
                                    # Don't require duration string to be perfect to avoid endless loop.
                                    if not is_gen:
                                        if ctx.rid not in song_index:
                                            song_index[ctx.rid] = []
                                        song_index[ctx.rid].append(row)
                                        if progress_callback: progress_callback(ctx.rid, "Hazır! Beklemede... ✅")
                                        
                        except Exception as parse_e:
                            logger.debug(f"Row {i} parse error in wait room: {parse_e}")
                            pass
                            
                    # Any ctx that is completely in the index is done waiting
                    for ctx in pending_ctxs:
                        if ctx.rid not in song_index:
                            if wait_loop_count > 2 and ctx.rid not in seen_ctx_rids:
                                logger.warning(f"[Aşama B] {ctx.rid} {wait_loop_count} şelale denemesi içinde bulunamadı, arama kutusuna gönderiliyor.")
                            else:
                                still_generating.append(ctx)
                            
                    pending_ctxs = still_generating
                    if not pending_ctxs:
                         break
                         
                    if progress_callback: progress_callback("global", f"Üretim Bekleniyor: {len(pending_ctxs)} şarkı kaldı... ⏳")
                    
                    try:
                        self.tab.mouse.wheel(0, 3000)
                    except Exception: pass
                    time.sleep(15)
                    
                    # Refresh occasionally
                    if (time.time() - start_wait) % 90 < 15:
                        logger.info("Refreshing page in wait room to rehydrate DOM elements...")
                        try:
                            self.tab.reload()
                            time.sleep(5)
                            self._ensure_v5_active()
                        except Exception: pass
                        
                found_count = len(song_index)
                not_found_ctxs = [c for c in valid_ctxs if c.rid not in song_index]
                logger.info(f"[Aşama B] Hazırlık tamamlandı. İndeks: {found_count} bulundu, {len(not_found_ctxs)} bulunamadı.")
                
                # ===== PHASE C: SEARCH FALLBACK FOR NOT-FOUND SONGS =====
                if not_found_ctxs:
                    logger.info(f"Aşama C: Sayfada bulunamayan {len(not_found_ctxs)} şarkı için arama motoru deneniyor: {[c.rid for c in not_found_ctxs]}")
                    if progress_callback: progress_callback("global", f"Arama ile {len(not_found_ctxs)} şarkı aranıyor... 🔍")
                    
                    for ctx in not_found_ctxs:
                        if self.stop_requested: break
                        
                        # SESSION CHECK
                        try: self.tab.url
                        except Exception:
                            logger.warning(f"{ctx.rid} aranmadan önce sekme çöktü. Kurtarılıyor...")
                            if self.persona_link: self._setup_persona_workflow()
                            else: self.browser.goto(self.base_url, page=self.tab)
                            time.sleep(5)
                            self._ensure_v5_active()
                            
                        logger.info(f"[Aşama C] {ctx.rid} ID'li şarkı aranıyor...")
                        if progress_callback: progress_callback(ctx.rid, f"Aranıyor... 🔍")
                        
                        if self._search_for_song(ctx.rid, ctx.title):
                            # Explicitly force lazy loading for search results
                            try:
                                time.sleep(2)
                                self.tab.keyboard.press("Home")
                                time.sleep(1)
                                self.tab.mouse.wheel(0, 3000)
                                time.sleep(2)
                                self.tab.keyboard.press("Home")
                                time.sleep(1)
                            except Exception: pass
                            
                            rows = self.tab.locator("div.clip-row")
                            occurrences = []
                            for i in range(rows.count()):
                                try:
                                    # Force visible via scroll into view block check
                                    r = rows.nth(i)
                                    if ctx.rid.lower() in r.inner_text().lower():
                                        occurrences.append(r)
                                except Exception: pass
                            
                            if occurrences:
                                logger.info(f"[Aşama C] Arama ile {ctx.rid} ID'li şarkıdan {len(occurrences)} versiyon bulundu.")
                                song_index[ctx.rid] = occurrences
                            else:
                                self._batch_stats["failed"] += 1
                                trigger_dashboard_update()
                                logger.warning(f"[Aşama C] {ctx.rid} arama sonucunda bile bulunamadı (Üretilemedi). Atlanıyor.")
                                if progress_callback: progress_callback(ctx.rid, "Üretilemedi ❌")
                                ctx.state = SongState.FAILED
                                self.update_row_status(ctx.row_idx, dl_status="failed", status="Üretilemedi")
                        else:
                            logger.warning(f"[Phase C] Search failed for {ctx.rid} (Üretilemedi). Skipping.")
                            ctx.state = SongState.FAILED
                            if progress_callback: progress_callback(ctx.rid, "Üretilemedi ❌")
                            self.update_row_status(ctx.row_idx, dl_status="failed", status="Üretilemedi")
                        
                        # Clear search after each song
                        try:
                            search_selectors = [
                                "input[aria-label='Search clips']",
                                "input[placeholder='Search']",
                                "input[aria-label='Search']",
                                "input[type='search']",
                                ".search-input"
                            ]
                            
                            search_input = None
                            for selector in search_selectors:
                                loc = self.tab.locator(selector).first
                                if loc.is_visible(timeout=500):
                                    search_input = loc
                                    break
                                    
                            if search_input:
                                search_input.fill("")
                                self.tab.keyboard.press("Escape")
                                time.sleep(1)
                                self.tab.keyboard.press("Escape")
                        except Exception: pass

                # ===== PHASE D: FINAL DOWNLOAD LOOP =====
                # Step 2: Download from index (no more DOM scanning needed!)
                for ctx in valid_ctxs:
                    if self.stop_requested: break
                    if ctx.rid not in song_index:
                        continue  # Was already marked as failed in Phase C
                    
                    # SESSION CHECK
                    try: self.tab.url
                    except Exception:
                        logger.warning(f"Tab crashed before downloading {ctx.rid}. Recovering...")
                        if self.persona_link: self._setup_persona_workflow()
                        else: self.browser.goto(self.base_url, page=self.tab)
                        time.sleep(5)
                        self._ensure_v5_active()
                    
                    suno_title = f"{ctx.rid}_{ctx.title}"
                    occurrences = song_index[ctx.rid]
                    
                    logger.info(f"[Aşama D] {ctx.rid} ID'li şarkı indiriliyor ({len(occurrences)} versiyon bulundu)...")
                    if progress_callback: progress_callback(ctx.rid, f"İndiriliyor... ⬇️")
                    ctx.state = SongState.DOWNLOADING
                    
                    s1, s2 = False, False
                    
                    # Playwright is strictly single-threaded per page/context. 
                    # Using ThreadPoolExecutor causes 'greenlet' thread-switch crashes.
                    # Execute sequentially instead.
                    try:
                        s1 = self._download_from_row(occurrences[0], suno_title, ctx.rid, "1")
                        if len(occurrences) > 1:
                            s2 = self._download_from_row(occurrences[1], suno_title, ctx.rid, "2")
                    except Exception as e:
                        logger.error(f"Sequential download failure for {ctx.rid}: {e}")

                    if s1 or s2:
                        self._batch_stats["success"] += 1
                        trigger_dashboard_update()
                        self.update_row_status(ctx.row_idx, status="Generated", dl_status="success", dl_attempts=0, music_status="İndirildi")
                        status_txt = f"{'1&2' if (s1 and s2) else ('1' if s1 else '2')} İndirildi! ✅"
                        if progress_callback: progress_callback(ctx.rid, status_txt)
                        completed_ids.append(ctx.rid)
                        ctx.state = SongState.VERIFIED if self.config.verify_downloads else SongState.SAVED
                        logger.info(f"✅ İndirme BAŞARILI ({ctx.rid}): s1={s1}, s2={s2}")
                    else:
                        self._batch_stats["failed"] += 1
                        trigger_dashboard_update()
                        self.update_row_status(ctx.row_idx, dl_status="failed", dl_attempts=1)
                        if progress_callback: progress_callback(ctx.rid, "İndirme Hatası! ❌")
                        ctx.state = SongState.FAILED
                        logger.warning(f"❌ İndirme BAŞARISIZ ({ctx.rid}): s1={s1}, s2={s2}")
                
                return len(completed_ids)

            return 0

        except Exception as e:
            logger.error(f"Batch Run Error: {e}")
            raise e
        finally:
            logger.info("Batch Run Finished.")

    def _generate_single_no_wait(self, row, progress_callback=None):
        """
        Fills the form and clicks Create, then verifies 'Generating' state appeared.
        Does NOT wait for completion.
        """
        prompt = row.get("prompt", "")
        lyrics = row.get("lyrics", "")
        style = row.get("suno_style", "")
        if not str(style).strip():
            style = row.get("style", "")
        
        rid = str(row.get("id", "song"))
        title = row.get("title", "Song")
        suno_title = f"{rid}_{title}"
        
        content = lyrics if str(lyrics).strip() else prompt
        
        try:
            # Custom Mode
            if not self.tab.locator("text='Lyrics'").first.is_visible():
                custom_btn = self.tab.locator("button:has-text('Custom')").first
                if custom_btn.is_visible():
                    custom_btn.click()
                    time.sleep(3)
            
            initial_count = self.tab.locator("div.clip-row").count()

            # Fills
            textareas = [t for t in self.tab.locator("textarea").all() if t.is_visible()]
            if not textareas: return False

            # Title Logic
            input_title = None
            title_loc = self.tab.locator("input[placeholder*='title' i]")
            for i in range(title_loc.count()):
                if title_loc.nth(i).is_visible():
                    input_title = title_loc.nth(i); break
            
            if not input_title:
                title_loc_alt = self.tab.locator("input[aria-label*='title' i]")
                for i in range(title_loc_alt.count()):
                    if title_loc_alt.nth(i).is_visible():
                        input_title = title_loc_alt.nth(i); break

            def fill_field(el, val):
                if self.turbo:
                    el.fill(str(val))
                else:
                    try:
                        self.browser.humanizer.type_text(self.tab, el, val)
                    except Exception:
                        el.fill(str(val))

            fill_field(textareas[0], content)
            
            style_textarea = self.tab.locator("textarea[placeholder*='style' i]").first
            if style_textarea.is_visible():
                fill_field(style_textarea, style)
            else:
                 if len(textareas) > 1: fill_field(textareas[1], style)
            
            if input_title:
                try: fill_field(input_title, suno_title)
                except Exception: pass
                
            # Adv Options
            self._setup_lyrics_mode()
            self._setup_advanced_options()

            # Click Create
            create_btn = self.tab.locator("button:has-text('Create')").filter(has_not_text="Custom").last
            if not create_btn.is_visible():
                create_btn = self.tab.get_by_role("button", name="Create").last

            if create_btn.is_enabled():
                create_btn.click()
                time.sleep(1 if self.turbo else 2) 
                # Update status in Excel
                try: self.update_row_status(rid, status="Sıraya Alındı")
                except Exception: pass

                # Wait for clip count increase OR "Generating"
                for i in range(15):
                    self._check_stop()
                    if self._detect_captcha():
                        logger.warning("CAPTCHA DETECTED DURING BATCH!")
                        if progress_callback: progress_callback(rid, "⚠️ CAPTCHA! Çözün... 🕒")
                        self._play_alert()
                        while self._detect_captcha():
                            self._check_stop()
                            time.sleep(3)
                    
                    current_count = self.tab.locator("div.clip-row").count()
                    if current_count > initial_count:
                        logger.info(f"Batch: Queued {rid}")
                        return True
                    
                    time.sleep(2)
                
                return False
            return False
        except UserStoppedException:
            logger.warning(f"Generation aborted by user during {rid}.")
            return False
        except Exception as e:
            logger.error(f"Generate Batch Error {rid}: {e}")
            return False


    def process_row(self, row, progress_callback=None):
        prompt = row.get("prompt", "")
        lyrics = row.get("lyrics", "")
        style = row.get("suno_style", "")
        if not str(style).strip():
            style = row.get("style", "")
        
        # [Requirement 7] Title Format: ID_Title
        rid = str(row.get("id", "song"))
        title = row.get("title", "Song")
        suno_title = f"{rid}_{title}"
        
        content = lyrics if str(lyrics).strip() else prompt
        
        try:
            # --- Ensure Custom Mode ---
            if not self.tab.locator("text='Lyrics'").first.is_visible():
                logger.info("Custom mode not detected, clicking 'Custom'...")
                custom_btn = self.tab.locator("button:has-text('Custom')").first
                if custom_btn.is_visible():
                    custom_btn.click()
                    time.sleep(3)
            
            # Record current clip count to prevent stale downloads
            initial_count = self.tab.locator("div.clip-row").count()
            logger.info(f"Current clip count before creation: {initial_count}")

            # Ensure we are indeed in custom mode
            textareas = [t for t in self.tab.locator("textarea").all() if t.is_visible()]
            if not textareas: 
                logger.error("No textareas found on Suno.")
                return False

            # Fills
            input_title = None
            title_loc = self.tab.locator("input[placeholder*='title' i]")
            for i in range(title_loc.count()):
                if title_loc.nth(i).is_visible():
                    input_title = title_loc.nth(i)
                    break
            
            if not input_title:
                title_loc_alt = self.tab.locator("input[aria-label*='title' i]")
                for i in range(title_loc_alt.count()):
                    if title_loc_alt.nth(i).is_visible():
                        input_title = title_loc_alt.nth(i)
                        break

            def fill_field(el, val):
                if not el or val is None: return
                try:
                    # Robust JS-based clear to ensure reactive state is reset
                    try:
                        el.scroll_into_view_if_needed()
                        self.tab.evaluate(r"""(el) => {
                            const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value')?.set || 
                                               Object.getOwnPropertyDescriptor(window.HTMLTextAreaElement.prototype, 'value')?.set;
                            if (nativeSetter) {
                                nativeSetter.call(el, '');
                            } else {
                                el.value = '';
                            }
                            el.dispatchEvent(new Event('input', { bubbles: true }));
                            el.dispatchEvent(new Event('change', { bubbles: true }));
                        }""", el)
                        time.sleep(0.2)
                    except Exception: pass

                    logger.info(f"Filling field with: {str(val)[:20]}...")
                    self.browser.humanizer.type_text(self.tab, el, val)
                except Exception as fe:
                    logger.warning(f"Humanizer failed: {fe}")
                    try:
                        el.scroll_into_view_if_needed()
                        el.fill(str(val))
                    except Exception: pass

            fill_field(textareas[0], content)
            
            # Style: try placeholder match first, then section-based lookup
            style_textarea = self.tab.locator("textarea[placeholder*='style' i]").first
            if style_textarea.is_visible():
                fill_field(style_textarea, style)
            else:
                style_found = self.tab.evaluate(r"""() => {
                    const headers = document.querySelectorAll('div[role="button"]');
                    for (const h of headers) {
                        if (h.textContent.trim() === 'Styles') {
                            let container = h.parentElement;
                            for (let d = 0; d < 5 && container; d++) {
                                const ta = container.querySelector('textarea');
                                if (ta && ta.offsetParent !== null) {
                                    ta.scrollIntoView({ block: 'center' });
                                    return true;
                                }
                                container = container.parentElement;
                            }
                        }
                    }
                    return false;
                }""")
                if style_found and len(textareas) > 1:
                    fill_field(textareas[1], style)
                elif len(textareas) > 1:
                    fill_field(textareas[1], style)
            
            # Title fill
            if input_title:
                try:
                    fill_field(input_title, suno_title)
                except Exception:
                    pass
                actual = input_title.input_value() if input_title.is_visible() else ""
                if str(suno_title) not in actual:
                    self.tab.evaluate("""(t) => {
                        const inputs = document.querySelectorAll('input[placeholder*="Song Title" i]');
                        for (const inp of inputs) {
                            if (inp.offsetParent !== null) {
                                inp.scrollIntoView({ block: 'center' });
                                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                                nativeSetter.call(inp, t);
                                inp.dispatchEvent(new Event('input', { bubbles: true }));
                                inp.dispatchEvent(new Event('change', { bubbles: true }));
                                return;
                            }
                        }
                    }""", str(suno_title))
                    time.sleep(1)

            # --- Persona & Advanced Options Setup ---
            self._setup_lyrics_mode()
            self._setup_advanced_options()

            # Click Create
            create_btn = self.tab.locator("button:has-text('Create')").filter(has_not_text="Custom").last
            if not create_btn.is_visible():
                create_btn = self.tab.get_by_role("button", name="Create").last

            if create_btn.is_enabled():
                logger.info(f"Clicking Create for: {suno_title}")
                create_btn.click()
                time.sleep(2) 
                # Update status in Excel
                try: self.update_row_status(rid, status="Sıraya Alındı")
                except Exception: pass

                # [Requirement 7 + CAPTCHA Alert]
                new_clip_appeared = False
                for i in range(25): # ~40-50s maximum wait
                    # 1. Check for success indicators
                    current_count = self.tab.locator("div.clip-row").count()
                    if current_count > initial_count:
                        new_clip_appeared = True
                        logger.info(f"New clip detected! Count: {current_count}")
                        break
                    # Check for "Generating" text in first row
                    if self.tab.locator("div.clip-row").first.locator("text='Generating'").is_visible(timeout=500):
                        new_clip_appeared = True
                        break
                    
                    # 2. Check for CAPTCHA / Challenge
                    if self._detect_captcha():
                        logger.warning("⚠️ CAPTCHA/Challenge detected on Suno!")
                        if progress_callback: progress_callback(rid, "⚠️ CAPTCHA ALGILANDI! Lütfen tarayıcıda çözün... 🕒")
                        self._play_alert()
                        # Pause until cleared
                        while self._detect_captcha() and not self.stop_requested:
                            time.sleep(3)
                        logger.info("Captcha cleared or disappeared. Resuming wait...")

                    time.sleep(2)
                
                if not new_clip_appeared:
                    if progress_callback: progress_callback(rid, "Suno Üretimi Başarısız (CAPTCHA olabilir) ❌")
                    return False

                # Trigger dual download
                logger.info(f"Triggering dual download for {rid}...")
                s1 = self._wait_and_download(suno_title, rid, progress_callback, index=0, suffix="1")
                s2 = self._wait_and_download(suno_title, rid, progress_callback, index=1, suffix="2")
                
                return s1 or s2
            
            logger.error("Create button not enabled or not found.")
            return False
        except Exception as e:
            logger.error(f"process_row failed: {e}")
            return False

    def close(self): pass


if __name__ == "__main__":
    suno = SunoGenerator(project_file="test.xlsx")
    suno.run()
