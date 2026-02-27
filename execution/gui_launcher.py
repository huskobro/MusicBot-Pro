import logging
import os
import sys
import re
import unicodedata
from concurrent.futures import ThreadPoolExecutor

# Configure logging IMMEDIATELY to catch all imports and initializations.
def get_safe_workspace():
    ws = os.path.expanduser("~/Documents/MusicBot_Workspace")
    try:
        if not os.path.exists(ws): os.makedirs(ws, exist_ok=True)
    except Exception: pass
    return ws

workspace = get_safe_workspace()
debug_log_path = os.path.join(workspace, "musicbot_debug.log")
crash_report_path = os.path.join(workspace, "crash_report.txt")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(debug_log_path, encoding='utf-8', mode='a')
    ],
    force=True
)
logger = logging.getLogger(__name__)

def _load_translations():
    """Load translations from JSON files. Falls back to minimal dict if files are missing."""
    import json as _json
    import sys as _sys
    translations = {}
    
    # In PyInstaller bundle, the entry script is at sys._MEIPASS root,
    # but the --add-data 'execution:execution' puts the json files inside 'execution' folder
    if getattr(_sys, 'frozen', False):
        _dir = os.path.join(_sys._MEIPASS, "execution")
    else:
        _dir = os.path.dirname(os.path.abspath(__file__))
        
    for lang, filename in [("English", "translations_en.json"), ("Turkish", "translations_tr.json")]:
        fpath = os.path.join(_dir, filename)
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                translations[lang] = _json.load(f)
        except Exception as e:
            logger.warning(f"Could not load {filename} from {fpath}: {e}")
            translations[lang] = {
                "title": "MusicBot Pro",
                "tab_gemini": "Gemini Lyrics",
                "tab_suno": "Suno Generator",
                "tab_visual": "Visuals & Video",
                "tab_batch": "Batch Processing",
                "tab_settings": "Settings"
            }
    return translations

TRANSLATIONS = _load_translations()

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, scrolledtext, filedialog
import time
import threading
import shutil
import openpyxl
import json
from openpyxl.styles import PatternFill
from browser_controller import BrowserController
import traceback

# PyInstaller bundle path handling
def get_bundle_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    # If not frozen, we assume we are in 'MusicBot/execution/gui_launcher.py'
    # So root is two levels up
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

bundle_dir = get_bundle_dir()

class GuiLogger(logging.Handler):
    """Custom logging handler that directs logs to a ScrolledText widget."""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.text_widget.tag_config("INFO", foreground="#333333")
        self.text_widget.tag_config("ERROR", foreground="#d9534f", font=("Consolas", 9, "bold"))
        self.text_widget.tag_config("WARNING", foreground="#f0ad4e")
        self.text_widget.tag_config("SYSTEM", foreground="#5bc0de", font=("Consolas", 9, "italic"))

    def emit(self, record):
        msg = self.format(record)
        def append():
            try:
                if self.text_widget.winfo_exists():
                    self.text_widget.configure(state='normal')
                    self.text_widget.insert(tk.END, msg + '\n', record.levelname)
                    self.text_widget.configure(state='disabled')
                    self.text_widget.yview(tk.END)
            except Exception:
                pass
        
        # Ensure thread safety by scheduling update on main loop
        try:
            if self.text_widget.winfo_exists():
                self.text_widget.after(0, append)
        except Exception:
            pass

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent, config, app_instance):
        super().__init__(parent)
        self.title(app_instance.t("settings"))
        self.geometry("950x750") # Widened to ensure all tabs are visible on all OSs
        self.config = config
        self.parent = parent
        self.app = app_instance
        
        # Initialize variables early to avoid AttributeErrors
        self.var_video_output_mode = tk.StringVar(value=config.get("video_output_mode", "profile"))
        
        # Root Configuration: Force Button Frame to Bottom using PACK (safer than grid for footers)
        # 1. Create Notebook (but don't pack yet)
        self.notebook = ttk.Notebook(self)
        
        # 2. Create and Pack Button Frame (BOTTOM) - Guaranteed Visibility
        f_btn = ttk.Frame(self, padding=10)
        f_btn.pack(side="bottom", fill="x")
        
        btn_save = ttk.Button(f_btn, text=self.app.t("save_all"), command=self.save_settings, style="Action.TButton")
        btn_save.pack(fill="x", ipady=12) 

        self.notebook.pack(side="top", fill="both", expand=True, padx=10, pady=(10, 5))

        # --- PRE-INITIALIZE ALL UI ATTRIBUTES TO PREVENT ATTRIBUTE ERRORS ---
        self.prompts_path = os.path.join(os.path.dirname(config.get("metadata_path", "")), "prompts.json")
        self.personas = []
        self.ent_preset_alias = None
        self.ent_artist_name = None
        self.ent_artist_style = None
        self.combo_preset_select = None
        self.ent_video_assets = None
        self.ent_video_custom_path = None
        self.combo_res = None
        self.spin_fps = None
        self.scale_intensity = None
        self.txt_lyrics = None
        self.txt_visual = None
        self.txt_video = None
        self.txt_art = None
        self.var_lyrics = None
        self.var_style = None
        self.var_visual = None
        self.var_video = None
        self.entry_delay = None
        self.entry_startup = None
        self.combo_lang = None
        self.combo_ui_lang = None
        self.var_log_at_start = None
        self.var_def_lyrics = None
        self.var_def_music = None
        self.var_def_art_p = None
        self.var_def_art_i = None
        self.var_def_video = None
        self.var_def_compilation = None
        self.var_persona_link_enabled = None
        self.combo_persona_select = None
        self.var_gender_enabled = None
        self.combo_gender = None
        self.scale_audio = None
        self.var_audio_enabled = None
        self.scale_weird = None
        self.var_weird_enabled = None
        self.scale_style = None
        self.var_style_enabled = None
        self.var_lyrics_mode_enabled = None
        self.combo_lyrics_mode = None
        self.var_humanizer_enabled = None
        self.var_h_gemini = None
        self.var_h_suno = None
        self.var_h_video = None
        self.combo_human_level = None
        self.scale_speed = None
        self.spin_retries = None
        self.var_adaptive = None
        self.combo_parallel_render = None
        self.combo_video_selection = None
        self.effect_vars = {}
        self.var_suno_batch = None
        self.var_batch_op = None

        # --- TAB 0: Profiles ---
        self.tab_presets = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_presets, text=self.app.t("profiles_tab"))
        
        f_pre_root = ttk.Frame(self.tab_presets, padding=15)
        f_pre_root.pack(fill="both", expand=True)

        f_presets = ttk.LabelFrame(f_pre_root, text=self.app.t("profile_mgmt"), padding=15)
        f_presets.pack(fill="x", padx=5, pady=5)

        # 1. Selection Row (Which profile are we viewing/using?)
        f_sel = ttk.Frame(f_presets)
        f_sel.pack(fill="x", pady=(0, 10))
        ttk.Label(f_sel, text=self.app.t("active_profile"), font=("Helvetica", 10, "bold")).pack(side="left")
        
        self.presets = config.get("artist_presets", {})
        self.combo_preset_select = ttk.Combobox(f_sel, values=list(self.presets.keys()), state="readonly")
        self.combo_preset_select.pack(side="left", fill="x", expand=True, padx=10)
        self.combo_preset_select.bind("<<ComboboxSelected>>", lambda e: self.load_preset())

        # 2. Management & Alias Row (Persona style)
        f_mgmt = ttk.Frame(f_presets)
        f_mgmt.pack(fill="x", pady=(0, 15))
        
        ttk.Label(f_mgmt, text=self.app.t("preset_alias")).pack(side="left")
        self.ent_preset_alias = ttk.Entry(f_mgmt)
        self.ent_preset_alias.pack(side="left", padx=5, fill="x", expand=True)
        
        # Action Buttons clustered next to the Alias entry
        btn_add = ttk.Button(f_mgmt, text="+", width=3, command=lambda: self.save_preset(silent=False))
        btn_add.pack(side="left", padx=1)
        btn_del = ttk.Button(f_mgmt, text="-", width=3, command=self.delete_preset)
        btn_del.pack(side="left", padx=1)
        btn_clear = ttk.Button(f_mgmt, text="✨", width=3, command=self.clear_preset_form)
        btn_clear.pack(side="left", padx=1)

        # 3. Details Grid (The "Old Beauty" - for artist specific fields)
        f_grid = ttk.Frame(f_presets)
        f_grid.pack(fill="x")

        # Artist Name
        ttk.Label(f_grid, text=self.app.t("artist_name_label")).grid(row=0, column=0, sticky="w", pady=5)
        self.ent_artist_name = ttk.Entry(f_grid)
        self.ent_artist_name.grid(row=0, column=1, sticky="ew", pady=5)
        self.ent_artist_name.bind("<FocusOut>", lambda e: self.auto_fill_project_path())

        # Artist Style
        ttk.Label(f_grid, text=self.app.t("artist_style_label")).grid(row=1, column=0, sticky="w", pady=5)
        self.ent_artist_style = ttk.Entry(f_grid)
        self.ent_artist_style.grid(row=1, column=1, sticky="ew", pady=5)

        # Project File (Multi-Excel)
        ttk.Label(f_grid, text=self.app.t("project_file_label")).grid(row=2, column=0, sticky="w", pady=5)
        f_proj = ttk.Frame(f_grid)
        f_proj.grid(row=2, column=1, sticky="ew", pady=5)
        self.ent_preset_project = ttk.Entry(f_proj)
        self.ent_preset_project.pack(side="left", fill="x", expand=True)
        ttk.Button(f_proj, text=self.app.t("browse"), width=8, command=self.browse_project_for_preset).pack(side="left", padx=(5, 0))
        
        # Help label for auto-fill logic
        self.lbl_path_hint = ttk.Label(f_grid, text="", font=("Helvetica", 8, "italic"), foreground="gray")
        self.lbl_path_hint.grid(row=3, column=1, sticky="w")

        f_grid.columnconfigure(1, weight=1)

        ttk.Label(f_pre_root, text=self.app.t("save_to_profile_note"), font=("Helvetica", 9, "italic"), foreground="gray", wraplength=800).pack(pady=10)
        
        # --- TAB 1: General ---
        self.tab_general = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_general, text=self.app.t("general"))
        
        # Scrollable area for General Tab
        canvas = tk.Canvas(self.tab_general, borderwidth=0, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_general, orient="vertical", command=canvas.yview)
        scroll_frame = ttk.Frame(canvas)
        
        scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # [User Request] Browser Operations Moved to Top
        f_browser = ttk.LabelFrame(scroll_frame, text=self.app.t("browser_action_label"), padding=10)
        f_browser.pack(fill="x", padx=10, pady=5)
        
        btn_grid = ttk.Frame(f_browser)
        btn_grid.pack(fill="x")
        
        ttk.Button(btn_grid, text=self.app.t("open_chrome_btn"), command=self.open_chrome, style="Action.TButton").pack(side="left", fill="x", expand=True, padx=2, pady=5)
        ttk.Button(btn_grid, text=self.app.t("reset_chrome_btn"), command=self.reset_chrome).pack(side="left", fill="x", expand=True, padx=2, pady=5)

        # 1. Default Checked Steps
        f_defaults = ttk.LabelFrame(scroll_frame, text=self.app.t("default_steps"), padding=10)
        f_defaults.pack(fill="x", padx=10, pady=5)
        
        self.var_def_lyrics = tk.BooleanVar(value=config.get("default_run_lyrics", True))
        ttk.Checkbutton(f_defaults, text="1. " + self.app.t("lyrics"), variable=self.var_def_lyrics).pack(anchor="w")
        self.var_def_music = tk.BooleanVar(value=config.get("default_run_music", True))
        ttk.Checkbutton(f_defaults, text="2. " + self.app.t("music"), variable=self.var_def_music).pack(anchor="w")
        self.var_def_art_p = tk.BooleanVar(value=config.get("default_run_art_prompt", True))
        ttk.Checkbutton(f_defaults, text="3. " + self.app.t("art_prompt"), variable=self.var_def_art_p).pack(anchor="w")
        self.var_def_art_i = tk.BooleanVar(value=config.get("default_run_art_image", True))
        ttk.Checkbutton(f_defaults, text="4. " + self.app.t("art_image"), variable=self.var_def_art_i).pack(anchor="w")
        self.var_def_video = tk.BooleanVar(value=config.get("default_run_video", False))
        ttk.Checkbutton(f_defaults, text="5. " + self.app.t("video"), variable=self.var_def_video).pack(anchor="w")
        
        self.var_def_compilation = tk.BooleanVar(value=config.get("default_run_compilation", True))
        ttk.Checkbutton(f_defaults, text="6. " + self.app.t("compilation"), variable=self.var_def_compilation).pack(anchor="w")

        # --- TAB 1.2: Active Workflow ---
        self.tab_workflow = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_workflow, text=self.app.t("active_run"))
        
        f_active = ttk.LabelFrame(self.tab_workflow, text=self.app.t("active_run"), padding=15)
        f_active.pack(fill="both", expand=True, padx=10, pady=10)
        
        ttk.Checkbutton(f_active, text="1. " + self.app.t("lyrics"), variable=self.app.var_run_lyrics).pack(anchor="w", pady=5)
        ttk.Checkbutton(f_active, text="2. " + self.app.t("music"), variable=self.app.var_run_music).pack(anchor="w", pady=5)
        ttk.Checkbutton(f_active, text="3. " + self.app.t("art_prompt"), variable=self.app.var_run_art_prompt).pack(anchor="w", pady=5)
        ttk.Checkbutton(f_active, text="4. " + self.app.t("art_image"), variable=self.app.var_run_art_image).pack(anchor="w", pady=5)
        ttk.Checkbutton(f_active, text="5. " + self.app.t("video"), variable=self.app.var_run_video).pack(anchor="w", pady=5)
        
        ttk.Checkbutton(f_active, text="6. " + self.app.t("compilation"), variable=self.app.var_run_compilation).pack(anchor="w", pady=5)
        
        ttk.Label(f_active, text=self.app.t("note_changes"), font=("Helvetica", 9, "italic"), foreground="gray").pack(pady=20)

        # 2. Gemini Generation Logic
        f_gemini = ttk.LabelFrame(scroll_frame, text=self.app.t("gemini_logic"), padding=10)
        f_gemini.pack(fill="x", padx=10, pady=5)
        
        self.var_lyrics = tk.BooleanVar(value=config.get("gemini_lyrics", True))
        ttk.Checkbutton(f_gemini, text=self.app.t("gen_lyrics_vs_title"), variable=self.var_lyrics).pack(anchor="w")
        self.var_style = tk.BooleanVar(value=config.get("gemini_style", True))
        ttk.Checkbutton(f_gemini, text=self.app.t("gen_music_style"), variable=self.var_style).pack(anchor="w")
        self.var_visual = tk.BooleanVar(value=config.get("gemini_visual", True))
        ttk.Checkbutton(f_gemini, text=self.app.t("gen_visual_prompts"), variable=self.var_visual).pack(anchor="w")
        self.var_video = tk.BooleanVar(value=config.get("gemini_video", False))
        ttk.Checkbutton(f_gemini, text=self.app.t("gen_video_prompts"), variable=self.var_video).pack(anchor="w")
        
        # 3. Automation Delays
        f_suno = ttk.LabelFrame(scroll_frame, text=self.app.t("automation_delays"), padding=10)
        f_suno.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(f_suno, text=self.app.t("suno_delay_label")).pack(anchor="w")
        self.entry_delay = tk.Spinbox(f_suno, from_=1, to_=120, width=10)
        self.entry_delay.delete(0, tk.END)
        self.entry_delay.insert(0, str(config.get("suno_delay", 15)))
        self.entry_delay.pack(fill="x", pady=2)
        
        ttk.Label(f_suno, text=self.app.t("startup_delay_label")).pack(anchor="w")
        self.entry_startup = tk.Spinbox(f_suno, from_=0, to_=120, width=10)
        self.entry_startup.delete(0, tk.END)
        self.entry_startup.insert(0, str(config.get("startup_delay", 5)))
        self.entry_startup.pack(fill="x", pady=2)
        
        # 4. Language
        f_lang = ttk.LabelFrame(scroll_frame, text=self.app.t("lang_reg"), padding=10)
        f_lang.pack(fill="x", padx=10, pady=5)
        ttk.Label(f_lang, text=self.app.t("target_lang_label")).pack(anchor="w")
        self.combo_lang = ttk.Combobox(f_lang, values=[
            "Turkish", "English", "German", "French", "Spanish", 
            "Italian", "Portuguese", "Thai", "Hindi", "Mexican Spanish",
            "Arabic", "Afro", "Japanese", "Balkan", "UK", "Indian",
            "Caribbean", "Nordic", "African"
        ], state="readonly")
        self.combo_lang.set(config.get("target_language", "Turkish"))
        self.combo_lang.pack(fill="x", pady=2)

        # UI Language
        f_ui_lang = ttk.LabelFrame(scroll_frame, text=self.app.t("ui_language"), padding=10)
        f_ui_lang.pack(fill="x", padx=10, pady=5)
        
        self.combo_ui_lang = ttk.Combobox(f_ui_lang, values=["Turkish", "English"], state="readonly")
        self.combo_ui_lang.set(config.get("ui_language", "Turkish"))
        self.combo_ui_lang.pack(fill="x", pady=2)

        # Activity Log Startup State
        f_startup_opts = ttk.LabelFrame(scroll_frame, text=self.app.t("startup_opts_label"), padding=10)
        f_startup_opts.pack(fill="x", padx=10, pady=5)
        self.var_log_at_start = tk.BooleanVar(value=config.get("log_open_at_start", False))
        ttk.Checkbutton(f_startup_opts, text=self.app.t("log_startup"), variable=self.var_log_at_start).pack(anchor="w")


        # --- TAB 1.5: Suno Adv ---
        self.tab_adv_suno = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_adv_suno, text=self.app.t("suno_adv_tab"))
        
        f_adv_suno = ttk.LabelFrame(self.tab_adv_suno, text=self.app.t("suno_adv_params"), padding=10)
        f_adv_suno.pack(fill="both", expand=True, padx=10, pady=10)


        # Persona Profile Section
        self.var_persona_link_enabled = tk.BooleanVar(value=config.get("suno_persona_link_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_persona_label"), variable=self.var_persona_link_enabled).grid(row=0, column=0, sticky="w", pady=2)
        
        # Frame for Persona Manager
        f_persona_mgr = ttk.Frame(f_adv_suno)
        f_persona_mgr.grid(row=0, column=1, sticky="ew", pady=2)
        
        # Load Personas
        self.personas = config.get("suno_personas", {}) # Dict: {Alias: Link}
        self.active_persona_alias = config.get("suno_active_persona", "")
        
        # Combobox for Selection
        self.combo_persona_select = ttk.Combobox(f_persona_mgr, values=list(self.personas.keys()), state="readonly")
        self.combo_persona_select.pack(side="top", fill="x", pady=(0, 2))
        if self.active_persona_alias in self.personas:
            self.combo_persona_select.set(self.active_persona_alias)
        
        # Profile Management Area
        f_pm_controls = ttk.Frame(f_persona_mgr)
        f_pm_controls.pack(side="top", fill="x")
        
        ttk.Label(f_pm_controls, text=self.app.t("alias_label")).pack(side="left")
        self.ent_pm_alias = ttk.Entry(f_pm_controls, width=10)
        self.ent_pm_alias.pack(side="left", padx=2)
        
        ttk.Label(f_pm_controls, text=self.app.t("link_label")).pack(side="left")
        self.ent_pm_link = ttk.Entry(f_pm_controls, width=30) # Increased width
        self.ent_pm_link.pack(side="left", padx=2, fill="x", expand=True)
        
        ttk.Button(f_pm_controls, text="+", width=3, command=self.add_persona).pack(side="left", padx=2)
        ttk.Button(f_pm_controls, text="-", width=3, command=self.delete_persona).pack(side="left", padx=2)
        ttk.Button(f_pm_controls, text="✎", width=3, command=self.edit_persona).pack(side="left", padx=2)
        ttk.Button(f_pm_controls, text="🧪", width=3, command=self.test_persona_link).pack(side="left", padx=2)

        # 2. Vocal Gender
        self.var_gender_enabled = tk.BooleanVar(value=config.get("vocal_gender_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_vocal_gender_label"), variable=self.var_gender_enabled).grid(row=1, column=0, sticky="w", pady=2)
        self.combo_gender = ttk.Combobox(f_adv_suno, values=[self.app.t("vocal_default"), self.app.t("vocal_none"), self.app.t("vocal_male"), self.app.t("vocal_female")], state="readonly")
        self.combo_gender.set(config.get("vocal_gender", self.app.t("vocal_default")))
        self.combo_gender.grid(row=1, column=1, sticky="ew", pady=2)

        # 3. Audio Influence (%)
        self.var_audio_enabled = tk.BooleanVar(value=config.get("audio_influence_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_audio_influence_label"), variable=self.var_audio_enabled).grid(row=2, column=0, sticky="w", pady=2)
        self.scale_audio = tk.Scale(f_adv_suno, from_=10, to_=90, orient="horizontal")
        self.scale_audio.set(config.get("audio_influence", 25))
        self.scale_audio.grid(row=2, column=1, sticky="ew", pady=2)

        # 4. Weirdness
        self.var_weird_enabled = tk.BooleanVar(value=config.get("weirdness_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_weirdness_label"), variable=self.var_weird_enabled).grid(row=3, column=0, sticky="w", pady=2)
        self.scale_weird = tk.Scale(f_adv_suno, from_=1, to_=100, orient="horizontal")
        self.scale_weird.set(50 if config.get("weirdness") == "Default" else int(config.get("weirdness", 50)))
        self.scale_weird.grid(row=3, column=1, sticky="ew", pady=2)

        # 5. Style Influence
        self.var_style_enabled = tk.BooleanVar(value=config.get("style_influence_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_style_influence_label"), variable=self.var_style_enabled).grid(row=4, column=0, sticky="w", pady=2)
        self.scale_style = tk.Scale(f_adv_suno, from_=1, to_=100, orient="horizontal")
        self.scale_style.set(50 if config.get("style_influence") == "Default" else int(config.get("style_influence", 50)))
        self.scale_style.grid(row=4, column=1, sticky="ew", pady=2)

        # 6. Lyrics Mode
        self.var_lyrics_mode_enabled = tk.BooleanVar(value=config.get("lyrics_mode_enabled", False))
        ttk.Checkbutton(f_adv_suno, text=self.app.t("enable_lyrics_mode_label"), variable=self.var_lyrics_mode_enabled).grid(row=5, column=0, sticky="w", pady=2)
        self.combo_lyrics_mode = ttk.Combobox(f_adv_suno, values=[self.app.t("vocal_default"), self.app.t("mode_manual"), self.app.t("mode_auto")], state="readonly")
        self.combo_lyrics_mode.set(config.get("lyrics_mode", self.app.t("vocal_default")))
        self.combo_lyrics_mode.grid(row=5, column=1, sticky="ew", pady=2)

        f_adv_suno.columnconfigure(1, weight=1)

        # --- TAB 2: Humanizer ---
        self.tab_humanizer = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_humanizer, text=self.app.t("humanizer"))
        
        f_human = ttk.LabelFrame(self.tab_humanizer, text=self.app.t("human_settings_label"), padding=10)
        f_human.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 1. Main Toggle
        self.var_humanizer_enabled = tk.BooleanVar(value=config.get("humanizer_enabled", True))
        ttk.Checkbutton(f_human, text=self.app.t("enable_humanizer_label"), variable=self.var_humanizer_enabled).grid(row=0, column=0, columnspan=2, sticky="w", pady=10)

        # 2. Phase-specific activity
        ttk.Label(f_human, text=self.app.t("activate_humanizer_label")).grid(row=1, column=0, sticky="w", pady=5)
        self.var_h_gemini = tk.BooleanVar(value=config.get("h_activate_gemini", True))
        ttk.Checkbutton(f_human, text=self.app.t("phase1_label"), variable=self.var_h_gemini).grid(row=2, column=0, sticky="w")
        self.var_h_suno = tk.BooleanVar(value=config.get("h_activate_suno", True))
        ttk.Checkbutton(f_human, text=self.app.t("phase2_label"), variable=self.var_h_suno).grid(row=2, column=1, sticky="w")
        self.var_h_video = tk.BooleanVar(value=config.get("h_activate_video", False))
        ttk.Checkbutton(f_human, text=self.app.t("phase3_label"), variable=self.var_h_video).grid(row=2, column=2, sticky="w")

        # 3. Levels and Speed
        ttk.Label(f_human, text=self.app.t("human_level_label")).grid(row=3, column=0, sticky="w", pady=5)
        self.combo_human_level = ttk.Combobox(f_human, values=[self.app.t("level_low"), self.app.t("level_medium"), self.app.t("level_high")], state="readonly")
        self.combo_human_level.set(config.get("humanizer_level", self.app.t("level_medium")))
        self.combo_human_level.grid(row=3, column=1, sticky="w", pady=5)
        
        ttk.Label(f_human, text=self.app.t("typing_speed_label")).grid(row=4, column=0, sticky="w", pady=5)
        self.scale_speed = tk.Scale(f_human, from_=0.05, to_=2.5, resolution=0.05, orient="horizontal")
        self.scale_speed.set(config.get("humanizer_speed", 1.0))
        self.scale_speed.grid(row=4, column=1, sticky="ew", pady=5)
        ttk.Label(f_human, text=self.app.t("speed_hint"), foreground="gray").grid(row=5, column=1, sticky="w")
        
        ttk.Label(f_human, text=self.app.t("max_retries_label")).grid(row=6, column=0, sticky="w", pady=5)
        self.spin_retries = tk.Spinbox(f_human, from_=0, to_=3, width=5)
        self.spin_retries.delete(0, tk.END)
        self.spin_retries.insert(0, str(config.get("humanizer_retries", 1)))
        self.spin_retries.grid(row=6, column=1, sticky="w", pady=5)
        
        self.var_adaptive = tk.BooleanVar(value=config.get("humanizer_adaptive", True))
        ttk.Checkbutton(f_human, text=self.app.t("enable_adaptive_label"), variable=self.var_adaptive).grid(row=7, column=0, columnspan=2, sticky="w", pady=5)
        
        f_human.columnconfigure(1, weight=1)

        # --- TAB 3: Video ---
        self.tab_video = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_video, text=self.app.t("video")) # Use t("video") for localization
        
        # Scrollable area for Video Tab (to match others)
        v_canvas = tk.Canvas(self.tab_video, borderwidth=0, highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(self.tab_video, orient="vertical", command=v_canvas.yview)
        v_scroll_frame = ttk.Frame(v_canvas)
        v_scroll_frame.bind("<Configure>", lambda e: v_canvas.configure(scrollregion=v_canvas.bbox("all")))
        v_canvas.create_window((0, 0), window=v_scroll_frame, anchor="nw")
        v_canvas.configure(yscrollcommand=v_scrollbar.set)
        v_canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

        # 1. Visual Effects (Multi-Select)
        f_v_effects = ttk.LabelFrame(v_scroll_frame, text=self.app.t("video_effect_label"), padding=10)
        f_v_effects.pack(fill="x", padx=10, pady=5)
        
        self.effect_vars = {}
        effects = ["Snow", "Rain", "Particles", "Glitch", "Ken Burns", "Vignette", "Audio Visualizer", "Bass Pulse"]
        saved_effects = config.get("video_effects", [config.get("video_effect", "None")])
        
        for i, eff in enumerate(effects):
            var = tk.BooleanVar(value=eff in saved_effects)
            self.effect_vars[eff] = var
            cb = ttk.Checkbutton(f_v_effects, text=self.app.t(eff.lower().replace(" ", "_")), variable=var)
            cb.grid(row=i//3, column=i%3, sticky="w", padx=10, pady=2)

        # 2. Quality & Performance Settings
        f_v_quality = ttk.LabelFrame(v_scroll_frame, text=self.app.t("video_quality_label"), padding=10)
        f_v_quality.pack(fill="x", padx=10, pady=5)
        
        # Parallel Render Count (NEW)
        ttk.Label(f_v_quality, text=self.app.t("video_parallel_label")).grid(row=0, column=0, sticky="w", pady=5)
        self.combo_parallel_render = ttk.Combobox(f_v_quality, values=["1", "2", "4", "6", "8"], state="readonly", width=5)
        self.combo_parallel_render.set(str(config.get("video_parallel_count", 1)))
        self.combo_parallel_render.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        # Video Selection Mode (NEW)
        ttk.Label(f_v_quality, text=self.app.t("video_selection_label")).grid(row=1, column=0, sticky="w", pady=5)
        self.combo_video_selection = ttk.Combobox(f_v_quality, values=[self.app.t("v_mode_1"), self.app.t("v_mode_2"), self.app.t("v_mode_both")], state="readonly")
        self.combo_video_selection.set(config.get("video_selection_mode", self.app.t("v_mode_both")))
        self.combo_video_selection.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        # FPS
        ttk.Label(f_v_quality, text=self.app.t("video_fps_label")).grid(row=2, column=0, sticky="w", pady=5)
        self.spin_fps = tk.Spinbox(f_v_quality, from_=1, to_=60, width=5)
        self.spin_fps.delete(0, tk.END)
        self.spin_fps.insert(0, str(config.get("video_fps", 30)))
        self.spin_fps.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Resolution
        ttk.Label(f_v_quality, text=self.app.t("video_res_label")).grid(row=3, column=0, sticky="w", pady=5)
        self.combo_res = ttk.Combobox(f_v_quality, values=[self.app.t("video_res_shorts"), self.app.t("video_res_hd"), self.app.t("video_res_sd")], state="readonly")
        self.combo_res.set(config.get("video_resolution", self.app.t("video_res_shorts")))
        self.combo_res.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        
        # Intensity
        ttk.Label(f_v_quality, text=self.app.t("video_intensity_label")).grid(row=4, column=0, sticky="w", pady=5)
        self.scale_intensity = tk.Scale(f_v_quality, from_=0.1, to_=2.0, resolution=0.1, orient="horizontal")
        self.scale_intensity.set(config.get("video_intensity", 1.0))
        self.scale_intensity.grid(row=4, column=1, sticky="ew", padx=5, pady=5)

        # Render Engine (NEW)
        ttk.Label(f_v_quality, text=self.app.t("video_engine_label")).grid(row=5, column=0, sticky="w", pady=5)
        self.combo_video_engine = ttk.Combobox(f_v_quality, values=["MoviePy", "FFmpeg"], state="readonly", width=10)
        self.combo_video_engine.set(config.get("video_engine", "FFmpeg"))
        self.combo_video_engine.grid(row=5, column=1, sticky="w", padx=5, pady=5)

        # 3. Assets & Output
        f_v_assets = ttk.LabelFrame(v_scroll_frame, text=self.app.t("video_assets_output_label"), padding=10)
        f_v_assets.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(f_v_assets, text=self.app.t("video_assets_label")).pack(anchor="w")
        f_path = ttk.Frame(f_v_assets)
        f_path.pack(fill="x", pady=2)
        self.ent_video_assets = ttk.Entry(f_path)
        self.ent_video_assets.insert(0, config.get("video_assets_path", ""))
        self.ent_video_assets.pack(side="left", fill="x", expand=True)
        ttk.Button(f_path, text="...", width=3, command=lambda: self.browse_folder(self.ent_video_assets)).pack(side="left", padx=2)
        ttk.Label(f_v_assets, text=self.app.t("video_assets_hint"), font=("Helvetica", 8, "italic"), foreground="gray").pack(anchor="w")

        ttk.Label(f_v_assets, text=self.app.t("video_output_mode_label")).pack(anchor="w", pady=(10,0))
        ttk.Radiobutton(f_v_assets, text=self.app.t("video_output_same_label"), variable=self.var_video_output_mode, value="profile").pack(anchor="w")
        ttk.Radiobutton(f_v_assets, text=self.app.t("video_output_custom_label"), variable=self.var_video_output_mode, value="custom").pack(anchor="w")
        
        f_custom_path = ttk.Frame(f_v_assets)
        f_custom_path.pack(fill="x", pady=2)
        self.ent_video_custom_path = ttk.Entry(f_custom_path)
        self.ent_video_custom_path.insert(0, config.get("video_custom_output_path", ""))
        self.ent_video_custom_path.pack(side="left", fill="x", expand=True)
        ttk.Button(f_custom_path, text="...", width=3, command=lambda: self.browse_folder(self.ent_video_custom_path)).pack(side="left", padx=2)

        # --- TAB 3: Prompts ---
        self.tab_prompts = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_prompts, text=self.app.t("prompts"))
        
        # Scrollable area for Prompts Tab
        p_canvas = tk.Canvas(self.tab_prompts, borderwidth=0, highlightthickness=0)
        p_scrollbar = ttk.Scrollbar(self.tab_prompts, orient="vertical", command=p_canvas.yview)
        p_scroll_frame = ttk.Frame(p_canvas)
        
        p_scroll_frame.bind(
            "<Configure>",
            lambda e: p_canvas.configure(scrollregion=p_canvas.bbox("all"))
        )
        p_canvas.create_window((0, 0), window=p_scroll_frame, anchor="nw")
        p_canvas.configure(yscrollcommand=p_scrollbar.set)
        
        p_canvas.pack(side="left", fill="both", expand=True)
        p_scrollbar.pack(side="right", fill="y")
        
        # Gemini Chat Mode
        f_gemini_mode = ttk.Frame(p_scroll_frame)
        f_gemini_mode.pack(fill="x", padx=10, pady=(5,0))
        ttk.Label(f_gemini_mode, text=self.app.t("gemini_chat_mode_label")).pack(side="left", padx=(0, 10))
        self.combo_gemini_mode = ttk.Combobox(f_gemini_mode, values=[self.app.t("gemini_mode_new"), self.app.t("gemini_mode_temp")], state="readonly", width=20)
        self.combo_gemini_mode.set(config.get("gemini_chat_mode", self.app.t("gemini_mode_new")))
        self.combo_gemini_mode.pack(side="left")

        # Lyrics
        ttk.Label(p_scroll_frame, text=self.app.t("lyrics_master_label"), font=("Helvetica", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
        self.txt_lyrics = scrolledtext.ScrolledText(p_scroll_frame, height=8, wrap=tk.WORD, font=("Consolas", 10))
        self.txt_lyrics.pack(fill="x", padx=10, pady=5)
        
        # Visual
        ttk.Label(p_scroll_frame, text=self.app.t("visual_master_label"), font=("Helvetica", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
        self.txt_visual = scrolledtext.ScrolledText(p_scroll_frame, height=8, wrap=tk.WORD, font=("Consolas", 10))
        self.txt_visual.pack(fill="x", padx=10, pady=5)
        
        # Video
        ttk.Label(p_scroll_frame, text=self.app.t("video_master_label"), font=("Helvetica", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
        self.txt_video = scrolledtext.ScrolledText(p_scroll_frame, height=8, wrap=tk.WORD, font=("Consolas", 10))
        self.txt_video.pack(fill="x", padx=10, pady=5)
        
        # Art (Thumbnail)
        ttk.Label(p_scroll_frame, text=self.app.t("art_master_label"), font=("Helvetica", 10, "bold")).pack(anchor="w", padx=10, pady=(10,0))
        self.txt_art = scrolledtext.ScrolledText(p_scroll_frame, height=8, wrap=tk.WORD, font=("Consolas", 10))
        self.txt_art.pack(fill="x", padx=10, pady=5)
        
        self.load_prompts_data()

        # --- TAB: Activity Log (MOVED TO END) ---
        self.tab_logs = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_logs, text=self.app.t("log"))
        
        self.log_disp = scrolledtext.ScrolledText(self.tab_logs, state='disabled', font=("Consolas", 9), bg="#ffffff", fg="#1e1e1e")
        self.log_disp.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Live Logger handler for Settings window
        self.settings_handler = GuiLogger(self.log_disp)
        formatter = logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S')
        self.settings_handler.setFormatter(formatter)
        logging.getLogger().addHandler(self.settings_handler)
        
        # Populate with current logs from main app
        if hasattr(self.app, "log_text"):
            self.log_disp.configure(state='normal')
            logs = self.app.log_text.get("1.0", tk.END).strip()
            self.log_disp.insert("1.0", logs + "\n--- Settings Window Log Start ---\n")
            self.log_disp.configure(state='disabled')
            self.log_disp.yview(tk.END)

        # Save Button (Locked at bottom via Grid Row 1)
        # NOW PACKED AT TOP OF INIT
        
        # Initialize with current active profile (Ref 2)
        initial_p = config.get("active_preset")
        if initial_p and initial_p in self.presets:
            self.combo_preset_select.set(initial_p)
            self.load_preset()
        
        # Cleanup handler on close
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        logging.getLogger().removeHandler(self.settings_handler)
        self.destroy()

    # --- Preset Methods ---
    def browse_project_for_preset(self):
        file_path = filedialog.askopenfilename(
            title=self.app.t("load_project"),
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            self.ent_preset_project.delete(0, tk.END)
            self.ent_preset_project.insert(0, file_path)

    def auto_fill_project_path(self):
        """Automatically suggest an Excel path based on the artist name."""
        artist_name = self.ent_artist_name.get().strip()
        if not artist_name:
            return
            
        # Convert to snake_case (lina_zahara)
        import re
        base_name = re.sub(r'[^a-zA-Z0-9]', '_', artist_name.lower())
        base_name = re.sub(r'_+', '_', base_name).strip('_')
        
        if not base_name:
            return
            
        filename = f"{base_name}.xlsx"
        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        project_path = os.path.join(workspace, filename)
        
        # Show hint regardless
        if hasattr(self, 'lbl_path_hint'):
            self.lbl_path_hint.config(text=f"{self.app.t('suggested_label')} {filename}")
            
        # Only auto-fill if the box is currently empty or contains a default-looking path
        current_path = self.ent_preset_project.get().strip()
        if not current_path or workspace in current_path:
            self.ent_preset_project.delete(0, tk.END)
            self.ent_preset_project.insert(0, project_path)
            
            # Warn if file doesn't exist
            if not os.path.exists(project_path):
                msg = self.app.t("msg_project_not_found").format(path=project_path)
                messagebox.showwarning(self.app.t("warning"), msg)

    def save_preset(self, silent=False):
        alias = self.ent_preset_alias.get().strip()
        if not alias:
            # If silent, we don't warn, we just skip saving preset
            if not silent:
                messagebox.showwarning(self.app.t("warning"), self.app.t("msg_enter_preset_alias"))
            return
        
        # Capture current prompts
        prompts = {
            "lyrics_master_prompt": self.txt_lyrics.get("1.0", tk.END).strip() if self.txt_lyrics else "",
            "visual_master_prompt": self.txt_visual.get("1.0", tk.END).strip() if self.txt_visual else "",
            "video_master_prompt": self.txt_video.get("1.0", tk.END).strip() if self.txt_video else "",
            "art_master_prompt": self.txt_art.get("1.0", tk.END).strip() if self.txt_art else ""
        }
        
        # Capture all configurable settings
        settings_snapshot = {}
        
        if self.var_lyrics: settings_snapshot["gemini_lyrics"] = self.var_lyrics.get()
        if self.var_style: settings_snapshot["gemini_style"] = self.var_style.get()
        if self.var_visual: settings_snapshot["gemini_visual"] = self.var_visual.get()
        if self.var_video: settings_snapshot["gemini_video"] = self.var_video.get()
        
        try:
            if self.entry_delay: settings_snapshot["suno_delay"] = int(self.entry_delay.get())
        except (ValueError, TypeError): settings_snapshot["suno_delay"] = 15
        
        try:
            if self.entry_startup: settings_snapshot["startup_delay"] = int(self.entry_startup.get())
        except (ValueError, TypeError): settings_snapshot["startup_delay"] = 5
        
        if self.combo_lang: settings_snapshot["target_language"] = self.combo_lang.get()
        if self.ent_artist_name: settings_snapshot["artist_name"] = self.ent_artist_name.get()
        if self.ent_artist_style: settings_snapshot["artist_style"] = self.ent_artist_style.get()
        if self.combo_gender: settings_snapshot["vocal_gender"] = self.combo_gender.get()
        if self.var_gender_enabled: settings_snapshot["vocal_gender_enabled"] = self.var_gender_enabled.get()
        
        try:
            if self.scale_audio: settings_snapshot["audio_influence"] = int(self.scale_audio.get())
        except (ValueError, TypeError): settings_snapshot["audio_influence"] = 25
        
        if self.var_audio_enabled: settings_snapshot["audio_influence_enabled"] = self.var_audio_enabled.get()
        
        try:
            if self.scale_weird: settings_snapshot["weirdness"] = int(self.scale_weird.get())
        except (ValueError, TypeError): settings_snapshot["weirdness"] = 50
        
        if self.var_weird_enabled: settings_snapshot["weirdness_enabled"] = self.var_weird_enabled.get()
        
        try:
            if self.scale_style: settings_snapshot["style_influence"] = int(self.scale_style.get())
        except (ValueError, TypeError): settings_snapshot["style_influence"] = 50
        if self.var_style_enabled: settings_snapshot["style_influence_enabled"] = self.var_style_enabled.get()
        if self.combo_lyrics_mode: settings_snapshot["lyrics_mode"] = self.combo_lyrics_mode.get()
        if self.var_lyrics_mode_enabled: settings_snapshot["lyrics_mode_enabled"] = self.var_lyrics_mode_enabled.get()
        if self.var_persona_link_enabled: settings_snapshot["suno_persona_link_enabled"] = self.var_persona_link_enabled.get()
        if self.combo_persona_select: settings_snapshot["suno_active_persona"] = self.combo_persona_select.get()
        
        # Defaults
        if self.var_def_lyrics: settings_snapshot["default_run_lyrics"] = self.var_def_lyrics.get()
        if self.var_def_music: settings_snapshot["default_run_music"] = self.var_def_music.get()
        if self.var_def_art_p: settings_snapshot["default_run_art_prompt"] = self.var_def_art_p.get()
        if self.var_def_art_i: settings_snapshot["default_run_art_image"] = self.var_def_art_i.get()
        if self.var_def_video: settings_snapshot["default_run_video"] = self.var_def_video.get()
        if self.var_def_compilation: settings_snapshot["default_run_compilation"] = self.var_def_compilation.get()

        # Video
        if self.effect_vars: settings_snapshot["video_effects"] = [eff for eff, var in self.effect_vars.items() if var and var.get()]
        if self.spin_fps: settings_snapshot["video_fps"] = int(self.spin_fps.get())
        if self.combo_res: settings_snapshot["video_resolution"] = self.combo_res.get()
        if self.scale_intensity: settings_snapshot["video_intensity"] = self.scale_intensity.get()
        if self.ent_video_assets: settings_snapshot["video_assets_path"] = self.ent_video_assets.get()
        if self.var_video_output_mode: settings_snapshot["video_output_mode"] = self.var_video_output_mode.get()
        if self.ent_video_custom_path: settings_snapshot["video_custom_output_path"] = self.ent_video_custom_path.get()
        if self.combo_parallel_render: settings_snapshot["video_parallel_count"] = int(self.combo_parallel_render.get())
        if self.combo_video_selection: settings_snapshot["video_selection_mode"] = self.combo_video_selection.get()
        if hasattr(self, 'combo_video_engine') and self.combo_video_engine: settings_snapshot["video_engine"] = self.combo_video_engine.get()
        if hasattr(self, 'combo_gemini_mode') and self.combo_gemini_mode: settings_snapshot["gemini_chat_mode"] = self.combo_gemini_mode.get()
        if hasattr(self, 'ent_preset_project'): settings_snapshot["project_file"] = self.ent_preset_project.get().strip()
        
        # Humanizer
        if self.var_humanizer_enabled: settings_snapshot["humanizer_enabled"] = self.var_humanizer_enabled.get()
        if self.var_h_gemini: settings_snapshot["h_activate_gemini"] = self.var_h_gemini.get()
        if self.var_h_suno: settings_snapshot["h_activate_suno"] = self.var_h_suno.get()
        if self.var_h_video: settings_snapshot["h_activate_video"] = self.var_h_video.get()
        if self.combo_human_level: settings_snapshot["humanizer_level"] = self.combo_human_level.get()
        if self.scale_speed: settings_snapshot["humanizer_speed"] = self.scale_speed.get()
        if self.spin_retries: settings_snapshot["humanizer_retries"] = int(self.spin_retries.get())
        if self.var_adaptive: settings_snapshot["humanizer_adaptive"] = self.var_adaptive.get()
        
        self.presets[alias] = {
            "settings": settings_snapshot,
            "prompts": prompts,
            "project_file": settings_snapshot.get("project_file", "")
        }
        self.update_preset_combo()
        self.combo_preset_select.set(alias)
        if not silent:
            messagebox.showinfo(self.app.t("success"), self.app.t("msg_preset_saved").format(alias=alias))

    def load_preset(self, silent=True):
        alias = self.combo_preset_select.get()
        if not alias or alias not in self.presets:
            return
        
        # Mark as active immediately (User requirement)
        self.config["active_preset"] = alias
        # Trigger auto-fill/hint update early
        self.auto_fill_project_path()
        
        # === FIX: Set cmb_profile directly (no 👤 prefix, combobox holds plain name) ===
        if hasattr(self.app, 'cmb_profile'):
            self.app.cmb_profile['values'] = list(self.presets.keys())  # Ensure list is up to date
            self.app.cmb_profile.set(alias)
        if hasattr(self.app, 'profile_var'):
            self.app.profile_var.set(alias)
        
        data = self.presets[alias]
        settings = data.get("settings", {})
        prompts = data.get("prompts", {})
        
        # Apply settings to UI
        if self.var_lyrics: self.var_lyrics.set(settings.get("gemini_lyrics", True))
        if self.var_style: self.var_style.set(settings.get("gemini_style", True))
        if self.var_visual: self.var_visual.set(settings.get("gemini_visual", True))
        if self.var_video: self.var_video.set(settings.get("gemini_video", False))
        
        if self.entry_delay:
            self.entry_delay.delete(0, tk.END)
            self.entry_delay.insert(0, settings.get("suno_delay", "15"))
        if self.entry_startup:
            self.entry_startup.delete(0, tk.END)
            self.entry_startup.insert(0, settings.get("startup_delay", "5"))
        if self.combo_lang: self.combo_lang.set(settings.get("target_language", "Turkish"))
        
        if self.ent_artist_name:
            self.ent_artist_name.delete(0, tk.END)
            self.ent_artist_name.insert(0, settings.get("artist_name", ""))
        if self.ent_artist_style:
            self.ent_artist_style.delete(0, tk.END)
            self.ent_artist_style.insert(0, settings.get("artist_style", ""))
        
        if self.combo_gender: self.combo_gender.set(settings.get("vocal_gender", "Default"))
        if self.var_gender_enabled: self.var_gender_enabled.set(settings.get("vocal_gender_enabled", False))
        if self.scale_audio: self.scale_audio.set(int(settings.get("audio_influence", 25)))
        if self.var_audio_enabled: self.var_audio_enabled.set(settings.get("audio_influence_enabled", False))
        if self.scale_weird: self.scale_weird.set(int(settings.get("weirdness", 50)))
        if self.var_weird_enabled: self.var_weird_enabled.set(settings.get("weirdness_enabled", False))
        if self.scale_style: self.scale_style.set(int(settings.get("style_influence", 50)))
        if self.var_style_enabled: self.var_style_enabled.set(settings.get("style_influence_enabled", False))
        if self.combo_lyrics_mode: self.combo_lyrics_mode.set(settings.get("lyrics_mode", "Default"))
        if self.var_lyrics_mode_enabled: self.var_lyrics_mode_enabled.set(settings.get("lyrics_mode_enabled", False))
        
        if self.combo_persona_select: self.combo_persona_select.set(settings.get("suno_active_persona", ""))
        
        if self.var_suno_batch: self.var_suno_batch.set(settings.get("suno_batch_mode", False))
        if self.var_batch_op: self.var_batch_op.set(settings.get("suno_batch_op_mode", "full"))

        # Defaults
        if self.var_def_lyrics: self.var_def_lyrics.set(settings.get("default_run_lyrics", True))
        if self.var_def_music: self.var_def_music.set(settings.get("default_run_music", True))
        if self.var_def_art_p: self.var_def_art_p.set(settings.get("default_run_art_prompt", True))
        if self.var_def_art_i: self.var_def_art_i.set(settings.get("default_run_art_image", True))
        if self.var_def_video: self.var_def_video.set(settings.get("default_run_video", False))
        if self.var_def_compilation: self.var_def_compilation.set(settings.get("default_run_compilation", True))
        
        # Apply Video Settings to UI
        target_effects = settings.get("video_effects", [settings.get("video_effect", "None")])
        if self.effect_vars:
            for eff, var in self.effect_vars.items():
                if var: var.set(eff in target_effects)
            
        if self.spin_fps:
            self.spin_fps.delete(0, tk.END)
            self.spin_fps.insert(0, str(settings.get("video_fps", 30)))
        if self.combo_res: self.combo_res.set(settings.get("video_resolution", self.app.t("video_res_shorts")))
        if self.scale_intensity: self.scale_intensity.set(float(settings.get("video_intensity", 1.0)))
        if self.ent_video_assets:
            self.ent_video_assets.delete(0, tk.END)
            self.ent_video_assets.insert(0, settings.get("video_assets_path", ""))
        if self.var_video_output_mode: self.var_video_output_mode.set(settings.get("video_output_mode", "profile"))
        if self.ent_video_custom_path:
            self.ent_video_custom_path.delete(0, tk.END)
            self.ent_video_custom_path.insert(0, settings.get("video_custom_output_path", ""))
        
        # Project File
        if hasattr(self, 'ent_preset_project') and self.ent_preset_project:
            self.ent_preset_project.delete(0, tk.END)
            self.ent_preset_project.insert(0, data.get("project_file", ""))
        
        if self.combo_parallel_render: self.combo_parallel_render.set(str(settings.get("video_parallel_count", 1)))
        if self.combo_video_selection: self.combo_video_selection.set(settings.get("video_selection_mode", self.app.t("v_mode_both")))
        if hasattr(self, 'combo_video_engine') and self.combo_video_engine: self.combo_video_engine.set(settings.get("video_engine", "FFmpeg"))
        if hasattr(self, 'combo_gemini_mode') and self.combo_gemini_mode: self.combo_gemini_mode.set(settings.get("gemini_chat_mode", self.app.t("gemini_mode_new")))
        
        # Humanizer
        if self.var_humanizer_enabled: self.var_humanizer_enabled.set(settings.get("humanizer_enabled", True))
        if self.var_h_gemini: self.var_h_gemini.set(settings.get("h_activate_gemini", True))
        if self.var_h_suno: self.var_h_suno.set(settings.get("h_activate_suno", True))
        if self.var_h_video: self.var_h_video.set(settings.get("h_activate_video", False))
        if self.combo_human_level: self.combo_human_level.set(settings.get("humanizer_level", self.app.t("level_medium")))
        if self.scale_speed: self.scale_speed.set(float(settings.get("humanizer_speed", 1.0)))
        if self.spin_retries:
            self.spin_retries.delete(0, tk.END)
            self.spin_retries.insert(0, str(settings.get("humanizer_retries", 1)))
        if self.var_adaptive: self.var_adaptive.set(settings.get("humanizer_adaptive", True))
        
        # Multi-Excel Support (Restore path and trigger load)
        p_file = settings.get("project_file")
        if p_file:
            if hasattr(self, 'ent_preset_project'):
                self.ent_preset_project.delete(0, tk.END)
                self.ent_preset_project.insert(0, p_file)
            
            # Switch the main application active project if it exists
            if os.path.exists(p_file):
                self.app.load_project_data(p_file)
            else:
                logger.warning(f"Profile '{alias}' project file not found: {p_file}")
        
        # Apply prompts to UI
        if self.txt_lyrics:
            self.txt_lyrics.delete("1.0", tk.END)
            self.txt_lyrics.insert("1.0", prompts.get("lyrics_master_prompt", ""))
        if self.txt_visual:
            self.txt_visual.delete("1.0", tk.END)
            self.txt_visual.insert("1.0", prompts.get("visual_master_prompt", ""))
        if self.txt_video:
            self.txt_video.delete("1.0", tk.END)
            self.txt_video.insert("1.0", prompts.get("video_master_prompt", ""))
        if self.txt_art:
            self.txt_art.delete("1.0", tk.END)
            self.txt_art.insert("1.0", prompts.get("art_master_prompt", ""))
        
        # Apply Alias to Entry (Crucial for user feedback)
        self.ent_preset_alias.delete(0, tk.END)
        self.ent_preset_alias.insert(0, str(alias))
        
        # Select in combobox to stay in sync if called externally
        if self.combo_preset_select.get() != alias:
            self.combo_preset_select.set(alias)
            
        # Refresh UI state
        self.update_idletasks()

    def clear_preset_form(self):
        """Clears the form to allow adding a new profile easily."""
        self.combo_preset_select.set('')
        self.ent_preset_alias.delete(0, tk.END)
        self.ent_artist_name.delete(0, tk.END)
        self.ent_artist_style.delete(0, tk.END)
        # We don't clear master prompts to allow users to use current ones as base
        messagebox.showinfo(self.app.t("info"), self.app.t("msg_enter_preset_alias"))

    def delete_preset(self):
        alias = self.combo_preset_select.get()
        if alias and alias in self.presets:
            if messagebox.askyesno(self.app.t("confirm"), self.app.t("msg_confirm_delete_preset").format(alias=alias)):
                del self.presets[alias]
                self.update_preset_combo()
                self.combo_preset_select.set('')
                self.ent_preset_alias.delete(0, tk.END)

    def update_preset_combo(self):
        preset_keys = list(self.presets.keys())
        self.combo_preset_select['values'] = preset_keys
        # === FIX: Also keep the main dashboard profile combobox in sync ===
        if hasattr(self.app, 'cmb_profile'):
            self.app.cmb_profile['values'] = preset_keys

    def load_prompts_data(self):
        import json
        if os.path.exists(self.prompts_path):
            try:
                with open(self.prompts_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.txt_lyrics.insert("1.0", data.get("lyrics_master_prompt", ""))
                    self.txt_visual.insert("1.0", data.get("visual_master_prompt", ""))
                    self.txt_video.insert("1.0", data.get("video_master_prompt", ""))
                    self.txt_art.insert("1.0", data.get("art_master_prompt", ""))
            except Exception as e:
                logger.error(f"Failed to load prompts data: {e}")

    def add_persona(self):
        alias = self.ent_pm_alias.get().strip()
        link = self.ent_pm_link.get().strip()
        if alias and link:
            self.personas[alias] = link
            self.update_persona_combo()
            self.combo_persona_select.set(alias)
            self.ent_pm_alias.delete(0, tk.END)
            self.ent_pm_link.delete(0, tk.END)

    def delete_persona(self):
        selection = self.combo_persona_select.get()
        if selection and selection in self.personas:
            del self.personas[selection]
            self.update_persona_combo()
            self.combo_persona_select.set('')

    def edit_persona(self):
        selection = self.combo_persona_select.get()
        if selection and selection in self.personas:
            link = self.personas[selection]
            # Load into entry fields
            self.ent_pm_alias.delete(0, tk.END)
            self.ent_pm_alias.insert(0, selection)
            self.ent_pm_link.delete(0, tk.END)
            self.ent_pm_link.insert(0, link)
            messagebox.showinfo("Bilgi", f"'{selection}' düzenleme için yüklendi.\nDeğişiklik yapıp '+' butonuna basın.")

    def test_persona_link(self):
        """Opens the entered link in system default browser to check for 404s."""
        link = self.ent_pm_link.get().strip()
        if not link:
            # If entry empty, try selected item
            selection = self.combo_persona_select.get()
            if selection and selection in self.personas:
                link = self.personas[selection]
        
        if link:
            import webbrowser
            webbrowser.open(link)
        else:
            messagebox.showwarning(self.app.t("warning"), "Lütfen test edilecek bir link girin veya listeden seçin.")

    def update_persona_combo(self):
        self.combo_persona_select['values'] = list(self.personas.keys())

    def open_chrome(self):
        self.app.open_chrome_profile()

    def reset_chrome(self):
        if messagebox.askyesno(self.app.t("confirm"), self.app.t("msg_confirm_reset_chrome")):
            self.app.reset_chrome_profile()

    def browse_folder(self, target_entry):
        folder = filedialog.askdirectory(title=self.app.t("msg_select_folder"))
        if folder:
            target_entry.delete(0, tk.END)
            target_entry.insert(0, folder)

    def save_settings(self):
        try:
            # 0. Automatically update active profile first (Ref 7) - SILENTLY
            self.save_preset(silent=True)
            self.config["artist_presets"] = self.presets
            
            # 1. Config Object Update
            if self.var_lyrics: self.config["gemini_lyrics"] = self.var_lyrics.get()
            if self.var_style: self.config["gemini_style"] = self.var_style.get()
            if self.var_visual: self.config["gemini_visual"] = self.var_visual.get()
            if self.var_video: self.config["gemini_video"] = self.var_video.get()
            
            if self.entry_delay: self.config["suno_delay"] = int(self.entry_delay.get())
            if self.entry_startup: self.config["startup_delay"] = int(self.entry_startup.get())
            if self.combo_lang: self.config["target_language"] = self.combo_lang.get()
            if self.combo_ui_lang: self.config["ui_language"] = self.combo_ui_lang.get()
            if self.var_log_at_start: self.config["log_open_at_start"] = self.var_log_at_start.get()
            
            if self.ent_artist_name: self.config["artist_name"] = self.ent_artist_name.get()
            if self.ent_artist_style: self.config["artist_style"] = self.ent_artist_style.get()
            
            # Humanizer Configs
            if self.var_humanizer_enabled: self.config["humanizer_enabled"] = self.var_humanizer_enabled.get()
            if self.var_h_gemini: self.config["h_activate_gemini"] = self.var_h_gemini.get()
            if self.var_h_suno: self.config["h_activate_suno"] = self.var_h_suno.get()
            if self.var_h_video: self.config["h_activate_video"] = self.var_h_video.get()
            if self.combo_human_level: self.config["humanizer_level"] = self.combo_human_level.get()
            if self.scale_speed: self.config["humanizer_speed"] = self.scale_speed.get()
            if self.spin_retries: self.config["humanizer_retries"] = int(self.spin_retries.get())
            if self.var_adaptive: self.config["humanizer_adaptive"] = self.var_adaptive.get()

            # Defaults Config
            if self.var_def_lyrics: self.config["default_run_lyrics"] = self.var_def_lyrics.get()
            if self.var_def_music: self.config["default_run_music"] = self.var_def_music.get()
            if self.var_def_art_p: self.config["default_run_art_prompt"] = self.var_def_art_p.get()
            if self.var_def_art_i: self.config["default_run_art_image"] = self.var_def_art_i.get()
            if self.var_def_compilation: self.config["default_run_compilation"] = self.var_def_compilation.get()
            
            # Suno Advanced
            if self.var_persona_link_enabled: self.config["suno_persona_link_enabled"] = self.var_persona_link_enabled.get()
            self.config["suno_personas"] = getattr(self, "personas", {})
            if self.combo_persona_select: self.config["suno_active_persona"] = self.combo_persona_select.get()
            
            if self.combo_gender: self.config["vocal_gender"] = self.combo_gender.get()
            if self.var_gender_enabled: self.config["vocal_gender_enabled"] = self.var_gender_enabled.get()
            if self.scale_audio: self.config["audio_influence"] = self.scale_audio.get()
            if self.var_audio_enabled: self.config["audio_influence_enabled"] = self.var_audio_enabled.get()
            if self.scale_weird: self.config["weirdness"] = self.scale_weird.get()
            if self.var_weird_enabled: self.config["weirdness_enabled"] = self.var_weird_enabled.get()
            if self.scale_style: self.config["style_influence"] = self.scale_style.get()
            if self.var_style_enabled: self.config["style_influence_enabled"] = self.var_style_enabled.get()
            if self.combo_lyrics_mode: self.config["lyrics_mode"] = self.combo_lyrics_mode.get()
            if self.var_lyrics_mode_enabled: self.config["lyrics_mode_enabled"] = self.var_lyrics_mode_enabled.get()
            
            # Video Configs
            if self.effect_vars: self.config["video_effects"] = [eff for eff, var in self.effect_vars.items() if var and var.get()]
            if self.spin_fps: self.config["video_fps"] = int(self.spin_fps.get())
            if self.combo_res: self.config["video_resolution"] = self.combo_res.get()
            if self.scale_intensity: self.config["video_intensity"] = self.scale_intensity.get()
            if self.ent_video_assets: self.config["video_assets_path"] = self.ent_video_assets.get()
            if self.var_video_output_mode: self.config["video_output_mode"] = self.var_video_output_mode.get()
            if self.ent_video_custom_path: self.config["video_custom_output_path"] = self.ent_video_custom_path.get()
            if self.combo_video_selection: self.config["video_selection_mode"] = self.combo_video_selection.get()
            if self.combo_parallel_render: self.config["video_parallel_count"] = int(self.combo_parallel_render.get())
            
            # 2. Prompts Data Update
            import json
            prompt_data = {
                "lyrics_master_prompt": self.txt_lyrics.get("1.0", tk.END).strip() if self.txt_lyrics else "",
                "visual_master_prompt": self.txt_visual.get("1.0", tk.END).strip() if self.txt_visual else "",
                "video_master_prompt": self.txt_video.get("1.0", tk.END).strip() if self.txt_video else "",
                "art_master_prompt": self.txt_art.get("1.0", tk.END).strip() if self.txt_art else ""
            }
            if self.prompts_path:
                with open(self.prompts_path, "w", encoding="utf-8") as f:
                    json.dump(prompt_data, f, indent=4, ensure_ascii=False)
            
            # 3. Request App Level Save
            # Note: app_instance must have save_settings(config)
            if hasattr(self.app, "save_settings"):
                self.app.save_settings(self.config)
            
            # Warn about language change
            if self.combo_ui_lang and self.combo_ui_lang.get() != self.app.config.get("ui_language"):
                messagebox.showinfo(self.app.t("ready"), self.app.t("msg_restart_lang"))

            messagebox.showinfo(self.app.t("success"), self.app.t("msg_settings_saved"))
            self.destroy()
        except Exception as e:
            messagebox.showerror(self.app.t("error"), self.app.t("msg_failed_to_save").format(error=e))

class TabState:
    """Helper to maintain UI elements and data for a specific profile tab."""
    def __init__(self, tab_frame, tree, scrollbar, profile_name):
        self.tab_frame = tab_frame
        self.tree = tree
        self.scrollbar = scrollbar
        self.profile_name = profile_name
        self.project_file = None
        self.selected_songs = set()
        self.song_steps = {} # {rid: [L, M, AP, AI, V]}
        self.filter_var = tk.StringVar()
        self.active_only_var = tk.BooleanVar(value=False)
        self.status_filter_var = tk.StringVar(value="All")

class MusicBotGUI:
    def t(self, key):
        lang = self.config.get("ui_language", "Turkish")
        return TRANSLATIONS.get(lang, TRANSLATIONS["Turkish"]).get(key, key)

    def __init__(self, root):
        self.root = root
        
        # 1. State & Logic Flags (Initialize FIRST)
        self.stop_requested = False
        self.active_tasks = {} # {profile_name: {"status": str, "start_time": float, "total": int, "current": int}}
        self._start_status_ticker()
        self.active_browsers = {} # {profile_name: browser_instance}
        self.xlsx_lock = threading.Lock() # Global lock for Excel writes
        
        self.tabs = {} # {profile_name: TabState}
        self.current_tab = None # Active TabState
        self._is_loading = False # Thread-safe loading flag
        
        self._search_timer = None
        self._drag_data = {"item": None}
        self.log_visible = tk.BooleanVar(value=False)
        self.current_song_var = tk.StringVar(value="")
        self.selected_songs = set()
        self.song_steps = {} # Per-song phase selection: {rid: [L, M, AP, AI]}

        # 2. Configuration (Default Values)
        input_path, _, _ = self.get_data_paths()
        self.config = {
            "gemini_lyrics": True, "gemini_style": True, "gemini_visual": True, "gemini_video": False,
            "suno_delay": 15, "startup_delay": 5,
            "metadata_path": input_path,
            "default_run_art_image": True,
            "ui_language": "Turkish",
            "log_open_at_start": False,
            "audio_influence": 25,
            "vocal_gender": "Default",
            "lyrics_mode": "Default",
            "weirdness": "Default",
            "style_influence": "Default",
            "artist_name": "",
            "artist_style": "",
            "artist_presets": {},
            "active_preset": "",
            "weirdness_enabled": False,
            "style_influence_enabled": False,
            "vocal_gender_enabled": False,
            "audio_influence_enabled": False,
            "video_parallel_count": 1,
            "video_selection_mode": "Both (_1 & _2)",
            "suno_batch_mode": False,
            "suno_batch_op_mode": "full"
        }
        
        # 3. Load Saved Settings
        self.load_settings()

        # 4. Finalize Window
        self.root.title(self.t("title"))
        self.root.geometry("1200x850") # Slightly larger for more columns

        # --- THEME (Modernized) ---

        # Workflow Logic Variables
        self.var_run_lyrics = tk.BooleanVar(value=self.config.get("default_run_lyrics", True))
        self.var_run_music = tk.BooleanVar(value=self.config.get("default_run_music", True))
        self.var_run_art_prompt = tk.BooleanVar(value=self.config.get("default_run_art_prompt", True))
        self.var_run_art_image = tk.BooleanVar(value=self.config.get("default_run_art_image", True))
        self.var_run_video = tk.BooleanVar(value=self.config.get("default_run_video", False))
        self.var_run_compilation = tk.BooleanVar(value=self.config.get("default_run_compilation", True))

        # Keyboard Bindings (Platform Aware)
        mod = "Command" if sys.platform == "darwin" else "Control"
        
        self.root.bind("<Return>", lambda e: self.start_process())
        self.root.bind("<Escape>", lambda e: self.stop_process())
        self.root.bind(f"<{mod}-f>", lambda e: self.ent_filter.focus())
        self.root.bind(f"<{mod}-a>", self.handle_ctrl_a)
        self.root.bind(f"<{mod}-s>", lambda e: self.save_settings())
        self.root.bind(f"<{mod}-r>", lambda e: self.load_project_data())
        self.root.bind("<space>", self.on_space_toggle)

        # Standard Text Operations Fix (Global for Mac)
        if sys.platform == "darwin":
            # Removed redundant manual <Command-c>, <Command-v>, <Command-x> bindings
            # because Tkinter on macOS handles these natively for Entry and Text widgets,
            # and manual generation causes the "double paste" bug.
            # Cmd+A is handled by handle_ctrl_a at root level, but bind_all is safer for deep widgets
            self.root.bind_all("<Command-a>", self.handle_ctrl_a)
        else:
            # Special case for Select All in entry fields for Windows/Linux
            self.root.bind_all("<Control-a>", self.handle_ctrl_a)

        # Apply Light Theme Styles (Constructs UI)
        self.setup_styles()
        
        # --- STYLES ---
    def _is_id_match(self, filename, song_id, variant=None):
        """
        Robustly checks if a filename matches a song_id, regardless of RTL/LTI rendering order.
        Example: 1051_Title.mp3, Title_1051.mp3, 1051_2.mp3
        """
        if not filename or not song_id: return False
        
        # Normalize to NFC (Mac uses NFD, Excel/Source usually NFC)
        fn_norm = unicodedata.normalize('NFC', filename).lower()
        sid_norm = unicodedata.normalize('NFC', str(song_id)).strip().lower()
        
        # 1. Direct Match (Exact)
        name_no_ext = os.path.splitext(fn_norm)[0]
        if name_no_ext == sid_norm: return True
        
        # 2. Regex Match (Word boundary or delimiter)
        # Using specific delimiters to avoid false positives with other numbers
        pattern = rf"(^|[ _\.\-])({re.escape(sid_norm)})([ _\.\-]|$)"
        match = re.search(pattern, fn_norm)
        
        if not match: return False
        
        # 3. Variant Check (Optional)
        if variant:
            v_str = unicodedata.normalize('NFC', str(variant)).strip().lower()
            # Variant is often at the end or delimited. 
            # We match if variant is a standalone numeric part (e.g. _1, -1, 1., .1)
            v_pattern = rf"[ _\.\-](variant)?{re.escape(v_str)}($|[ _\.\-])"
            return bool(re.search(v_pattern, name_no_ext))
            
        return True

    def _get_materials_report(self, profile_name, project_file, song_ids):
        """
        Optimized $O(N+M)$ scanning to prevent UI freezes.
        Returns: {song_id: {"materials": int, "video_exists": bool, "has_r": bool, "has_m1": bool, "has_m2": bool}}
        """
        report = {str(sid): {"materials": 0, "video_exists": False, "has_r": False, "has_m1": False, "has_m2": False} for sid in song_ids}
        if not project_file or not os.path.exists(project_file): return report

        # Fix: Ensure we use the profile_name passed to this method, which is the CORRECT profile directory
        output_media = os.path.join(os.path.dirname(project_file), "output_media", profile_name)
        videos_dir = os.path.join(output_media, "videos")
        
        # 1. Map Media Files (R, M1, M2)
        if os.path.exists(output_media):
            try:
                media_files = [f.lower() for f in os.listdir(output_media) if os.path.isfile(os.path.join(output_media, f))]
                for f in media_files:
                    id_match = re.search(r'(^|[_ \.\-])(\d+)([_ \.\-]|$)', f)
                    if id_match:
                        found_id = id_match.group(2)
                        if found_id in report:
                            ext = f.split(".")[-1]
                            name_no_ext = os.path.splitext(f)[0]
                            if ext in ["png", "jpg", "jpeg"]:
                                if not report[found_id]["has_r"]:
                                    report[found_id]["has_r"] = True
                                    report[found_id]["materials"] += 1
                            elif ext in ["mp3", "wav"]:
                                if "_1" in f or "variant1" in f:
                                    if not report[found_id]["has_m1"]:
                                        report[found_id]["has_m1"] = True
                                        report[found_id]["materials"] += 1
                                elif "_2" in f or "variant2" in f:
                                    if not report[found_id]["has_m2"]:
                                        report[found_id]["has_m2"] = True
                                        report[found_id]["materials"] += 1
                                elif name_no_ext == found_id:
                                    if not report[found_id]["has_m1"]:
                                        report[found_id]["has_m1"] = True
                                        report[found_id]["materials"] += 1
            except Exception as e:
                logger.error(f"Error scanning media: {e}")

        # 2. Map Video Files
        if os.path.exists(videos_dir):
            try:
                video_files = [f.lower() for f in os.listdir(videos_dir) if f.lower().endswith(".mp4")]
                for vf in video_files:
                    id_match = re.search(r'(^|[_ \.\-])(\d+)([_ \.\-]|$)', vf)
                    if id_match:
                        found_id = id_match.group(2)
                        if found_id in report:
                            report[found_id]["video_exists"] = True
            except Exception as e:
                logger.error(f"Error scanning videos: {e}")
                
        return report

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam") # Clam supports color customization better
        
        # Colors
        bg_color = "#f0f2f5"
        accent_color = "#4a90e2" 
        text_color = "#333333"
        white = "#ffffff"
        
        self.root.configure(bg=bg_color)
        
        # General
        style.configure("TFrame", background=bg_color)
        style.configure("TLabel", background=bg_color, foreground=text_color, font=("Helvetica", 11))
        style.configure("Header.TLabel", font=("Helvetica", 14, "bold"), foreground="#1a1a1a")
        style.configure("TLabelframe", background=bg_color, borderwidth=1)
        style.configure("TLabelframe.Label", background=bg_color, foreground=text_color, font=("Helvetica", 10, "bold"))
        
        # Buttons
        style.configure("TButton", font=("Helvetica", 10), padding=6, borderwidth=1, focuscolor="none")
        style.map("TButton", background=[("active", "#e1e4e8")])
        style.configure("Action.TButton", font=("Helvetica", 11, "bold"), background=white, foreground=accent_color)
        
        # Treeview
        style.configure("Treeview", 
                        background=white, 
                        foreground=text_color, 
                        fieldbackground=white, 
                        rowheight=30, 
                        font=("Helvetica", 10))
        style.configure("Treeview.Heading", font=("Helvetica", 10, "bold"), background="#e1e4e8")
        style.map("Treeview", background=[("selected", accent_color)], foreground=[("selected", "white")])

        # --- LAYOUT ---
        self.setup_ui()
        
    def setup_ui(self):
        # Top Bar
        self.f_top = ttk.Frame(self.root, padding=10)
        self.f_top.pack(fill="x")
        
        ttk.Button(self.f_top, text=self.t("settings"), command=self.open_settings).pack(side="right", padx=5)
        ttk.Button(self.f_top, text=self.t("refresh"), command=self._refresh_all).pack(side="right", padx=5)
        
        ttk.Label(self.f_top, text="MusicBot Pro", style="Header.TLabel").pack(side="left", padx=5)
        
        # Unified "Load Project" Button
        self.btn_new = ttk.Button(self.f_top, text=self.t("new_project"), command=self.create_new_project)
        self.btn_new.pack(side="left", padx=5)
        
        self.btn_load = ttk.Button(self.f_top, text=self.t("load_project"), command=self.load_project_file)
        self.btn_load.pack(side="left", padx=5)
        
        # Project Status Label
        self.lbl_project_text = tk.StringVar(value=self.t("no_project"))
        self.lbl_project = ttk.Label(self.f_top, textvariable=self.lbl_project_text, font=("Helvetica", 10, "italic"), foreground="gray")
        self.lbl_project.pack(side="left", padx=10)
        
        ttk.Button(self.f_top, text=self.t("open_project_btn"), command=self.open_excel_file).pack(side="left", padx=2)

        # Filter (Enhanced)
        self.f_filter = ttk.Frame(self.root, padding="5 0 5 10")
        self.f_filter.pack(fill="x", padx=10)
        ttk.Label(self.f_filter, text=self.t("search")).pack(side="left")
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", self.apply_filter)
        self.ent_filter = ttk.Entry(self.f_filter, textvariable=self.filter_var)
        self.ent_filter.pack(side="left", fill="x", expand=True, padx=5)

        # Select All / Deselect All / Active Only
        self.f_sel_controls = ttk.Frame(self.f_filter)
        self.f_sel_controls.pack(side="right")
        
        ttk.Button(self.f_sel_controls, text=self.t("select_all"), command=self.select_all, width=12).pack(side="left", padx=2)
        ttk.Button(self.f_sel_controls, text=self.t("deselect_all"), command=self.deselect_all, width=12).pack(side="left", padx=2)
        
        self.var_active_only = tk.BooleanVar(value=False)
        ttk.Checkbutton(self.f_sel_controls, text=self.t("active_only"), variable=self.var_active_only, command=self.apply_filter).pack(side="left", padx=5)

        # Status Filter Dropdown
        ttk.Label(self.f_sel_controls, text=self.t("filter_status")).pack(side="left", padx=(10, 2))
        self.status_filter_var = tk.StringVar(value="All")
        self.cmb_status = ttk.Combobox(self.f_sel_controls, textvariable=self.status_filter_var, 
                                       state="readonly", width=18)
        # Update values with new options
        self.cmb_status['values'] = [
            self.t("f_all"), self.t("f_done"), self.t("f_pending"),
            self.t("f_no_lyrics"), self.t("f_no_music"), self.t("f_no_art"), self.t("f_no_video"),
            self.t("f_missing_r"), self.t("f_missing_m1"), self.t("f_missing_m2")
        ]
        self.cmb_status.current(0)
        self.cmb_status.bind("<<ComboboxSelected>>", self.apply_filter)
        self.cmb_status.pack(side="left", padx=2)

        # Active Run Controls (MODERNIZED & INTEGRATED TO DASHBOARD)
        self.f_run_ops = ttk.LabelFrame(self.root, text=self.t("global_settings"), padding=5)
        self.f_run_ops.pack(fill="x", padx=10, pady=(0, 5))
        
        ttk.Checkbutton(self.f_run_ops, text=self.t("lyrics"), variable=self.var_run_lyrics, command=self._on_global_step_change).pack(side="left", padx=10)
        ttk.Checkbutton(self.f_run_ops, text=self.t("music"), variable=self.var_run_music, command=self._on_global_step_change).pack(side="left", padx=10)
        ttk.Checkbutton(self.f_run_ops, text=self.t("art_prompt"), variable=self.var_run_art_prompt, command=self._on_global_step_change).pack(side="left", padx=10)
        ttk.Checkbutton(self.f_run_ops, text=self.t("art_image"), variable=self.var_run_art_image, command=self._on_global_step_change).pack(side="left", padx=10)
        ttk.Checkbutton(self.f_run_ops, text=self.t("video"), variable=self.var_run_video, command=self._on_global_step_change).pack(side="left", padx=10)
        ttk.Checkbutton(self.f_run_ops, text=self.t("compilation"), variable=self.var_run_compilation, command=self._on_global_step_change).pack(side="left", padx=10)

        # Hidden vars for _quick_save_advanced compatibility (Vocal/Lyrics moved to status bar info only)
        self.var_vocal_gender_enabled = tk.BooleanVar(value=self.config.get("vocal_gender_enabled", False))
        self.var_vocal_gender = tk.StringVar(value=self.config.get("vocal_gender", self.t("vocal_default")))
        self.var_lyrics_mode_enabled = tk.BooleanVar(value=self.config.get("lyrics_mode_enabled", False))
        self.var_lyrics_mode = tk.StringVar(value=self.config.get("lyrics_mode", self.t("vocal_default")))

        # Batch Mode Section
        ttk.Separator(self.f_run_ops, orient="vertical").pack(side="left", fill="y", padx=10)
        
        self.var_suno_batch = tk.BooleanVar(value=self.config.get("suno_batch_mode", False))
        ttk.Checkbutton(self.f_run_ops, text=self.t("batch"), variable=self.var_suno_batch, command=self._quick_save_advanced).pack(side="left", padx=5)
        
        self.var_batch_op = tk.StringVar(value=self.config.get("suno_batch_op_mode", "full"))
        ttk.Radiobutton(self.f_run_ops, text="Full", variable=self.var_batch_op, value="full", command=self._quick_save_advanced).pack(side="left", padx=2)
        ttk.Radiobutton(self.f_run_ops, text="Gen", variable=self.var_batch_op, value="gen_only", command=self._quick_save_advanced).pack(side="left", padx=2)
        ttk.Radiobutton(self.f_run_ops, text="DL", variable=self.var_batch_op, value="dl_only", command=self._quick_save_advanced).pack(side="left", padx=2)
        
        self.var_turbo = tk.BooleanVar(value=True)
        ttk.Checkbutton(self.f_run_ops, text="Turbo", variable=self.var_turbo).pack(side="left", padx=5)

        # Main Table Container (converted to Notebook for tabs)
        self.f_dashboard = ttk.Frame(self.root)
        self.f_dashboard.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.notebook = ttk.Notebook(self.f_dashboard)
        self.notebook.pack(fill="both", expand=True)
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        
        # Initial call to load data (moved to after everything is ready)
        self.root.after(100, self.load_project_data)

        # --- STATUS & CONTROL (CLEAN UI) ---
        self.f_bottom = ttk.Frame(self.root, padding=10)
        self.f_bottom.pack(fill="x", side="bottom")
        
        self.status_bar = ttk.Frame(self.f_bottom)
        self.status_bar.pack(fill="x", pady=(0, 5))
        
        self.lbl_badge = tk.Label(self.status_bar, text=f" ● {self.t('badge_idle')} ", font=("Helvetica", 9, "bold"), bg="#888888", fg="white", padx=8, pady=2)
        self.lbl_badge.pack(side="left")
        
        self.status_var = tk.StringVar(value=self.t("ready"))
        ttk.Label(self.status_bar, textvariable=self.status_var, font=("Helvetica", 10, "italic")).pack(side="left", padx=10)
        
        # [User Request] Active Profile Display Combobox at bottom
        ttk.Label(self.status_bar, text="👤", font=("Helvetica", 10, "bold")).pack(side="right", padx=(5, 10))
        self.profile_var = tk.StringVar(value=self.config.get('active_preset', 'Default'))
        self.cmb_profile = ttk.Combobox(self.status_bar, textvariable=self.profile_var, state="readonly", width=15)
        self.cmb_profile['values'] = list(self.config.get("artist_presets", {"Default": {}}).keys())
        self.cmb_profile.bind("<<ComboboxSelected>>", self._on_profile_changed)
        self.cmb_profile.pack(side="right", padx=0)

        # Profile Info Badge (shows key settings at a glance, updates on profile switch)
        self.profile_info_var = tk.StringVar(value="")
        self.lbl_profile_info = ttk.Label(
            self.status_bar, textvariable=self.profile_info_var,
            font=("Helvetica", 9), foreground="#555555", cursor="hand2"
        )
        self.lbl_profile_info.pack(side="right", padx=(0, 8))
        self.lbl_profile_info.bind("<Button-1>", lambda e: self.open_settings())
        self._refresh_profile_badge()  # Fill badge on load
        
        ttk.Label(self.status_bar, text="|", foreground="gray").pack(side="left", padx=5)

        self.f_song_info = ttk.Frame(self.f_bottom)
        self.f_song_info.pack(fill="x", pady=2)
        self.lbl_current_song = ttk.Label(self.f_song_info, textvariable=self.current_song_var, font=("Helvetica", 10, "bold"), foreground="#4a90e2", wraplength=1000)
        self.lbl_current_song.pack(side="left", fill="x", expand=True)

        # Active Profile Display (Removed - Now handled by cmb_profile widget above)

        # Action Buttons
        self.f_btns = ttk.Frame(self.f_bottom)
        self.f_btns.pack(fill="x")
        
        self.btn_run = ttk.Button(self.f_btns, text=self.t("start"), style="Action.TButton", command=self.start_process)
        self.btn_run.pack(side="left", fill="x", expand=True, padx=2)
        
        self.btn_stop = ttk.Button(self.f_btns, text=self.t("stop"), command=self.stop_process)
        self.btn_stop.pack(side="left", fill="x", expand=True, padx=2)
        self.btn_stop.configure(state="disabled")

        # --- STATISTICS DASHBOARD (PHASE 6) ---
        self.f_stats = tk.Frame(self.f_bottom, bg="#f8f9fa", bd=1, relief="solid")
        self.f_stats.pack(fill="x", pady=(5, 5), ipady=4)
        
        self.var_stat_total = tk.StringVar(value="📊 Total: 0")
        self.var_stat_success = tk.StringVar(value="✅ Success: 0")
        self.var_stat_failed = tk.StringVar(value="❌ Failed: 0")
        self.var_stat_time = tk.StringVar(value="⏱ Time: 00:00")
        self.var_stat_threads = tk.StringVar(value="⚙️ Threads: 0")
        
        tk.Label(self.f_stats, textvariable=self.var_stat_total, font=("Helvetica", 10, "bold"), bg="#f8f9fa", fg="#2c3e50").pack(side="left", expand=True)
        tk.Label(self.f_stats, textvariable=self.var_stat_success, font=("Helvetica", 10, "bold"), bg="#f8f9fa", fg="#27ae60").pack(side="left", expand=True)
        tk.Label(self.f_stats, textvariable=self.var_stat_failed, font=("Helvetica", 10, "bold"), bg="#f8f9fa", fg="#c0392b").pack(side="left", expand=True)
        tk.Label(self.f_stats, textvariable=self.var_stat_time, font=("Helvetica", 10, "bold"), bg="#f8f9fa", fg="#2980b9").pack(side="left", expand=True)
        tk.Label(self.f_stats, textvariable=self.var_stat_threads, font=("Helvetica", 10, "italic"), bg="#f8f9fa", fg="#8e44ad").pack(side="left", expand=True)

        # --- COLLAPSIBLE LOGS (NEW) ---
        self.f_log_container = ttk.Frame(self.root)
        self.f_log_container.pack(fill="x", side="bottom", padx=10, pady=(0, 5))
        
        self.btn_toggle_log = ttk.Button(self.f_log_container, text=self.t("show_log"), command=self.toggle_logs)
        self.btn_toggle_log.pack(fill="x")
        
        self.f_log_content = ttk.Frame(self.f_log_container)
        # Initially hidden:
        self.log_text = scrolledtext.ScrolledText(self.f_log_content, height=10, font=("Consolas", 9), bg="#ffffff")
        self.log_text.pack(fill="both", expand=True)

        # Connect Logger
        self.gui_handler = GuiLogger(self.log_text)
        formatter = logging.Formatter('%(asctime)s - %(message)s', datefmt='%H:%M:%S')
        self.gui_handler.setFormatter(formatter)
        logging.getLogger().addHandler(self.gui_handler)
        
        # Data Cache
        self.all_songs = {} 
        self.filtered_ids = []
        
        # Initial load is handled via root.after in setup
        
        # Handle log visibility at startup
        if self.config.get("log_open_at_start", False):
            self.root.after(100, self.toggle_logs)

    def load_settings(self):
        """Loads settings from settings.json or settings.json.bak in workspace."""
        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        settings_path = os.path.join(workspace, "settings.json")
        backup_path = settings_path + ".bak"
        
        # Determine which file to load
        target_path = settings_path
        if not os.path.exists(settings_path) and os.path.exists(backup_path):
            target_path = backup_path
            logger.info("Main settings not found, attempting to load from backup.")
            
        if os.path.exists(target_path):
            try:
                with open(target_path, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if content:
                        saved_config = json.loads(content)
                        if isinstance(saved_config, dict):
                            self.config.update(saved_config)
                            logger.info(self.t("log_settings_loaded") + (f" (from backup)" if target_path == backup_path else ""))
                            
                            # Restore last profile
                            last_p = self.config.get("last_active_profile")
                            if last_p and last_p in self.config.get("artist_presets", {}):
                                self.config["active_preset"] = last_p
                                if hasattr(self, "profile_var"):
                                    self.profile_var.set(last_p)
                                if hasattr(self, "cmb_profile"):
                                    self.cmb_profile['values'] = list(self.config.get("artist_presets", {}).keys())
                                    self.cmb_profile.set(last_p)
                                if hasattr(self, "lbl_active_profile"):
                                    self.lbl_active_profile.config(text=last_p)
                        
                        # Refresh Excel path label if exists (Logic moved to tab-aware methods)
            except Exception as e:
                logger.error(f"❌ {self.t('log_settings_fail').format(error=e)}")
                # If settings are corrupted, we just continue with defaults
                # If we failed on the main file, we could try backup as a second pass
                if target_path == settings_path and os.path.exists(backup_path):
                    logger.info("Attempting recovery from backup file...")
                    try:
                        with open(backup_path, "r", encoding="utf-8") as f:
                            saved_config = json.loads(f.read().strip())
                            if isinstance(saved_config, dict):
                                self.config.update(saved_config)
                                logger.info("✅ Recovered settings from backup.")
                    except Exception as e2:
                        logger.error(f"❌ Backup recovery also failed: {e2}")

    def save_settings(self, new_config=None):
        """Saves current config to settings.json in workspace."""
        if new_config:
            self.config.update(new_config)
        
        # Sync dashboard variables to config
        if hasattr(self, "var_suno_batch"):
            self.config["suno_batch_mode"] = self.var_suno_batch.get()
        if hasattr(self, "var_batch_op"):
            self.config["suno_batch_op_mode"] = self.var_batch_op.get()

        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        os.makedirs(workspace, exist_ok=True)
        settings_path = os.path.join(workspace, "settings.json")
        backup_path = settings_path + ".bak"
        
        try:
            # Create backup of existing config before modifying
            if os.path.exists(settings_path):
                import shutil
                shutil.copy2(settings_path, backup_path)
                
            # Save last profile for persistence
            if "active_preset" in self.config:
                self.config["last_active_profile"] = self.config["active_preset"]
                if hasattr(self, "lbl_active_profile"):
                    self.lbl_active_profile.config(text=self.config["active_preset"])
                
            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            logger.info("✅ Settings saved to settings.json (backup updated)")
        except Exception as e:
            logger.error(f"Failed to save settings: {e}")

    def update_progress(self, rid, text, profile_name="Default"):
        """Updates the progress text for a specific row ID in a specific tab."""
        tab = self.tabs.get(profile_name)
        if not tab: return
        
        if tab.tree.exists(rid):
            # Update the main progress column
            tab.tree.set(rid, "progress", text)
            
            # Check for Download Status keywords
            text_lower = text.lower()
            if any(k in text_lower for k in ["iniyor", "indiriliyor"]):
                tab.tree.set(rid, "dl_status", "⏳")
            elif any(k in text_lower for k in ["indirildi", "downloaded"]):
                tab.tree.set(rid, "dl_status", "✅")
            elif any(k in text_lower for k in ["indirilemedi", "hata", "başarısız", "failed", "bulunamadı"]):
                if "hata" in text_lower or "indirilemedi" in text_lower or "bulunamadı" in text_lower:
                    tab.tree.set(rid, "dl_status", "❌")
            
            # If it's a completion mark, we might want to refresh the bar
            if "✅" in text or "Done" in text:
                # Re-fetch song state and update
                pass 
        
    def open_settings(self):
        # Pass the path to prompts.json for the PromptEditor
        prompts_path = self.get_prompts_path()
        self.config["metadata_path"] = prompts_path # This is a bit of a hack, but PromptEditor expects metadata_path
        SettingsDialog(self.root, self.config, self)

    def create_new_project(self):
        """Creates a new project file from template."""
        
        # 1. Define Workspace
        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        if not os.path.exists(workspace): os.makedirs(workspace)
        
        # 2. Ask User for Filename
        initial_file = f"Project_{int(time.time())}.xlsx"
        path = filedialog.asksaveasfilename(
            initialdir=workspace,
            initialfile=initial_file,
            title=self.t("new_project"),
            filetypes=[("Excel Files", "*.xlsx")]
        )
        
        if not path: return # User cancelled
        
        if not path.endswith(".xlsx"): path += ".xlsx"
        
        # 3. Create File (Using openpyxl directly to be self-contained)
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Music Project"
            
            headers = [
                "id", "prompt", "style", "title", "lyrics", "status", 
                "visual_prompt", "video_prompt", "suno_style", "cover_art_prompt", "cover_art_path",
                "lyrics_status", "music_status", "art_status", "video_status"
            ]
            ws.append(headers)
            # Add example row
            ws.append(["EXAMPLE_01", "A happy song about coding", "Pop", "Code Joy", "", "Pending", "", "", "", ""])
            
            wb.save(path)
            logger.info(f"Created new project: {path}")
            
            # 4. Load it immediately
            self.load_project_data(path)
            messagebox.showinfo(self.t("success"), self.t("msg_new_project_created"))
            
        except Exception as e:
            logger.error(f"Failed to create project: {e}")
            messagebox.showerror(self.t("error"), f"{self.t('msg_failed_to_create_project')}: {e}")

    def load_project_file(self):
        """Opens file dialog to select a project file."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title=self.t("load_project"),
            filetypes=[("Excel Files", "*.xlsx")],
            initialdir=os.path.expanduser("~/Documents")
        )
        if path:
            self.config["last_project"] = path
            self.load_project_data(path)

    def open_excel_file(self):
        """Opens the current project Excel file with the default system application."""
        import subprocess
        
        path = getattr(self, "project_path", None)
        if not path:
             path = self.config.get("last_project")

        if path and os.path.exists(path):
            try:
                if sys.platform == "darwin":
                    subprocess.call(["open", path])
                elif sys.platform == "win32":
                    os.startfile(path)
                else: # Linux
                    subprocess.call(["xdg-open", path])
                logger.info(f"Opening Excel file: {path}")
            except Exception as e:
                logger.error(f"Failed to open Excel: {e}")
                messagebox.showerror(self.t("error"), f"Excel açılamadı: {e}")
        else:
            messagebox.showwarning(self.t("warning"), self.t("msg_load_first"))

    def _refresh_all(self):
        """Full refresh: reload Excel data + scan filesystem for materials for the current tab."""
        tab = self.current_tab
        if not tab: return
        self.load_project_data(profile_name=tab.profile_name)
        self.scan_materials()
        self.apply_filter()

    def load_project_data(self, file_path=None, profile_name=None):
        """Loads Excel data into the specified or active tab in a background thread."""
        active_profile = profile_name if profile_name else self.config.get("active_preset", "Default")
        tab = self._get_or_create_tab(active_profile)
        self.current_tab = tab
        
        # Select the tab in notebook immediately on main thread
        self.notebook.select(tab.tab_frame)
        self.status_var.set(f"{self.t('loading')}...")
        
        if not file_path:
            # 1. Try profile-specific project
            presets = self.config.get("artist_presets", {})
            active_data = presets.get(active_profile, {})
            file_path = active_data.get("project_file")
            
            # 2. Fallback to global last_project
            if not file_path:
                file_path = self.config.get("last_project")
            
        if not file_path or not os.path.exists(file_path):
            self.lbl_project_text.set(self.t("no_project"))
            tab.project_file = None
            tab.all_songs = {}
            for item in tab.tree.get_children():
                tab.tree.delete(item)
            self.status_var.set(self.t("ready"))
            return

        if self._is_loading:
            logger.warning("Load already in progress. Skipping.")
            return
            
        self._is_loading = True
        tab.project_file = file_path
        self.project_path = file_path
        self.config["last_project"] = file_path
        
        fname = os.path.basename(file_path)
        self.lbl_project_text.set(f"📊 {fname}")
        
        # RUN HEAVY I/O IN BACKGROUND
        threading.Thread(target=self._load_project_worker, args=(file_path, tab, active_profile), daemon=True).start()

    def _load_project_worker(self, file_path, tab, active_profile):
        """Worker thread for load_project_data."""
        try:
            # Ensure Output Logic
            self.ensure_project_structure(file_path)
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = {str(cell.value).strip().lower(): i for i, cell in enumerate(ws[1]) if cell.value}
            
            # Robust column mapping for pre-checks
            visual_col = headers.get("visual_prompt") or headers.get("visual prompt") or headers.get("visual_prompts")
            video_col = headers.get("video_prompt") or headers.get("video prompt") or headers.get("video_prompts")
            lyrics_col = headers.get("lyrics")

            rows_data = []
            song_ids = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                id_idx = headers.get('id', 0)
                if id_idx is not None and id_idx < len(row):
                    rid = str(row[id_idx]) if row[id_idx] is not None else ""
                    if not rid: continue
                    rows_data.append(row)
                    song_ids.append(rid)

            # Batched Material Check
            materials_report = self._get_materials_report(active_profile, file_path, song_ids)
            
            # Prepare internal data
            new_all_songs = {}
            new_all_songs_data = {}
            new_song_steps = {}
            
            for row in rows_data:
                id_idx = headers.get('id', 0)
                rid = str(row[id_idx])
                
                rep = materials_report.get(rid, {"materials": 0, "video_exists": False})
                stat_idx = headers.get('status')
                status_val = str(row[stat_idx]) if stat_idx is not None and stat_idx < len(row) and row[stat_idx] else ""
                
                # Look up robust indices
                t_idx = headers.get('title') or headers.get('başlık / prompt') or headers.get('başlık') or headers.get('baslik') or 1
                p_idx = headers.get('prompt') or headers.get('lyrics_master_prompt') or headers.get('lyrics') or None
                s_idx = headers.get('style') or headers.get('tarz (style)') or headers.get('tarz') or 2
                ss_idx = headers.get('suno_style') or None
                
                # Fetch values
                t_val = str(row[t_idx]).replace('\n', ' ').strip() if t_idx < len(row) and row[t_idx] is not None else ""
                p_val = str(row[p_idx]).replace('\n', ' ').strip() if p_idx is not None and p_idx < len(row) and row[p_idx] is not None else ""
                s_val = str(row[s_idx]).replace('\n', ' ').strip() if s_idx < len(row) and row[s_idx] is not None else ""
                ss_val = str(row[ss_idx]).replace('\n', ' ').strip() if ss_idx is not None and ss_idx < len(row) and row[ss_idx] is not None else ""
                
                # Apply fallback logic
                title_val = t_val if t_val else (p_val if p_val else "-")
                style_val = s_val if s_val else (ss_val if ss_val else "-")
                
                new_all_songs[rid] = rid
                new_all_songs_data[rid] = {
                    "title": title_val,
                    "style": style_val,
                    "status": status_val,
                    "lyrics": (lyrics_col is not None and lyrics_col < len(row) and row[lyrics_col]),
                    "music": (headers.get('music_url_1') is not None and headers.get('music_url_1') < len(row) and row[headers.get('music_url_1')]),
                    "art": (headers.get('cover_art_path') is not None and headers.get('cover_art_path') < len(row) and row[headers.get('cover_art_path')]),
                    "video": (headers.get('video_path') is not None and headers.get('video_path') < len(row) and row[headers.get('video_path')]),
                    "video_exists": rep.get("video_exists", False),
                    "dl_status": str(row[headers.get('dl_status')]) if headers.get('dl_status') is not None and headers.get('dl_status') < len(row) else "",
                    "material_status": f"📦 {rep['materials']}" if rep["materials"] > 0 else "-",
                    # Pre-check flags (O(1) lookup during start)
                    "has_lyrics": bool(lyrics_col is not None and lyrics_col < len(row) and row[lyrics_col] and str(row[lyrics_col]).strip()),
                    "has_visual": bool(visual_col is not None and visual_col < len(row) and row[visual_col] and str(row[visual_col]).strip()),
                    "has_video_prompt": bool(video_col is not None and video_col < len(row) and row[video_col] and str(row[video_col]).strip()),
                    # Material flags
                    "has_r": rep.get("has_r", False),
                    "has_m1": rep.get("has_m1", False),
                    "has_m2": rep.get("has_m2", False)
                }
                
                # Default Phase Selection
                new_song_steps[rid] = [
                    self.var_run_lyrics.get(),
                    self.var_run_music.get(),
                    self.var_run_art_prompt.get(),
                    self.var_run_art_image.get(),
                    self.var_run_video.get()
                ]

            # MARSHAL TO MAIN THREAD
            def _finalize_ui():
                tab.all_songs = new_all_songs
                tab.all_songs_data = new_all_songs_data
                for rid, steps in new_song_steps.items():
                    if rid not in tab.song_steps:
                        tab.song_steps[rid] = steps
                
                # Sync global song_steps too
                self.song_steps.update(tab.song_steps)
                
                self.apply_filter()
                self.status_var.set(self.t("ready"))
                self._is_loading = False
                logger.info(f"Loaded {len(tab.all_songs_data)} songs into tab: {active_profile}")

            self.root.after(0, _finalize_ui)
            
        except Exception as e:
            self._is_loading = False
            logger.error(f"Failed to load project: {e}", exc_info=True)
            self.root.after(0, lambda: self.status_var.set(f"Error: {e}"))

    def _check_video_exists(self, rid, profile_name):
        """Helper to specifically check for .mp4 Existence in the profile's video directory."""
        tab = self.tabs.get(profile_name)
        if not tab or not tab.project_file: return False
        
        output_media = os.path.join(os.path.dirname(tab.project_file), "output_media", profile_name)
        videos_dir = os.path.join(output_media, "videos")
        if not os.path.exists(videos_dir): return False
        
        try:
            video_files = [f.lower() for f in os.listdir(videos_dir) if f.lower().endswith(".mp4")]
            rid_l = str(rid).lower()
            return any(self._is_id_match(vf, rid_l) for vf in video_files)
        except: return False
    def _count_materials(self, rid, profile_name):
        """Helper to count materials for a song in a specific profile's directory."""
        tab = self.tabs.get(profile_name)
        if not tab or not tab.project_file: return 0
        
        output_media = os.path.join(os.path.dirname(tab.project_file), "output_media", profile_name)
        if not os.path.exists(output_media): return 0
        
        all_files = [f.lower() for f in os.listdir(output_media)]
        count = 0
        rid_l = str(rid).lower()
        
        # Check image (R)
        if any(self._is_id_match(f, rid_l) and f.split(".")[-1] in ["png", "jpg", "jpeg"] for f in all_files):
            count += 1
            
        # Check audio (M1, M2)
        has_m1 = any(f.endswith((".mp3", ".wav")) and (os.path.splitext(f)[0] == rid_l or self._is_id_match(f, rid_l, variant="1")) for f in all_files)
        has_m2 = any(f.endswith((".mp3", ".wav")) and self._is_id_match(f, rid_l, variant="2") for f in all_files)
        
        if has_m1: count += 1
        if has_m2: count += 1
        
        return count

    def _on_tab_changed(self, event):
        """Handle tab switch logic."""
        selected_id = self.notebook.select()
        if not selected_id: return
        
        tab_name = self.notebook.tab(selected_id, "text")
        if tab_name in self.tabs:
            self.current_tab = self.tabs[tab_name]
            # Synchronize the profile combo box if it's not already correct
            if hasattr(self, 'cmb_profile') and self.cmb_profile.get() != tab_name:
                self.cmb_profile.set(tab_name)
            
            # Refresh project label
            if self.current_tab.project_file and os.path.exists(self.current_tab.project_file):
                self.project_path = self.current_tab.project_file
                fname = os.path.basename(self.current_tab.project_file)
                self.lbl_project_text.set(f"📊 {fname}")
            else:
                self.lbl_project_text.set(self.t("no_project"))

    def _get_or_create_tab(self, profile_name):
        """Returns the TabState for a profile, creating it if needed."""
        if profile_name in self.tabs:
            return self.tabs[profile_name]
        
        # Create Frame for Tab
        f_tab = ttk.Frame(self.notebook)
        self.notebook.add(f_tab, text=profile_name)
        
        # Setup Treeview inside Tab
        columns = ("sel", "id", "title", "style", "progress", "lyrics", "music", "art", "dl_status", "video_status", "materials", "run_l", "run_m", "run_ap", "run_ai", "run_v")
        tree = ttk.Treeview(f_tab, columns=columns, show="headings", selectmode="extended")
        
        # Bind events
        tree.bind("<Button-1>", self.on_tree_click)
        tree.bind("<Button-2>", self.on_right_click)
        tree.bind("<Button-3>", self.on_right_click)
        tree.bind("<Motion>", self.on_tree_hover)
        tree.bind("<Leave>", self.on_tree_leave)
        
        # Style Tags
        tree.tag_configure("hover", background="#eef6ff") 
        tree.tag_configure("done", foreground="gray", font=("Helvetica", 9, "italic"))
        tree.tag_configure("missing", foreground="#e67e22") 
        tree.tag_configure("error", background="#ffcccc", foreground="red") 
        
        # Reorder events
        tree.bind("<ButtonPress-1>", self.on_reorder_start, add="+")
        tree.bind("<B1-Motion>", self.on_reorder_motion)
        tree.bind("<ButtonRelease-1>", self.on_reorder_stop)

        # Setup Columns
        tree.heading("sel", text="✔")
        tree.heading("id", text="ID", command=lambda: self.sort_tree("id", False))
        tree.heading("title", text=self.t("column_title"), command=lambda: self.sort_tree("title", False))
        tree.heading("style", text=self.t("column_style"))
        tree.heading("progress", text=self.t("column_progress"))
        tree.heading("lyrics", text="LYR", command=lambda: self.sort_tree("lyrics", False))
        tree.heading("music", text="MUS", command=lambda: self.sort_tree("music", False))
        tree.heading("art", text="ART", command=lambda: self.sort_tree("art", False))
        tree.heading("dl_status", text="DL", command=lambda: self.sort_tree("dl_status", False))
        tree.heading("video_status", text="VID", command=lambda: self.sort_tree("video_status", False))
        tree.heading("materials", text=self.t("column_materials"), command=lambda: self.sort_tree("materials", False))
        tree.heading("run_l", text="L")
        tree.heading("run_m", text="M")
        tree.heading("run_ap", text="AP")
        tree.heading("run_ai", text="AI")
        tree.heading("run_v", text="V")
        
        # Columns Config
        tree.column("sel", width=30, anchor="center")
        tree.column("id", width=50, anchor="center")
        tree.column("title", width=200)
        tree.column("style", width=100)
        tree.column("progress", width=120)
        tree.column("lyrics", width=40, anchor="center")
        tree.column("music", width=40, anchor="center")
        tree.column("art", width=40, anchor="center")
        tree.column("dl_status", width=40, anchor="center")
        tree.column("video_status", width=40, anchor="center")
        tree.column("materials", width=80, anchor="center")
        tree.column("run_l", width=30, anchor="center")
        tree.column("run_m", width=30, anchor="center")
        tree.column("run_ap", width=30, anchor="center")
        tree.column("run_ai", width=30, anchor="center")
        tree.column("run_v", width=30, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(f_tab, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        tree.pack(fill="both", expand=True)
        
        tab_state = TabState(f_tab, tree, scrollbar, profile_name)
        self.tabs[profile_name] = tab_state
        return tab_state

    def scan_materials(self, profile_name=None):
        """Scans filesystem to detect material presence and existing videos. Updates internal cache in a background thread."""
        active_profile = profile_name if profile_name else (self.current_tab.profile_name if self.current_tab else None)
        if not active_profile: return
        
        tab = self.tabs.get(active_profile)
        if not tab or not tab.project_file: return
        
        project_file = tab.project_file
        song_ids = list(tab.all_songs_data.keys())
        if not song_ids: return

        def _scan_worker():
            try:
                # Fix: Use active_profile instead of profile_name to avoid NoneType join error
                report = self._get_materials_report(active_profile, project_file, song_ids)
                
                def _update_ui():
                    changed = False
                    for rid, data in report.items():
                        if rid in tab.all_songs_data:
                            new_m = f"📦 {data['materials']}" if data['materials'] > 0 else "-"
                            new_v = data["video_exists"]
                            
                            # Update granular material flags too
                            if tab.all_songs_data[rid].get("material_status") != new_m or \
                               tab.all_songs_data[rid].get("video_exists") != new_v or \
                               tab.all_songs_data[rid].get("has_r") != data.get("has_r") or \
                               tab.all_songs_data[rid].get("has_m1") != data.get("has_m1") or \
                               tab.all_songs_data[rid].get("has_m2") != data.get("has_m2"):
                               
                                tab.all_songs_data[rid]["material_status"] = new_m
                                tab.all_songs_data[rid]["video_exists"] = new_v
                                tab.all_songs_data[rid]["has_r"] = data.get("has_r", False)
                                tab.all_songs_data[rid]["has_m1"] = data.get("has_m1", False)
                                tab.all_songs_data[rid]["has_m2"] = data.get("has_m2", False)
                                changed = True
                    
                    if changed:
                        self.apply_filter()

                self.root.after(0, _update_ui)
            except Exception as e:
                logger.error(f"Background material scan failed: {e}")

        threading.Thread(target=_scan_worker, daemon=True).start()

    def ensure_project_structure(self, path):
        """Adds missing columns to a raw input file to make it a Project File."""
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            
            # Map current headers
            headers = {str(cell.value).lower(): cell.column for cell in ws[1] if cell.value}
            
            # Required Columns
            required = [
                "id", "prompt", "style", "title", "lyrics", "status", 
                "visual_prompt", "video_prompt", "suno_style", "cover_art_prompt", "cover_art_path",
                "dl_status", "dl_attempts", "lyrics_status", "music_status", "art_status", "video_status"
            ]
            
            updates = False
            for col in required:
                if col not in headers:
                    new_idx = ws.max_column + 1
                    ws.cell(row=1, column=new_idx, value=col)
                    headers[col] = new_idx
                    updates = True
            
            if updates:
                wb.save(path)
                logger.info(f"Project structure initialized for {os.path.basename(path)}")
        except Exception as e:
            logger.error(f"Structure init error: {e}")

    def update_project_excel(self, rid, project_file=None, **kwargs):
        """Atomic helper to update specific columns for a song ID in the Excel file."""
        if project_file:
            path = project_file
        else:
            tab = self.current_tab
            if not tab or not tab.project_file: return
            path = tab.project_file
            
        if not path or not os.path.exists(path): return

        def _do_update():
            try:
                wb = openpyxl.load_workbook(path)
                ws = wb.active
                headers = {str(cell.value).strip().lower(): cell.column for cell in ws[1] if cell.value}
                
                # Ensure keys exist
                for key in kwargs.keys():
                    if key not in headers:
                        new_idx = ws.max_column + 1
                        ws.cell(row=1, column=new_idx, value=key)
                        headers[key] = new_idx
                
                target_row = None
                id_col = headers.get("id", 1)
                for r in range(2, ws.max_row + 1):
                    if str(ws.cell(row=r, column=id_col).value).strip() == str(rid).strip():
                        target_row = r
                        break
                
                if target_row:
                    for key, val in kwargs.items():
                        ws.cell(row=target_row, column=headers[key], value=val)
                    wb.save(path)
            except Exception as e:
                logger.error(f"Excel update error for {rid}: {e}")

        if self.xlsx_lock:
            with self.xlsx_lock: _do_update()
        else:
            _do_update()
    def get_progress_bar(self, current, total=3):
        """Returns a text-based progress bar."""
        try:
            percent = (current / total) * 100
            filled_len = int(10 * current // total)
            bar = "▰" * filled_len + "▱" * (10 - filled_len)
            return f"{bar} {int(percent)}%"
        except Exception: return "▱▱▱▱▱▱▱▱▱▱ 0%"

    def apply_filter(self, *args, profile_name=None):
        """Debounced filter call. If profile_name is None, uses current_tab."""
        if hasattr(self, '_filter_after_id'):
            self.root.after_cancel(self._filter_after_id)
        # Pass profile_name through after
        self._filter_after_id = self.root.after(300, lambda: self._do_filter(profile_name=profile_name))
            
    def _do_filter(self, profile_name=None):
        tab = self.tabs.get(profile_name) if profile_name else self.current_tab
        if not tab: return
        
        # If refreshing current tab, sync with global UI. If other tab, use its stored state.
        if tab == (self.current_tab if hasattr(self, "current_tab") else None):
            query = self.filter_var.get().strip()
            active_only = self.var_active_only.get()
            s_filter = self.status_filter_var.get()
            # Save to tab for background use
            tab.filter_var.set(query)
            tab.active_only_var.set(active_only)
            tab.status_filter_var.set(s_filter)
        else:
            query = tab.filter_var.get().strip()
            active_only = tab.active_only_var.get()
            s_filter = tab.status_filter_var.get()
        
        query_lower = query.lower()
        
        # --- ID RANGE DETECTION ---
        id_range_min = None
        id_range_max = None
        import re as _re
        range_match = _re.match(r'^(\d+)\s*-\s*(\d+)$', query)
        if range_match:
            id_range_min = int(range_match.group(1))
            id_range_max = int(range_match.group(2))
            query_lower = "" 

        for item in tab.tree.get_children():
            tab.tree.delete(item)
            
        display_items = []
        # We need all_songs but scoped to the profile or at least filtered correctly.
        # Actually, all_songs is loaded per tab in load_project_data now.
        # BUT we need to ensure it's stored in TabState.
        
        if not hasattr(tab, "all_songs_data"): 
            # Fallback if not cached, but load_project_data should have handled it
            tab.all_songs_data = {} 
            
        for rid, s in getattr(tab, "all_songs_data", {}).items():
            # 1. Text Query OR ID Range Filter
            if id_range_min is not None:
                try:
                    rid_num = int(rid)
                    if not (id_range_min <= rid_num <= id_range_max): continue
                except ValueError: continue
            elif query_lower and query_lower not in str(s.get("title", "")).lower() and query_lower not in rid.lower():
                continue
            
            # 2. Active Only Filter
            if active_only and rid not in tab.selected_songs:
                continue

            # 3. Status Filter
            done_cnt = sum([1 for k in ["lyrics", "music", "art"] if s.get(k)])
            video_done = s.get("video") or s.get("video_exists")

            if s_filter == self.t("f_done") and (done_cnt < 3 or not video_done): continue
            elif s_filter == self.t("f_pending") and (done_cnt == 3 and video_done): continue
            elif s_filter == self.t("f_no_lyrics") and s.get("lyrics"): continue
            elif s_filter == self.t("f_no_music") and s.get("music"): continue
            elif s_filter == self.t("f_no_art") and s.get("art"): continue
            elif s_filter == self.t("f_no_video"):
                if video_done: continue
                # Yalnızca malzemesi tam (resim ve ses var) ama videosu olmayanları göster
                if not s.get("has_r") or not (s.get("has_m1") or s.get("has_m2")):
                    continue
            
            # Material specific filters (using report flags)
            elif s_filter == self.t("f_missing_r") and s.get("has_r"): continue
            elif s_filter == self.t("f_missing_m1") and s.get("has_m1"): continue
            elif s_filter == self.t("f_missing_m2") and s.get("has_m2"): continue

            # Symbols
            s_lyrics = "✅" if s.get("lyrics") else "⚪"
            s_music = "✅" if s.get("music") else "⚪"
            s_art = "✅" if s.get("art") else "⚪"
            video_done = s.get("video") or s.get("video_exists")
            s_v = "✅" if video_done else "⚪"
            s_dl = "✅" if "indirildi" in str(s.get("dl_status", "")).lower() else "⚪"
            
            s_sel = "☑️" if rid in tab.selected_songs else "☐"
            prog_bar = self.get_progress_bar(done_cnt + (1 if video_done else 0), 4)
            
            steps = tab.song_steps.get(rid, self.get_global_steps())
            s_rl = "🟣" if steps[0] else "☐"
            s_rm = "🔵" if steps[1] else "☐"
            s_rap = "🟡" if steps[2] else "☐"
            s_rai = "🟢" if steps[3] else "☐"
            s_rv = "🟠" if (len(steps) > 4 and steps[4]) else "☐"

            row_tags = []
            if video_done: row_tags.append("done")
            elif "error" in str(s.get("status", "")).lower(): row_tags.append("error")
            
            display_items.append({
                "rid": rid, "values": (s_sel, rid, s.get("title", "-"), s.get("style", "-"), prog_bar, s_lyrics, s_music, s_art, s_dl, s_v, s.get("material_status", "-"), s_rl, s_rm, s_rap, s_rai, s_rv),
                "tags": tuple(row_tags),
                "sort": (1 if video_done else 0, rid)
            })

        display_items.sort(key=lambda x: x["sort"])
        for item in display_items:
             tab.tree.insert("", "end", iid=item["rid"], values=item["values"], tags=item["tags"])

    def select_all(self):
        tab = self.current_tab
        if not tab: return
        for itm in tab.tree.get_children():
            tab.selected_songs.add(itm)
        self.apply_filter()

    def tree_clear_selected(self):
        tab = self.current_tab
        if not tab: return
        tab.selected_songs.clear()
        self.apply_filter()

    def get_global_steps(self):
        """Returns the current state of UI global checkboxes for song step defaults."""
        return [
            self.var_run_lyrics.get(), 
            self.var_run_music.get(),
            self.var_run_art_prompt.get(),
            self.var_run_art_image.get(),
            self.var_run_video.get()
        ]

    def _on_global_step_change(self, *args):
        """When a global step checkbox is toggled, override all song-specific selections so UI Orbs sync instantly."""
        tab = self.current_tab
        if tab:
            tab.song_steps.clear() # Clear specific overrides to force fallback to get_global_steps()
        self.apply_filter()

    def deselect_all(self):
        tab = self.current_tab
        if not tab: return
        tab.selected_songs.clear()
        self.apply_filter()

    def handle_ctrl_a(self, event):
        """Intelligent Select All: Text in Entry vs Songs in Tree."""
        widget = self.root.focus_get()
        if isinstance(widget, (tk.Entry, tk.Text, ttk.Entry)):
            widget.event_generate("<<SelectAll>>")
            return "break"
        self.select_all()
        return "break"

    def sort_tree(self, col, reverse):
        """Sorts the treeview of the active tab by column."""
        tab = self.current_tab
        if not tab: return
        
        # Get data from active tree
        l = [(tab.tree.set(k, col), k) for k in tab.tree.get_children("")]
        
        try:
            l.sort(key=lambda x: float(re.sub(r'[^\d.]', '', x[0])), reverse=reverse)
        except (ValueError, TypeError):
            l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tab.tree.move(k, "", index)

        tab.tree.heading(col, command=lambda: self.sort_tree(col, not reverse))

    def update_dashboard_stats(self, **kwargs):
        """Updates the Tkinter statistics dashboard."""
        if "total" in kwargs:
            self.var_stat_total.set(f"📊 Total: {kwargs['total']}")
        if "success" in kwargs:
            self.var_stat_success.set(f"✅ Success: {kwargs['success']}")
        if "failed" in kwargs:
            self.var_stat_failed.set(f"❌ Failed: {kwargs['failed']}")
        if "time" in kwargs:
            self.var_stat_time.set(f"⏱ Time: {kwargs['time']}")
        if "threads" in kwargs:
            self.var_stat_threads.set(f"⚙️ Threads: {kwargs['threads']}")
        self.root.update_idletasks()
            
    def set_badge(self, text, bg, fg="white"):
        self.lbl_badge.config(text=f" ● {text} ", bg=bg, fg=fg)

    def toggle_logs(self):
        if self.log_visible.get():
            self.f_log_content.pack_forget()
            self.btn_toggle_log.config(text=self.t("show_log"))
            self.log_visible.set(False)
        else:
            self.f_log_content.pack(fill="x", pady=5)
            self.btn_toggle_log.config(text=self.t("hide_log"))
            self.log_visible.set(True)

    def on_space_toggle(self, event):
        """Selected rows toggle checkbox with space in the active tab."""
        tab = self.current_tab
        if not tab: return
        
        selected = tab.tree.selection()
        if not selected: return
        
        for item_id in selected:
            if item_id in tab.selected_songs:
                tab.selected_songs.remove(item_id)
                tab.tree.set(item_id, "sel", "☐")
            else:
                tab.selected_songs.add(item_id)
                tab.tree.set(item_id, "sel", "☑️")

    def _on_profile_changed(self, event=None):
        """Standard profile switch: Switches or creates a tab and loads data."""
        new_profile = self.profile_var.get()
        if not new_profile: return
        
        self.config["active_preset"] = new_profile
        
        # Switch to the tab (will create if missing via load_project_data)
        self.load_project_data()
        
        self.save_settings()
        self._refresh_profile_badge()
        logger.info(f"Dashboard Profile Changed: {new_profile}")

    def _refresh_profile_badge(self):
        """Shows a compact info chip with per-profile details from the preset data."""
        try:
            parts = []
            
            # Read from preset's own settings (not global config)
            presets = self.config.get("chrome_presets", {})
            active = self.config.get("active_preset", "")
            preset_data = presets.get(active, {})
            s = preset_data.get("settings", {})
            
            # Fallback to global config if preset has no settings yet
            if not s:
                s = self.config
            
            # Target Language
            lang = s.get("target_language", self.config.get("target_language", ""))
            if lang:
                parts.append(f"🌐 {lang}")
            
            # Vocal Gender
            if s.get("vocal_gender_enabled", False):
                gender = s.get("vocal_gender", "")
                if gender and gender != "Default":
                    parts.append(f"🎤 {gender}")
            
            # Active Persona
            persona = s.get("suno_active_persona", "")
            if persona:
                parts.append(f"👤 {persona}")
            
            # Artist Name
            artist = s.get("artist_name", "")
            if artist:
                parts.append(f"🎨 {artist}")
            
            # Batch Mode
            if s.get("suno_batch_mode", False):
                op = s.get("suno_batch_op_mode", "full")
                parts.append(f"⚡ Batch:{op}")
            
            # Lyrics Mode
            if s.get("lyrics_mode_enabled", False):
                mode = s.get("lyrics_mode", "")
                if mode and mode != "Default":
                    parts.append(f"📜 {mode}")
            
            # Weirdness
            if s.get("weirdness_enabled", False):
                weird = s.get("weirdness", "")
                if weird and str(weird) != "Default":
                    parts.append(f"☄️ W:{weird}")
            
            # Audio Influence
            if s.get("audio_influence_enabled", False):
                ai = s.get("audio_influence", "")
                if ai:
                    parts.append(f"🔊 AI:{ai}%")
            
            # Project file indicator
            pf = preset_data.get("project_file", "")
            if pf:
                parts.append(f"📄 {os.path.basename(pf)}")
            
            badge_text = "  │  ".join(parts) if parts else "ℹ️ Ayarlar için tıkla"
            self.profile_info_var.set(badge_text)
        except Exception:
            pass

    def _quick_save_advanced(self, event=None):
        """Instantly saves Advanced Options changes to config."""
        self.config["vocal_gender_enabled"] = self.var_vocal_gender_enabled.get()
        self.config["vocal_gender"] = self.var_vocal_gender.get()
        self.config["lyrics_mode_enabled"] = self.var_lyrics_mode_enabled.get()
        self.config["lyrics_mode"] = self.var_lyrics_mode.get()
        self.config["suno_batch_mode"] = self.var_suno_batch.get()
        self.config["suno_batch_op_mode"] = self.var_batch_op.get()
        self.save_settings()

    def play_chime(self):
        """Plays a notification sound based on OS."""
        try:
            if sys.platform == "darwin":
                os.system("afplay /System/Library/Sounds/Glass.aiff &")
            elif sys.platform == "win32":
                import winsound
                winsound.PlaySound("SystemAsterisk", winsound.SND_ALIAS)
        except Exception: pass



    # --- DRAG & DROP REORDERING ---
    def on_reorder_start(self, event):
        tab = self.current_tab
        if not tab: return
        item = tab.tree.identify_row(event.y)
        if item: self._drag_data["item"] = item

    def on_reorder_motion(self, event):
        pass

    def on_reorder_stop(self, event):
        tab = self.current_tab
        if not tab: return
        target = tab.tree.identify_row(event.y)
        source = self._drag_data.get("item")
        if target and source and target != source:
            tab.tree.move(source, "", tab.tree.index(target))
        self._drag_data["item"] = None
        
    def on_right_click(self, event):
        tab = self.current_tab
        if not tab: return
        item_id = tab.tree.identify_row(event.y)
        if item_id:
            tab.tree.selection_set(item_id)
            tab.tree.focus(item_id)
            try:
                self.tree_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.tree_menu.grab_release()

    def on_tree_click(self, event):
        tab = self.current_tab
        if not tab: return
        region = tab.tree.identify("region", event.x, event.y)
        if region == "cell":
            column = tab.tree.identify_column(event.x)
            item_id = tab.tree.identify_row(event.y)
            if not item_id: return

            if column == "#11": return
                
            if column not in ["#12", "#13", "#14", "#15", "#16"]:
                if item_id in tab.selected_songs:
                    tab.selected_songs.remove(item_id)
                    tab.tree.set(item_id, "sel", "☐")
                else:
                    tab.selected_songs.add(item_id)
                    tab.tree.set(item_id, "sel", "☑️")
                
                tab.tree.selection_set(item_id)
                tab.tree.focus(item_id)
            
            elif column in ["#12", "#13", "#14", "#15", "#16"]: # L, M, AP, AI, V
                idx = int(column[1:]) - 12
                # UI Global sync:
                steps = tab.song_steps.get(item_id, self.get_global_steps()).copy()
                while len(steps) < 5: steps.append(False)
                
                steps[idx] = not steps[idx]
                tab.song_steps[item_id] = steps
                
                char = "☐"
                if steps[idx]:
                    char = ["🟣", "🔵", "🟡", "🟢", "🟠"][idx]

                col_name = tab.tree.cget("columns")[idx + 11] # Adjust index mapping
                tab.tree.set(item_id, col_name, char)
            
            return "break"
            
    def on_tree_hover(self, event):
        tab = self.current_tab
        if not tab: return
        item = tab.tree.identify_row(event.y)
        for i in tab.tree.tag_has("hover"):
            tags = list(tab.tree.item(i, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                tab.tree.item(i, tags=tuple(tags))
        if item:
            tags = list(tab.tree.item(item, "tags"))
            if "hover" not in tags:
                tags.append("hover")
                tab.tree.item(item, tags=tuple(tags))

    def on_tree_leave(self, event):
        tab = self.current_tab
        if not tab: return
        for i in tab.tree.tag_has("hover"):
            tags = list(tab.tree.item(i, "tags"))
            if "hover" in tags:
                tags.remove("hover")
                tab.tree.item(i, tags=tuple(tags))

    def start_process(self):
        try:
            self._start_process_internal()
        except Exception as e:
            logger.error(f"Critical error in start_process: {e}", exc_info=True)
            import traceback
            messagebox.showerror(self.t("error"), f"Beklenmeyen bir hata oluştu:\n{e}\n\n{traceback.format_exc()}")

    def _start_process_internal(self):
        tab = self.current_tab
        if not tab or not tab.project_file:
            messagebox.showerror(self.t("error"), self.t("msg_load_first"))
            return

        profile_name = tab.profile_name
        if profile_name in self.active_tasks:
            messagebox.showwarning(self.t("warning"), f"'{profile_name}' {self.t('msg_already_running')}")
            return

        if not os.path.exists(tab.project_file):
             logger.error(f"❌ Project file not found: {tab.project_file}")
             return

        # Check if at least one step is selected
        has_step = any([
            self.var_run_lyrics.get(), 
            self.var_run_music.get(), 
            self.var_run_art_prompt.get(), 
            self.var_run_art_image.get(), 
            self.var_run_video.get(), 
            self.var_run_compilation.get()
        ])
        
        # If no main step is selected but "Batch Mode" is enabled, we implicitly enable "Music" (Step 2) 
        # to fulfill the user's intent to run a Batch Suno operation.
        if not has_step:
            if hasattr(self, "var_suno_batch") and self.var_suno_batch.get():
                logger.info("Batch mode active with no explicit phase selected. Auto-enabling Music phase.")
                self.var_run_music.set(True)
            else: # This else block was removed by the instruction, but keeping it for logical consistency.
                logger.error(self.t("msg_no_steps"))
                messagebox.showwarning(self.t("warning"), self.t("msg_select_step_warn"))
                return

        # Check for saved session state before taking new selection
        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        state_file = os.path.join(workspace, "session_state.json")
        resumed = False
        target_ids = []
        
        if os.path.exists(state_file):
            try:
                with open(state_file, "r", encoding="utf-8") as f:
                    state = json.load(f)
                if state.get("project_file") == getattr(self, 'project_path', '') and state.get("target_ids"):
                    if messagebox.askyesno("Devam Et (Resume)", "Yarıda kalmış bir işlem bulundu. Kaldığınız yerden (önceki seçimlerinizle) devam edilsin mi?\n(Hayır derseniz, sıfırdan başlarsınız.)"):
                        target_ids = state["target_ids"]
                        resumed = True
                    else:
                        try: os.remove(state_file)
                        except: pass
            except: pass

        if not resumed:
            # 1. Use manual checkboxes if any
            target_ids = list(tab.selected_songs)
            
            # 2. If no checkboxes, use tree selection
            if not target_ids:
                selected_items = tab.tree.selection()
                if selected_items:
                    target_ids = list(selected_items)
                else:
                    # If filter is active and nothing selected, ask to process all filtered
                    # We need filtered_ids to be tab-aware too
                    filtered_ids = [k for k in tab.all_songs_data.keys()] # Simplest fallback
                    if messagebox.askyesno(self.t("confirm"), self.t("msg_confirm_process_all")):
                        target_ids = filtered_ids
        
        if not target_ids and not any(str(s.get("status", "")).upper() == "YARIM" for s in tab.all_songs_data.values()):
            logger.warning("No target songs selected. Process cannot start.")
            messagebox.showwarning(self.t("warning"), "Lütfen işlenecek şarkıları seçin!\n(İşlem yapılacak satırları işaretleyin veya filtreyi kullanarak 'Hepsini İşle' deyin.)")
            return

        # [Requirement 6] Strict Sequential Processing by ID
        try:
            target_ids.sort(key=lambda x: (float(re.sub(r'[^\d.]', '', x)) if re.search(r'\d', x) else 0, x))
        except (ValueError, TypeError):
            target_ids.sort()

        # --- GLOBAL YARIM AUTO-SCAN (POST-SORT) ---
        yarim_ids = []
        if yarim_ids:
            logger.info(f"Auto-Retry: Found {len(yarim_ids)} YARIM items. Appending to the end of target list with Lyrics phase forced ON.")
            target_ids.extend(yarim_ids)
            
        # --- Update song_steps for ALL target items from CURRENT dashboard state ---
        # This ensures that global checkbox changes are applied to selected items
        current_global_steps = [
            self.var_run_lyrics.get(),
            self.var_run_music.get(),
            self.var_run_art_prompt.get(),
            self.var_run_art_image.get(),
            self.var_run_video.get()
        ]
        for rid in target_ids:
            # For YARIM items, we already set s_steps[0]=True above, but let's be safe
            if rid in yarim_ids:
                steps = list(current_global_steps)
                steps[0] = True
                tab.song_steps[rid] = steps
            else:
                tab.song_steps[rid] = list(current_global_steps)
        
        # Sync global copy
        self.song_steps.update(tab.song_steps)

        # --- Evaluate Re-generation (Main Thread) ---
        force_update = False
        any_gemini_requested = False
        for rid in target_ids:
            s_steps = self.song_steps.get(rid)
            if s_steps is None:
                s_steps = [
                    self.var_run_lyrics.get(),
                    self.var_run_music.get(),
                    self.var_run_art_prompt.get(),
                    self.var_run_art_image.get(),
                    self.var_run_video.get()
                ]
            
            if s_steps[0] or s_steps[2]:
                any_gemini_requested = True
                break

        if any_gemini_requested:
            has_data = self._check_existing_data(target_ids)
            if has_data:
                # Ask user if they want to re-generate (Must be on main thread!)
                if messagebox.askyesno(self.t("confirm"), self.t("msg_confirm_regen")):
                    force_update = True

        self.stop_requested = False
        
        # --- SNAPSHOT CONFIGURATION FOR BACKGROUND ENGINE ---
        import copy
        config_snapshot = copy.deepcopy(self.config)
        config_snapshot["project_file"] = getattr(self, 'project_path', '')
        config_snapshot["global_steps"] = [
            self.var_run_lyrics.get(), 
            self.var_run_music.get(),
            self.var_run_art_prompt.get(),
            self.var_run_art_image.get(),
            self.var_run_video.get(),
            self.var_run_compilation.get()
        ]
        # ---------------------------------------------------
        
        # Check if the active preset has a custom project file (just in case it's not loaded)
        active_preset_data = self.config.get("artist_presets", {}).get(profile_name, {})
        if active_preset_data.get("project_file"):
            config_snapshot["project_file"] = active_preset_data["project_file"]

        # Register task (Rich structure for multiplexer)
        self.active_tasks[profile_name] = {
            "status": "Starting...",
            "start_time": time.time(),
            "total": len(target_ids),
            "current": 0
        }
        self._update_composite_status() # Update status bar immediately

        # We NO LONGER call self.disable_buttons() to lock the whole UI.
        # Minimal button sync:
        self.btn_stop.config(state="normal")
        
        # Pass the config snapshot to the background thread
        threading.Thread(target=self.run_process, args=(target_ids, force_update, config_snapshot, profile_name), daemon=True).start()

    def stop_process(self):
        self.stop_requested = True
        self.status_var.set(self.t("msg_stopping"))
        self.set_badge(self.t("badge_stopping"), "#ffaa00")
        # Iterate through all active browsers and stop them
        for profile_name, browser_instance in getattr(self, 'active_browsers', {}).items():
            try:
                browser_instance.stop()
                logger.warning(f"[{profile_name}] {self.t('log_browser_stop')}")
            except Exception:
                logger.error(f"Error stopping browser for profile {profile_name}", exc_info=True)
        # Clear the active_browsers dictionary
        setattr(self, 'active_browsers', {})
        
        # Safety: guarantee buttons re-enable even if worker thread hangs
        def _safety_enable():
            if str(self.btn_run.cget("state")) == "disabled":
                logger.warning("Safety timer: force-enabling buttons after stop.")
                self.enable_buttons()
        self.root.after(5000, _safety_enable)

    def _check_existing_data(self, target_ids):
        """Optimized: Checks if any of the target songs already have data using memory cache instead of re-loading Excel."""
        tab = self.current_tab
        if not tab or not hasattr(tab, "all_songs_data"): return False
        
        for rid in target_ids:
            rid_s = str(rid)
            s_data = tab.all_songs_data.get(rid_s)
            if not s_data: continue
            
            # Use specific steps for THIS song
            s_steps = tab.song_steps.get(rid_s)
            if s_steps is None:
                s_steps = [
                    self.var_run_lyrics.get(), 
                    self.var_run_music.get(),
                    self.var_run_art_prompt.get(),
                    self.var_run_art_image.get(),
                    self.var_run_video.get()
                ]

            # 1. Lyrics Check
            if s_steps[0] and s_data.get("has_lyrics"):
                return True
                
            # 2. Visual Prompt Check (Phase 3: Art Prompt)
            if s_steps[2]:
                if s_data.get("has_visual"):
                    return True
                if self.config.get("gemini_video") and s_data.get("has_video_prompt"):
                    return True
                    
        return False

    def run_process(self, target_ids, force_update=False, snapshot_config=None, profile_name="Default"):
        start_time = time.time()
        total_songs = len(target_ids)
        self.active_tasks[profile_name] = {
            "status": self.t("ready"),
            "start_time": start_time,
            "total": total_songs,
            "current": 0
        }
        # Local video queue for this run instance
        video_render_queue = []
        
        # Use snapshotted config if provided, fallback to live
        conf = snapshot_config if snapshot_config is not None else self.config

        try:
            # Save session state before starting long operation
            workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
            state_file = os.path.join(workspace, "session_state.json")
            try:
                with open(state_file, "w", encoding="utf-8") as f:
                    json.dump({
                        "project_file": getattr(self, 'project_path', ''),
                        "target_ids": target_ids
                    }, f)
            except Exception as e:
                logger.error(f"Failed to save session state: {e}")

            # Unified Project Path (Profile-based: output_media/[Profile_Name])
            project_file = conf.get("project_file", self.project_path)
            workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
            workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
            output_media = os.path.join(os.path.dirname(project_file), "output_media", profile_name)
            if not os.path.exists(output_media): os.makedirs(output_media, exist_ok=True)
            
            tab = self.tabs.get(profile_name)
            def get_song_steps(sid):
                if tab:
                    return tab.song_steps.get(sid) or conf.get("global_steps", [False, False, False, False, False, False])
                return conf.get("global_steps", [False, False, False, False, False, False])

            last_progress_time = {}
            def progress_callback(rid, text):
                if rid == "global":
                    if profile_name in self.active_tasks:
                        self.active_tasks[profile_name]["status"] = text
                    self._update_composite_status()
                else:
                    now = time.time()
                    # Force update for final states/emojis or rate-limit progress strings to max 3 FPS (0.3s)
                    force = any(e in text for e in ["✅", "❌", "🖼️", "🎵", "⚠️"])
                    if force or (now - last_progress_time.get(rid, 0) > 0.3):
                        last_progress_time[rid] = now
                        self.root.after(0, lambda t=text: self.update_progress(rid, t, profile_name=profile_name))
            
            # =================================================================================================
            # DECISION POINT: Batch Mode vs Sequential Mode
            # =================================================================================================
            # Sequential vs Batch
            if conf.get("suno_batch_enabled", False):
                self._run_process_batch_mode(target_ids, project_file, output_media, workspace, start_time, progress_callback, force_update, snapshot_config=conf, profile_name=profile_name, video_queue=video_render_queue)
            else:
                self._run_process_sequential_mode(target_ids, project_file, output_media, workspace, start_time, progress_callback, force_update, snapshot_config=conf, profile_name=profile_name, video_queue=video_render_queue)

            # Success cleanup (for non-video steps)
            if profile_name in self.active_tasks:
                self.active_tasks[profile_name]["status"] = self.t("ready")
            self._update_composite_status()

            logger.info(f"Main steps for '{profile_name}' finished in {time.time()-start_time:.1f}s")
            
            # If ONLY video was selected and queue is empty, inform user
            only_video = all(not get_song_steps(sid)[0] and not get_song_steps(sid)[1] and not get_song_steps(sid)[2] and not get_song_steps(sid)[3] and get_song_steps(sid)[4] for sid in target_ids)
            if only_video and not video_render_queue:
                 self.root.after(0, lambda: messagebox.showwarning(self.t("warning"), self.t("msg_no_materials") if hasattr(self, "t") else "No materials found for video rendering!"))
                 return

            # --- Phase 2: Parallel Video Rendering ---
            if video_render_queue and not self.stop_requested:
                parallel_count = int(self.config.get("video_parallel_count", 1))
                logger.info(f"Starting Parallel Video Rendering: {len(video_render_queue)} tasks, pool size: {parallel_count}")
                if profile_name in self.active_tasks:
                    self.active_tasks[profile_name]["status"] = f"Rendering {len(video_render_queue)} Videos ({parallel_count} parallel)..."
                self._update_composite_status()
                
                from video_generator import VideoGenerator
                
                def run_parallel_task(task):
                    if self.stop_requested: return
                    try:
                        vgen = VideoGenerator(output_dir=task["output_dir"])
                        success = vgen.generate_video(
                            audio_path=task["audio_path"],
                            image_path=task["image_path"],
                            output_filename=task["output_filename"],
                            **task["params"],
                            progress_callback=progress_callback
                        )
                        if success:
                            progress_callback(task["rid"], "Video Rendered ✅")
                            self.update_project_excel(task["rid"], video_status="TAMAM")
                            # Move used materials to 'done' (Ref: output_media/[profile]/done)
                            try:
                                done_dir = os.path.join(task["profile_dir"], "done")
                                if not os.path.exists(done_dir): os.makedirs(done_dir, exist_ok=True)
                                # Audio
                                if os.path.exists(task["audio_path"]):
                                    shutil.move(task["audio_path"], os.path.join(done_dir, os.path.basename(task["audio_path"])))
                                # Image
                                if os.path.exists(task["image_path"]):
                                    shutil.move(task["image_path"], os.path.join(done_dir, os.path.basename(task["image_path"])))
                            except Exception as me:
                                logger.warning(f"Failed to move files to done/: {me}")
                        else:
                            logger.error(f"Render failed for {task['rid']} (returned False)")
                            progress_callback(task["rid"], "Render Failed ❌")
                            self.update_project_excel(task["rid"], video_status="HATA")
                    except Exception as ve:
                        logger.error(f"Parallel Render Error ({task['rid']}): {ve}")
                        progress_callback(task["rid"], "Render Error ❌")
                        self.update_project_excel(task["rid"], video_status="HATA")

                # Run pool
                with ThreadPoolExecutor(max_workers=parallel_count) as pool:
                    pool.map(run_parallel_task, video_render_queue)
                
                logger.info("Parallel Video Rendering Phase Completed.")
                self.scan_materials(profile_name=profile_name) # Refresh status icons for THIS profile

            # --- Step 6: Video Merger (Compilation) ---
            global_steps = conf.get("global_steps", [False]*6)
            run_compilation = global_steps[5] if len(global_steps) > 5 else False
            
            if run_compilation and not self.stop_requested:
                try:
                    logger.info(self.t("log_merge_start"))
                    if profile_name in self.active_tasks:
                        self.active_tasks[profile_name]["status"] = self.t("compilation")
                    self._update_composite_status()
                    
                    video_dir = os.path.join(output_media, "videos")
                    if self.config.get("video_output_mode") == "custom":
                        video_dir = self.config.get("video_custom_output_path") or os.path.join(workspace, "Output_Videos")
                    
                    if os.path.exists(video_dir):
                        # Get all .mp4 files in the video dir, excluding existing compilations
                        all_vids = [os.path.join(video_dir, f) for f in os.listdir(video_dir) if f.endswith(".mp4") and not f.startswith("Compilation_")]
                        all_vids.sort() # Sort by name (ids usually)
                        
                        if all_vids:
                            from video_merger import VideoMerger
                            merger = VideoMerger(output_dir=video_dir)
                            compilation_name = f"Compilation_{profile_name}_{int(time.time())}.mp4"
                            success = merger.merge_videos(all_vids, compilation_name)
                            if success:
                                logger.info(self.t("log_merge_success").format(path=compilation_name))
                            else:
                                logger.error(self.t("log_merge_fail").format(error="Process Interrupted or Failed"))
                        else:
                            logger.warning(self.t("log_merge_no_files").format(path=video_dir))
                except Exception as e:
                    logger.error(self.t("log_merge_fail").format(error=str(e)))

            self.root.after(0, lambda: self.current_song_var.set(""))
            self.play_chime()

        except Exception as e:
            err_msg = str(e)
            logger.error(f"Critical Process Error: {err_msg}")
            self.root.after(0, lambda m=err_msg: messagebox.showerror(self.t("msg_critical_error"), m))
        finally:
            # Cleanup task registration for THIS profile
            if profile_name in self.active_tasks:
                del self.active_tasks[profile_name]
            
            # Close browser for THIS profile if exists and was managed by this task
            if profile_name in getattr(self, "active_browsers", {}):
                try:
                    self.active_browsers[profile_name].stop()
                    del self.active_browsers[profile_name]
                except: pass
            
            if not getattr(self, "stop_requested", False):
                # Only clean session state if NO other tasks are running? 
                # Or just for this specific project? For now, we clean if it matches.
                workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
                state_file = os.path.join(workspace, "session_state.json")
                try: 
                    if os.path.exists(state_file): os.remove(state_file)
                except: pass
            
            self._update_composite_status()
            
            if not self.active_tasks:
                self.enable_buttons(profile_name=profile_name)
                logger.info(f"🏁 ALL BACKGROUND TASKS COMPLETED.")
                
                # Show completion message safely after thread exits to avoid macOS Tkinter freeze
                def _show_done():
                    messagebox.showinfo(self.t("msg_done_title"), self.t("msg_done_info"))
                self.root.after(200, _show_done)
            else:
                self._refresh_all_tab(profile_name)
                logger.info(f"✔️ Task for '{profile_name}' finished. Remaining: {list(self.active_tasks.keys())}")

    def _start_status_ticker(self):
        """Start a periodic update for the status bar to show live elapsed time."""
        self._update_composite_status()
        self.root.after(1000, self._start_status_ticker)

    def _update_composite_status(self):
        """Multiplex active task statuses to the footer status bar with timing."""
        if not hasattr(self, "active_tasks") or not self.active_tasks:
            self.root.after(0, lambda: self.status_var.set(self.t("ready")))
            return
            
        now = time.time()
        combined = []
        for p, data in self.active_tasks.items():
            status = data.get("status", "...")
            start = data.get("start_time", now)
            total = data.get("total", 0)
            current = data.get("current", 0)
            
            # Elapsed
            elapsed_sec = int(now - start)
            e_h = elapsed_sec // 3600
            e_m = (elapsed_sec % 3600) // 60
            e_s = elapsed_sec % 60
            e_str = f"{e_h:02d}:{e_m:02d}:{e_s:02d}" if e_h > 0 else f"{e_m:02d}:{e_s:02d}"
            
            # ETA
            eta_str = "--:--"
            if current > 0 and total > 0:
                avg_per_item = elapsed_sec / current
                remaining_items = total - current
                eta_sec = int(avg_per_item * remaining_items)
                r_h = eta_sec // 3600
                r_m = (eta_sec % 3600) // 60
                r_s = eta_sec % 60
                eta_str = f"{r_h:02d}:{r_m:02d}:{r_s:02d}" if r_h > 0 else f"{r_m:02d}:{r_s:02d}"
            
            info = f"[{p}] {status} ({current}/{total}) | {self.t('elapsed_label').format(time=e_str)} | {self.t('eta_label').format(time=eta_str)}"
            combined.append(info)
        
        task_count_text = self.t("tasks_active").format(count=len(combined))
        
        if len(combined) > 1:
            # Cycle through tasks if multiple
            idx = int(now % (len(combined) * 3) // 3) # Switch every 3 seconds
            status_text = f"⚙️ {task_count_text} | " + combined[idx]
        else:
            status_text = combined[0]
            
        self.root.after(0, lambda: self.status_var.set(status_text))



    def get_prompts_path(self):
        """Returns the path to prompts.json in the workspace, initializing it if needed."""
        workspace = os.path.expanduser("~/Documents/MusicBot_Workspace")
        os.makedirs(workspace, exist_ok=True)
        prompts_path = os.path.join(workspace, "prompts.json")
        
        if not os.path.exists(prompts_path):
            # Try to migrate from old location
            old_path = os.path.join(os.getcwd(), "data", "prompts.json")
            if os.path.exists(old_path):
                import shutil
                shutil.copy(old_path, prompts_path)
            else:
                # Fallback default
                try:
                    import json
                    with open(prompts_path, "w", encoding="utf-8") as f:
                         json.dump({
                            "lyrics_master_prompt": "Sen profesyonel bir şarkı sözü yazarı ve müzik prodüktörüsün...",
                            "art_master_prompt": "Create a high-quality YouTube music thumbnail..."
                        }, f, indent=4)
                except Exception: pass
        return prompts_path

    def get_data_paths(self):
        docs_dir = os.path.expanduser("~/Documents/MusicBot_Workspace")
        os.makedirs(docs_dir, exist_ok=True)
        
        input_path = os.path.join(docs_dir, "input_songs.xlsx")
        output_path = os.path.join(docs_dir, "output_results.xlsx")
        output_dir = os.path.join(docs_dir, "output_media")
        prompts_path = os.path.join(docs_dir, "prompts.json")
        
        # Init prompts.json if missing
        if not os.path.exists(prompts_path):
            # Try to find default prompts in bundle or source
            bundle_prompts = os.path.join(bundle_dir, "data", "prompts.json")
            
            if os.path.exists(bundle_prompts):
                import shutil
                try:
                    shutil.copy(bundle_prompts, prompts_path)
                except Exception as e:
                    logger.error(f"Failed to copy prompts: {e}")
            else:
                # Fallback: Create with basic defaults if bundle also missing
                try:
                    import json
                    with open(prompts_path, "w", encoding="utf-8") as f:
                        json.dump({
                            "lyrics_master_prompt": "Sen profesyonel bir şarkı sözü yazarı ve müzik prodüktörüsün...",
                            "art_master_prompt": "Create a high-quality YouTube music thumbnail..."
                        }, f, indent=4)
                except Exception as e:
                    logger.error(f"Failed to create default prompts.json: {e}")

        # Init Input if missing
        if not os.path.exists(input_path):
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["id", "prompt", "style"])
                wb.save(input_path)
            except Exception as e:
                logger.error(f"Failed to create default input_songs.xlsx: {e}")
            
        return input_path, output_path, output_dir

    def open_xlsx(self, type="input"):
        input_path, output_path, _ = self.get_data_paths()
        path = input_path if type == "input" else output_path
        if os.path.exists(path):
            os.system(f"open '{path}'")
        else:
            messagebox.showinfo("Wait", "File not created yet.")

    def open_image_folder(self):
        _, output_path, _ = self.get_data_paths()
        img_dir = os.path.join(os.path.dirname(output_path), "images")
        if not os.path.exists(img_dir):
            os.makedirs(img_dir, exist_ok=True)
        os.system(f"open '{img_dir}'")

    def open_chrome_profile(self):
        """Launches the ACTUAL Google Chrome binary for manual login (Undetectable)."""
        
        # Determine current profile name/path
        profile_name = self.config.get("active_preset", "Default")
        if not profile_name: profile_name = "Default"
        profile_name = profile_name.strip().replace(" ", "_")
        
        # Base path for profiles
        base_path = os.path.expanduser("~/Documents/MusicBot_Workspace/chrome_profiles")
        profile_path = os.path.join(base_path, profile_name)
        os.makedirs(profile_path, exist_ok=True)
        
        def _launch():
            try:
                # 1. Notify User via Log
                logger.info(self.t("log_chrome_start"))
                logger.info(f"Using Profile: {profile_name} at {profile_path}")
                                     
                # 2. Start Native Chrome with SPECIFIC PROFILE
                temp_bc = BrowserController(headless=False, profile_path=profile_path)
                success = temp_bc.launch_native_chrome()
                
                if success:
                    logger.info(f"✅ Native Chrome launched for profile '{profile_name}'.")
                    logger.info("Please login manually and CLOSE the browser before starting the bot.")
                else:
                    logger.error("❌ Failed to launch native Chrome.")
                
            except Exception as e:
                logger.error(f"❌ Failed to open browser: {e}")

        # Threads
        threading.Thread(target=_launch, daemon=True).start()

    def reset_chrome_profile(self):
        """Resets ONLY the current profile's Chrome data."""
        
        profile_name = self.config.get("active_preset", "Default")
        if not profile_name: profile_name = "Default"
        profile_name = profile_name.strip().replace(" ", "_")
            
        base_path = os.path.expanduser("~/Documents/MusicBot_Workspace/chrome_profiles")
        profile_path = os.path.join(base_path, profile_name)

        if not messagebox.askyesno(self.t("confirm"), f"Are you sure you want to reset the Chrome profile for '{profile_name}'?\n\nThis will delete cookies and logins for THIS profile only."):
            return

        try:
            # First, try to kill any chrome instances for safety
            if profile_name in getattr(self, "active_browsers", {}):
                try: 
                    self.active_browsers[profile_name].stop()
                    del self.active_browsers[profile_name]
                except Exception: pass
            
            logger.info(f"Resetting Chrome profile at: {profile_path}")
            
            if os.path.exists(profile_path):
                import shutil
                shutil.rmtree(profile_path, ignore_errors=True)
                logger.info(f"✅ Profile '{profile_name}' reset successfully.")
                messagebox.showinfo(self.t("success"), f"Profile '{profile_name}' reset successfully.")
            else:
                logger.warning(f"Profile path not found: {profile_path}")
                messagebox.showinfo("Info", "Profile directory did not exist (already clean).")
                
        except Exception as e:
            logger.error(f"Failed to reset profile: {e}")
            messagebox.showerror("Error", f"Failed to reset profile: {e}")

    def disable_buttons(self):
        self.btn_run.config(state="disabled")
        self.btn_stop.config(state="normal")
        # Keep tabs and settings enabled for background work

    def enable_buttons(self, profile_name=None):
        self.stop_requested = False  # Reset flag so engine can restart
        self.root.after(0, lambda: self.btn_run.config(state="normal"))
        self.root.after(0, lambda: self.btn_stop.config(state="disabled"))
        self.root.after(0, lambda: self.status_var.set(self.t("ready")))
        self.root.after(0, lambda: self.set_badge(self.t("badge_idle"), "#888888"))
        
        if profile_name:
            self.root.after(0, lambda: self._refresh_all_tab(profile_name))
        else:
            self.root.after(0, self._refresh_all)

    def _refresh_all_tab(self, profile_name):
        """Helper to refresh a specific tab from background thread."""
        tab = self.tabs.get(profile_name)
        if tab:
            # We must be careful not to switch current_tab if we just want to update data
            # But load_project_data might need to know which tab to fill
            self.load_project_data(profile_name=profile_name)
            self.scan_materials(profile_name=profile_name)
            self.apply_filter(profile_name=profile_name)

    def _run_process_sequential_mode(self, target_ids, project_file, output_media, workspace, start_time, progress_callback, force_update, snapshot_config=None, profile_name="Default", video_queue=None):
        """Process each song 1-by-1 through all selected steps with a persistent browser."""
        total_songs = len(target_ids)
        # Using local video_queue passed from run_process
        conf = snapshot_config if snapshot_config is not None else self.config
        global_human = conf.get("humanizer_enabled", True)

        common_browser = None
        tab = self.tabs.get(profile_name)
        if not tab:
            logger.error(f"Tab for profile {profile_name} not found!")
            return

        try:
            for idx, song_id in enumerate(target_ids):
                if self.stop_requested: break
                
                if profile_name in self.active_tasks:
                    self.active_tasks[profile_name]["current"] = idx + 1
                
                song_data = tab.all_songs_data.get(song_id, {})
                title = str(song_data.get("title") or "Unknown")
                title_short = title[:50] + ".." if len(title) > 50 else title
                self.root.after(0, lambda id=song_id, t=title_short: self.current_song_var.set(f"{self.t('working_on')} {id} ({t})"))
                self.root.after(0, lambda: self.set_badge(self.t("badge_active"), "#00aa00"))

                # Determine steps for THIS song
                s_steps = tab.song_steps.get(song_id)
                if s_steps is None:
                    s_steps = conf.get("global_steps", [False]*6)[:5]
                while len(s_steps) < 5: s_steps.append(False)

                # Start browser IF NEEDED
                browser_needed = s_steps[0] or s_steps[1] or s_steps[2]
                if browser_needed and not common_browser and not self.stop_requested:
                    try:
                        base_path = os.path.expanduser("~/Documents/MusicBot_Workspace/chrome_profiles")
                        profile_path = os.path.join(base_path, profile_name.strip().replace(" ", "_"))
                        from browser_controller import BrowserController
                        h_conf = {"level": conf.get("humanizer_level", "MEDIUM"), "speed": conf.get("humanizer_speed", 1.0), "retries": conf.get("humanizer_retries", 1), "adaptive": conf.get("humanizer_adaptive", True)}
                        common_browser = BrowserController(headless=False, profile_path=profile_path, humanizer_config=h_conf)
                        common_browser.start()
                        self.active_browsers[profile_name] = common_browser
                    except Exception as b_err:
                        logger.error(f"Sequential browser start failed: {b_err}")

                try:
                    # --- Step 1: Lyrics & Gemini ---
                    if s_steps[0] and not self.stop_requested and common_browser:
                        # --- PROMPT ISOLATION: Get master prompts from this profile's preset snapshot ---
                        artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                        preset_prompts = artist_preset.get("prompts", {})
                        m_prompts = {
                            "lyrics_master_prompt": preset_prompts.get("lyrics_master_prompt"),
                            "visual_master_prompt": preset_prompts.get("visual_master_prompt"),
                            "video_master_prompt": preset_prompts.get("video_master_prompt"),
                            "art_master_prompt": preset_prompts.get("art_master_prompt")
                        }
                        
                        common_browser.humanizer_enabled = global_human and conf.get("h_activate_gemini", True)
                        from gemini_prompter import GeminiPrompter
                        gemini = GeminiPrompter(
                            project_file=project_file, 
                            browser=common_browser, 
                            use_gemini_lyrics=s_steps[0], 
                            generate_visual=s_steps[2], 
                            generate_video=s_steps[4] if len(s_steps) > 4 else False, 
                            language=conf.get("target_language", "Turkish"), 
                            chat_mode=conf.get("gemini_chat_mode", self.t("gemini_mode_new")), 
                            xlsx_lock=self.xlsx_lock,
                            master_prompts=m_prompts
                        )
                        gemini.run(target_ids=[song_id], progress_callback=progress_callback, force_update=force_update)
                    
                    # --- Step 2: Music ---
                    if s_steps[1] and not self.stop_requested and common_browser:
                        common_browser.humanizer_enabled = global_human and self.config.get("h_activate_suno", True)
                        
                        # --- SUNO ISOLATION: Get persona and gender from this profile's preset snapshot ---
                        artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                        preset_settings = artist_preset.get("settings", {})
                        p_alias = preset_settings.get("suno_active_persona", "")
                        p_link = conf.get("suno_personas", {}).get(p_alias, "") if preset_settings.get("suno_persona_link_enabled") else ""
                        if not p_link:
                            # Fallback if using old structure
                            p_link = preset_settings.get("persona_link", "")

                        v_gender = preset_settings.get("vocal_gender", "Default")
                        a_infl = preset_settings.get("audio_influence", 25)
                        wrd = preset_settings.get("weirdness", 50)
                        s_infl = preset_settings.get("style_influence", 50)
                        l_mode = preset_settings.get("lyrics_mode", "Default")
                        
                        from suno_generator import SunoGenerator
                        suno = SunoGenerator(
                            project_file=project_file, 
                            output_dir=output_media, 
                            delay=conf.get("suno_delay", 15), 
                            browser=common_browser, 
                            xlsx_lock=self.xlsx_lock,
                            persona_link=p_link,
                            vocal_gender=v_gender,
                            audio_influence=a_infl,
                            weirdness=wrd,
                            style_influence=s_infl,
                            lyrics_mode=l_mode
                        )
                        suno.run(target_ids=[song_id], progress_callback=progress_callback, force_update=force_update)
                    
                    # --- Step 3: Art Proxies ---
                    if s_steps[2] and not s_steps[0] and common_browser:
                        # --- PROMPT ISOLATION: Get master prompts from this profile's preset snapshot ---
                        artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                        preset_prompts = artist_preset.get("prompts", {})
                        m_prompts = {
                            "lyrics_master_prompt": preset_prompts.get("lyrics_master_prompt"),
                            "visual_master_prompt": preset_prompts.get("visual_master_prompt"),
                            "video_master_prompt": preset_prompts.get("video_master_prompt"),
                            "art_master_prompt": preset_prompts.get("art_master_prompt")
                        }
                        
                        from gemini_prompter import GeminiPrompter
                        ga = GeminiPrompter(
                            project_file=project_file, 
                            browser=common_browser, 
                            xlsx_lock=self.xlsx_lock,
                            master_prompts=m_prompts
                        )
                        ga.generate_art_prompts(target_ids=[song_id], progress_callback=progress_callback)
                    
                    # --- Step 4: Art Images ---
                    if s_steps[3] and not self.stop_requested:
                        # --- PROMPT ISOLATION: Get master prompts from this profile's preset snapshot ---
                        artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                        preset_prompts = artist_preset.get("prompts", {})
                        m_prompts = {
                            "lyrics_master_prompt": preset_prompts.get("lyrics_master_prompt"),
                            "visual_master_prompt": preset_prompts.get("visual_master_prompt"),
                            "video_master_prompt": preset_prompts.get("video_master_prompt"),
                            "art_master_prompt": preset_prompts.get("art_master_prompt")
                        }
                        
                        from gemini_prompter import GeminiPrompter
                        gi = GeminiPrompter(
                            project_file=project_file, 
                            xlsx_lock=self.xlsx_lock,
                            master_prompts=m_prompts
                        )
                        gi.generate_art_images(target_ids=[song_id], progress_callback=progress_callback)

                    # --- Step 5: Video ---
                    if len(s_steps) > 4 and s_steps[4] and not self.stop_requested:
                        # Optimization: Pre-scan directory once per chunk/song loop
                        all_output_media_files = []
                        if os.path.exists(output_media):
                            all_output_media_files = os.listdir(output_media)
                        self._prepare_video_task(song_id, output_media, workspace, progress_callback, video_queue=video_queue, all_files=all_output_media_files)
                except Exception as e:
                    logger.error(f"Error in song {song_id}: {e}")
                    progress_callback(song_id, "Hata! ❌")
        finally:
            if common_browser:
                try: 
                    common_browser.stop()
                    if profile_name in self.active_browsers: del self.active_browsers[profile_name]
                except: pass
            self.active_browser = None

    def _prepare_video_task(self, song_id, output_media, workspace, progress_callback=None, video_queue=None, all_files=None):
        """Optimized helper to prepare a video task for a single song."""
        if not os.path.exists(output_media): 
            if progress_callback: progress_callback(song_id, "Medya Klasörü Yok ❌")
            return
        
        if all_files is None:
            all_files = os.listdir(output_media)
            
        try:
            all_materials = []
            found_audio = []
            
            # Selection logic: _1, _2 or Both
            selection_mode = self.config.get("video_selection_mode", "Both")
            suffix_map = {
                self.t("v_mode_1"): "_1",
                self.t("v_mode_2"): "_2",
                "Only _1": "_1",
                "Only _2": "_2"
            }
            target_suffix = suffix_map.get(selection_mode)
            variant_filter = target_suffix.replace("_", "") if target_suffix else None
            
            common_id_suffix = None
            if variant_filter:
                # If we have a variant filter (1 or 2), try matching _1 or _2 explicitly
                common_id_suffix = f"_{variant_filter}"

            for f in all_files:
                f_l = f.lower()
                if self._is_id_match(f_l, song_id):
                    if f_l.endswith((".mp3", ".wav", ".png", ".jpg", ".jpeg")):
                        all_materials.append(f)
                    
                    if f_l.endswith((".mp3", ".wav")):
                        if not variant_filter or self._is_id_match(f_l, song_id, variant=variant_filter):
                            found_audio.append(f)

            found_audio = sorted(list(set(found_audio)))
            
            if found_audio:
                v_params = {
                    "effect_types": self.config.get("video_effects", [self.config.get("video_effect", "None")]),
                    "fps": int(self.config.get("video_fps", 30)),
                    "resolution": self.config.get("video_resolution", self.t("video_res_shorts")),
                    "intensity": int(self.config.get("video_intensity", 1.0) * 50)
                }
                
                for aud_file in found_audio:
                    if self.stop_requested: break
                    aud_full_path = os.path.join(output_media, aud_file)
                    aud_base = os.path.splitext(aud_file)[0]
                    
                    if not hasattr(self, "video_render_queue"): self.video_render_queue = []

                    # Find matching image logic 
                    img_path = None
                    assets_folder = self.config.get("video_assets_path")
                    
                    # Extract variant from audio file name
                    aud_variant = None
                    if "_1" in aud_file.lower() or "variant1" in aud_file.lower(): aud_variant = "1"
                    elif "_2" in aud_file.lower() or "variant2" in aud_file.lower(): aud_variant = "2"

                    # 1. Search in output_media
                    for f in all_files:
                        f_l = f.lower()
                        if f_l.endswith((".png", ".jpg", ".jpeg")) and self._is_id_match(f_l, song_id, variant=aud_variant):
                            img_path = os.path.join(output_media, f); break
                    
                    # 2. Fallback
                    if not img_path:
                        for f in all_files:
                            f_l = f.lower()
                            if f_l.endswith((".png", ".jpg", ".jpeg")) and self._is_id_match(f_l, song_id):
                                img_path = os.path.join(output_media, f); break
                                
                    # 3. Assets Fallback
                    if not img_path and assets_folder and os.path.exists(assets_folder):
                        try:
                            asset_files = os.listdir(assets_folder)
                            for f in asset_files:
                                f_l = f.lower()
                                if f_l.endswith((".png", ".jpg", ".jpeg")) and self._is_id_match(f_l, song_id):
                                    img_path = os.path.join(assets_folder, f); break
                        except Exception: pass
                    
                    if img_path:
                        video_out_dir = os.path.join(output_media, "videos")
                        if self.config.get("video_output_mode") == "custom":
                            video_out_dir = self.config.get("video_custom_output_path") or os.path.join(workspace, "Output_Videos")
                        if not os.path.exists(video_out_dir): os.makedirs(video_out_dir, exist_ok=True)
                        
                    if video_queue is not None:
                        video_queue.append({
                            "rid": song_id,
                            "audio_path": aud_full_path,
                            "image_path": img_path,
                            "output_dir": video_out_dir,
                            "profile_dir": output_media,
                            "output_filename": f"{aud_base}.mp4",
                            "params": v_params
                        })
                    else:
                        if progress_callback: progress_callback(song_id, "Malzeme Eksik (Resim Yok) 🖼️")
                        logger.warning(f"No matching image found for {aud_file}")
            else:
                 if progress_callback: progress_callback(song_id, "Malzeme Eksik (Ses Yok) 🎵")
                 logger.warning(f"No matching audio found for song {song_id}")
                 
        except Exception as e:
            logger.error(f"Error preparing video task for {song_id}: {e}")
            if progress_callback: progress_callback(song_id, "Hazırlık Hatası ❌")

    def _run_process_batch_mode(self, target_ids, project_file, output_media, workspace, start_time, progress_callback, force_update, snapshot_config=None, profile_name="Default", video_queue=None):
        """Batch executed flow: Lyrics(All) -> Suno_Batch(All) -> Art(All) -> Video(All)"""
        
        conf = snapshot_config if snapshot_config is not None else self.config
        tab = self.tabs.get(profile_name)
        if not tab:
            logger.error(f"Tab for profile {profile_name} not found during batch mode!")
            return
            
        def get_song_steps(sid):
            return tab.song_steps.get(sid) or conf.get("global_steps", [False]*6)[:5]

        try:
            # 1. Establish SINGLE Browser Session for Phases 1 & 2 & 3
            # We use one browser session for everything to avoid startups/shutdowns
            # ---------------- PROFILE LOGIC ----------------
            profile_name = profile_name.strip().replace(" ", "_")
            
            base_path = os.path.expanduser("~/Documents/MusicBot_Workspace/chrome_profiles")
            profile_path = os.path.join(base_path, profile_name)
            # -----------------------------------------------

            from browser_controller import BrowserController
            h_conf = {
                "level": conf.get("humanizer_level", "MEDIUM"),
                "speed": conf.get("humanizer_speed", 1.0),
                "retries": conf.get("humanizer_retries", 1),
                "adaptive": conf.get("humanizer_adaptive", True)
            }
            batch_browser = BrowserController(headless=False, profile_path=profile_path, humanizer_config=h_conf)
            batch_browser.start()
            self.active_browsers[profile_name] = batch_browser
            global_human = conf.get("humanizer_enabled", True)
            
            # --- PHASE 1: LYRICS & PREP (Gemini) ---
            gemini_ids = [id for id in target_ids if get_song_steps(id)[0]] # Step 1 enabled
            if gemini_ids:
                logger.info(f"Batch Phase 1: Gemini Lyrics ({len(gemini_ids)} songs)")
                batch_browser.humanizer_enabled = global_human and self.config.get("h_activate_gemini", True)
                
                # --- PROMPT ISOLATION: Get master prompts from this profile's preset snapshot ---
                artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                preset_prompts = artist_preset.get("prompts", {})
                m_prompts = {
                    "lyrics_master_prompt": preset_prompts.get("lyrics_master_prompt"),
                    "visual_master_prompt": preset_prompts.get("visual_master_prompt"),
                    "video_master_prompt": preset_prompts.get("video_master_prompt"),
                    "art_master_prompt": preset_prompts.get("art_master_prompt")
                }
                
                from gemini_prompter import GeminiPrompter
                gemini = GeminiPrompter(
                    project_file=project_file, 
                    browser=batch_browser,
                    use_gemini_lyrics=conf.get("gemini_lyrics", True),
                    generate_visual=False, 
                    generate_video=False,
                    generate_style=conf.get("gemini_style", False),
                    startup_delay=conf.get("startup_delay", 5),
                    language=conf.get("target_language", "Turkish"),
                    chat_mode=conf.get("gemini_chat_mode", self.t("gemini_mode_new")),
                    xlsx_lock=self.xlsx_lock,
                    master_prompts=m_prompts
                )
                # Gemini doesn't have internal batch support yet, so we loop but REUSE browser
                for idx, song_id in enumerate(gemini_ids):
                    if self.stop_requested: break
                    if profile_name in self.active_tasks:
                        self.active_tasks[profile_name]["current"] = idx + 1
                    progress_callback(song_id, "Gemini İşleniyor... ✍️")
                    gemini.run(target_ids=[song_id], progress_callback=progress_callback, force_update=force_update)
            
            # --- PHASE 2: SUNO BATCH (Gen & Download) ---
            suno_ids = [id for id in target_ids if get_song_steps(id)[1]] # Step 2 enabled
            if suno_ids and not self.stop_requested:
                logger.info(f"Batch Phase 2: Suno Batch ({len(suno_ids)} songs)")
                batch_browser.humanizer_enabled = global_human and self.config.get("h_activate_suno", True)
                
                # --- SUNO ISOLATION: Get persona and gender from this profile's preset snapshot ---
                artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                preset_settings = artist_preset.get("settings", {})
                p_alias = preset_settings.get("suno_active_persona", "")
                p_link = conf.get("suno_personas", {}).get(p_alias, "") if preset_settings.get("suno_persona_link_enabled") else ""
                if not p_link:
                    # Fallback if using old structure
                    p_link = preset_settings.get("persona_link", "")

                v_gender = preset_settings.get("vocal_gender", "Default")
                a_infl = preset_settings.get("audio_influence", 25)
                wrd = preset_settings.get("weirdness", 50)
                s_infl = preset_settings.get("style_influence", 50)
                l_mode = preset_settings.get("lyrics_mode", "Default")
                
                from suno_generator import SunoGenerator
                suno = SunoGenerator(
                    project_file=project_file, 
                    output_dir=output_media,
                    delay=conf.get("suno_delay", 15),
                    startup_delay=conf.get("startup_delay", 5),
                    browser=batch_browser,
                    audio_influence=a_infl if preset_settings.get("audio_influence_enabled") else "Default",
                    vocal_gender=v_gender if preset_settings.get("vocal_gender_enabled") else "Default",
                    weirdness=wrd if preset_settings.get("weirdness_enabled") else "Default",
                    style_influence=s_infl if preset_settings.get("style_influence_enabled") else "Default",
                    lyrics_mode=l_mode if preset_settings.get("lyrics_mode_enabled") else "Default",
                    persona_link=p_link,
                    turbo=self.var_turbo.get(),
                    xlsx_lock=self.xlsx_lock
                )
                # Use the new Batch Method with OP MODE!
                op_mode = self.config.get("suno_batch_op_mode", "full")
                logger.info(f"Batch Suno Op Mode: {op_mode}")
                suno.run_batch(target_ids=suno_ids, progress_callback=progress_callback, stats_callback=self.update_dashboard_stats, force_update=force_update, op_mode=op_mode)

            # --- PHASE 3: ART (Prompts & Images) ---
            art_p_ids = [id for id in target_ids if get_song_steps(id)[2]] # Step 3 enabled
            art_i_ids = [id for id in target_ids if get_song_steps(id)[3]] # Step 4 enabled
            
            if (art_p_ids or art_i_ids) and not self.stop_requested:
                logger.info("Batch Phase 3: Art Generation")
                # --- PROMPT ISOLATION: Get master prompts from this profile's preset snapshot ---
                artist_preset = snapshot_config.get("artist_presets", {}).get(profile_name, {}) if snapshot_config else {}
                preset_prompts = artist_preset.get("prompts", {})
                m_prompts = {
                    "lyrics_master_prompt": preset_prompts.get("lyrics_master_prompt"),
                    "visual_master_prompt": preset_prompts.get("visual_master_prompt"),
                    "video_master_prompt": preset_prompts.get("video_master_prompt"),
                    "art_master_prompt": preset_prompts.get("art_master_prompt")
                }
                
                from gemini_prompter import GeminiPrompter
                gemini_art = GeminiPrompter(
                    project_file=project_file,
                    output_dir=output_media,
                    browser=batch_browser, 
                    startup_delay=conf.get("startup_delay", 5),
                    language=conf.get("target_language", "Turkish"),
                    chat_mode=conf.get("gemini_chat_mode", self.t("gemini_mode_new")),
                    xlsx_lock=self.xlsx_lock,
                    master_prompts=m_prompts
                )
                
                # Prompts via Browser
                if art_p_ids:
                    for song_id in art_p_ids:
                        if self.stop_requested: break
                        gemini_art.generate_art_prompts(target_ids=[song_id], progress_callback=progress_callback)
                
                # Images (Local, but using same logical flow)
                if art_i_ids:
                    for song_id in art_i_ids:
                        if self.stop_requested: break
                        gemini_art.generate_art_images(target_ids=[song_id], progress_callback=progress_callback)

            # Clean up browser
            batch_browser.stop()
            if profile_name in self.active_browsers:
                del self.active_browsers[profile_name]

            # --- PHASE 4: VIDEO PREP & QUEUE ---
            video_ids = [id for id in target_ids if get_song_steps(id)[4]] # Step 5 enabled
            if video_ids and not self.stop_requested:
                self.video_render_queue = []
                all_output_media_files = []
                if os.path.exists(output_media):
                    all_output_media_files = os.listdir(output_media)
                for song_id in video_ids:
                    self._prepare_video_task(song_id, output_media, workspace, progress_callback, all_files=all_output_media_files)

        except Exception as e:
            logger.error(f"Batch Mode Error: {e}")
            self.root.after(0, lambda m=str(e): messagebox.showerror(self.t("error"), m))
        finally:
            if profile_name in getattr(self, "active_browsers", {}):
                try: 
                    self.active_browsers[profile_name].stop()
                    del self.active_browsers[profile_name]
                except Exception: pass

    def load_data(self):
        # Dummy load_data for the finally block
        pass

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = MusicBotGUI(root)
        root.mainloop()
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        # Log to console
        print(f"CRITICAL STARTUP ERROR:\n{error_msg}")
        # Log to crash report file
        try:
            ws_path = os.path.expanduser("~/Documents/MusicBot_Workspace")
            os.makedirs(ws_path, exist_ok=True)
            with open(os.path.join(ws_path, "crash_report.txt"), "w", encoding="utf-8") as f:
                f.write(f"Timestamp: {time.ctime()}\n")
                f.write(error_msg)
        except Exception as file_e:
            print(f"Failed to write crash report: {file_e}")
        sys.exit(1)
