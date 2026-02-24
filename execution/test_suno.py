import unittest
import os
import shutil
import openpyxl

from suno_config import SunoConfig
from suno_excel import SunoExcelMixin
from suno_downloader import SunoDownloaderMixin

class DummyConfig:
    def __init__(self):
        self.backup_interval = -1
        self.max_backups = 3
        self.min_file_size = 100  # Small size for testing
        self.retry_count = 1

class DummyGenerator(SunoExcelMixin, SunoDownloaderMixin):
    def __init__(self, metadata_path):
        self.metadata_path = metadata_path
        self.config = DummyConfig()

class TestSunoComponents(unittest.TestCase):
    def setUp(self):
        self.test_dir = "test_env"
        os.makedirs(self.test_dir, exist_ok=True)
        self.excel_path = os.path.join(self.test_dir, "test.xlsx")
        
        # Create a valid minimal Excel file for testing
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id", "prompt", "status", "dl_status", "dl_attempts"])
        ws.append(["123", "test prompt", "", "", ""])
        wb.save(self.excel_path)
        
        self.gen = DummyGenerator(self.excel_path)

    def tearDown(self):
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)

    def test_config_defaults(self):
        config = SunoConfig()
        self.assertEqual(config.retry_count, 3)
        self.assertEqual(config.max_backups, 3)
        self.assertIn("wav", config.format_preference)

    def test_excel_update_status(self):
        self.gen.update_row_status(row_idx=2, status="Completed", dl_status="Success", dl_attempts=1)
        
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=3).value, "Completed")
        self.assertEqual(ws.cell(row=2, column=4).value, "Success")
        self.assertEqual(ws.cell(row=2, column=5).value, 1)
        
    def test_excel_backup_creation(self):
        self.gen.update_row_status(row_idx=2, status="Completed")
        backup_dir = os.path.join(self.test_dir, ".backups")
        
        self.assertTrue(os.path.exists(backup_dir))
        self.assertTrue(os.path.exists(os.path.join(backup_dir, "test.xlsx.bak1")))

    def test_excel_recovery_from_corruption(self):
        # 1. Create a valid backup by doing one save
        self.gen.update_row_status(row_idx=2, status="Initial Valid")
        
        # 2. Corrupt the active excel file explicitly
        with open(self.excel_path, "wb") as f:
            f.write(b"CORRUPT DATA NOT A ZIP FILE")
            
        # 3. Attempt to save again. The try/except in update_row_status 
        # should catch the BadZipFile and trigger _recover_excel_from_backup!
        self.gen.update_row_status(row_idx=2, status="Post-Corrupt Save")
        
        # 4. Check if we can load it normally now (it should have been recovered)
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            recovered = True
        except:
            recovered = False
            
        self.assertTrue(recovered, "File was not recovered from backup structure after corruption!")

    def test_download_verification_wav(self):
        wav_path = os.path.join(self.test_dir, "test.wav")
        with open(wav_path, "wb") as f:
            f.write(b"RIFF\x00\x00\x00\x00WAVEfmt ")
            f.write(b"\x00" * 200) # padding to pass min_file_size
            
        self.assertTrue(self.gen._verify_audio_file(wav_path, "wav"))
        self.assertFalse(self.gen._verify_audio_file(wav_path, "mp3"))
        
    def test_download_verification_mp3(self):
        mp3_path = os.path.join(self.test_dir, "test.mp3")
        with open(mp3_path, "wb") as f:
            f.write(b"ID3\x03\x00\x00\x00\x00\x00\x00")
            f.write(b"\x00" * 200)
            
        self.assertTrue(self.gen._verify_audio_file(mp3_path, "mp3"))
        self.assertFalse(self.gen._verify_audio_file(mp3_path, "wav"))
        
    def test_download_verification_too_small(self):
        tiny_path = os.path.join(self.test_dir, "tiny.wav")
        with open(tiny_path, "wb") as f:
            f.write(b"RIFF....WAVE") # Size is < 100 bytes limit
            
        self.assertFalse(self.gen._verify_audio_file(tiny_path, "wav"))

if __name__ == '__main__':
    unittest.main()
